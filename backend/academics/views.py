from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes, parser_classes, action
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django.db.models import Q
from django.db import IntegrityError, transaction
from django.db.models import Sum, Avg
from django.http import HttpResponse
from django.conf import settings
from django.core.cache import cache
from django.template.loader import render_to_string
import threading
from io import BytesIO, StringIO
from datetime import date
from django.utils import timezone
import uuid
import requests
import secrets
from django.core.files.base import ContentFile
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False
import csv, io
from django_filters.rest_framework import DjangoFilterBackend
from communications.utils import send_sms, send_email_safe, send_email_safe_html, send_email_with_attachment, log_delivery, create_messages_for_users, resolve_default_sender_id
from communications.models import DeliveryLog, Message, MessageRecipient
from finance.models import Invoice, Payment
from .models import (
    Class, Student, Competency, Assessment, Attendance, TeacherProfile, Subject, SubjectComponent,
    Exam, ExamResult, AcademicYear, Term, Stream, LessonPlan, ClassSubjectTeacher, SubjectGradingBand, StageGradingBand,
    Room, TimetableEntry, TimetableTemplate, PeriodSlotTemplate, TimetablePlan, TimetableClassConfig,
    ClassSubjectQuota, TeacherAvailability, TimetableVersion, TeacherDuty, StudentClassHistory
)
from .serializers import (
    ClassSerializer, StudentSerializer, StudentListSerializer, CompetencySerializer, AssessmentSerializer, AttendanceSerializer,
    TeacherProfileSerializer, SubjectSerializer, SubjectComponentSerializer, ExamSerializer,
    ExamResultSerializer, AcademicYearSerializer, TermSerializer, StreamSerializer, LessonPlanSerializer,
    ClassSubjectTeacherSerializer, SubjectGradingBandSerializer, StageGradingBandSerializer, RoomSerializer, TimetableEntrySerializer,
    TimetableTemplateSerializer, PeriodSlotTemplateSerializer, TimetablePlanSerializer, TimetableClassConfigSerializer,
    ClassSubjectQuotaSerializer, TeacherAvailabilitySerializer, TimetableVersionSerializer, TeacherDutySerializer
)

class IsTeacherOrAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated):
            return False
        role = getattr(user, 'role', None)
        # Treat Django staff/superuser as admin-equivalent for read access
        if user.is_staff or user.is_superuser:
            return True
        return role in ('teacher', 'admin', 'finance')

class IsAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and (
            request.user.role == 'admin' or request.user.is_staff or request.user.is_superuser
        )

class IsAdminOrTeacherReadOnly(permissions.BasePermission):
    """Allow teachers to perform safe (read-only) requests; only admins can modify.
    Additionally, a teacher with TeacherProfile.can_manage_timetable=True may modify
    timetable-related resources. Admin is role 'admin' or staff/superuser.
    """
    def has_permission(self, request, view):
        user = getattr(request, 'user', None)
        if not (user and user.is_authenticated):
            return False
        is_admin = bool(getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser)
        if request.method in permissions.SAFE_METHODS:
            return is_admin or getattr(user, 'role', None) == 'teacher'
        # For write methods, allow timetable-managing teachers
        if not is_admin and getattr(user, 'role', None) == 'teacher':
            try:
                prof = TeacherProfile.objects.filter(user=user).first()
                return bool(getattr(prof, 'can_manage_timetable', False))
            except Exception:
                return False
        return is_admin


class TeacherDutyViewSet(viewsets.ModelViewSet):
    queryset = TeacherDuty.objects.all()
    serializer_class = TeacherDutySerializer
    permission_classes = [IsTeacherOrAdmin]

    def get_queryset(self):
        qs = super().get_queryset().select_related('teacher', 'created_by')
        user = getattr(self.request, 'user', None)
        school = getattr(user, 'school', None)
        if school:
            qs = qs.filter(school=school)
        # Teachers only see their own by default
        if getattr(user, 'role', None) == 'teacher' and self.request.method in permissions.SAFE_METHODS:
            qs = qs.filter(teacher=user)
        # Filters
        if self.request.query_params.get('mine'):
            qs = qs.filter(teacher=user)
        status_param = self.request.query_params.get('status')
        if status_param:
            qs = qs.filter(status=status_param)
        teacher_param = self.request.query_params.get('teacher')
        if teacher_param:
            try:
                qs = qs.filter(teacher_id=int(teacher_param))
            except Exception:
                qs = qs.filter(teacher_id=teacher_param)
        return qs

    def get_serializer_class(self):
        # Use lightweight serializer for list responses
        if getattr(self, 'action', None) == 'list':
            return StudentListSerializer
        return super().get_serializer_class()

    def perform_create(self, serializer):
        user = getattr(self.request, 'user', None)
        if not (user and (user.role == 'admin' or user.is_staff or user.is_superuser)):
            raise ValidationError({'detail': 'Only admin can create duties'})
        school = getattr(user, 'school', None)
        serializer.save(created_by=user, school=school)

    def perform_update(self, serializer):
        user = getattr(self.request, 'user', None)
        is_admin = bool(user and (user.role == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            # Allow assignee to update status only
            instance = self.get_object()
            data = serializer.validated_data
            if instance.teacher_id != getattr(user, 'id', None):
                raise ValidationError({'detail': 'Not allowed'})
            # Restrict updates
            allowed = {k: v for k, v in data.items() if k in ('status',)}
            for k, v in allowed.items():
                setattr(instance, k, v)
            instance.save(update_fields=list(allowed.keys()) or ['status'])
            return instance
        serializer.save()

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='mark-done')
    def mark_done(self, request, pk=None):
        duty = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (user.role == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin and duty.teacher_id != getattr(user, 'id', None):
            return Response({'detail': 'Not allowed'}, status=status.HTTP_403_FORBIDDEN)
        duty.status = 'done'
        duty.save(update_fields=['status'])
        return Response({'detail': 'Marked done'})

class StreamViewSet(viewsets.ModelViewSet):
    queryset = Stream.objects.all()
    serializer_class = StreamSerializer
    permission_classes = [IsAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        return qs

    def perform_create(self, serializer):
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if not school:
            raise ValidationError({'school': 'School is required. Set your user.school in Django admin.'})
        serializer.save(school=school)

    @action(detail=True, methods=['post'], url_path='resync-classes')
    def resync_classes(self, request, pk=None):
        """Force-refresh names of classes under this stream to match the current stream name.
        Useful after renaming a stream when some classes may be stale.
        """
        stream = self.get_object()
        from .models import Class as ClassModel
        updated = 0
        for c in ClassModel.objects.filter(stream=stream).only('id','grade_level','stream'):
            # Trigger Class.save() name regeneration
            c.save(update_fields=['name'])
            updated += 1
        return Response({'detail': 'ok', 'updated': updated})


class ClassSubjectTeacherViewSet(viewsets.ModelViewSet):
    queryset = ClassSubjectTeacher.objects.all()
    serializer_class = ClassSubjectTeacherSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]

    def get_queryset(self):
        qs = super().get_queryset().select_related('klass','subject','teacher')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(klass__school=school)
        # Optional filters
        klass = self.request.query_params.get('klass')
        if klass:
            qs = qs.filter(klass_id=klass)
        subject = self.request.query_params.get('subject')
        if subject:
            qs = qs.filter(subject_id=subject)
        return qs

    def perform_create(self, serializer):
        # Validate the chosen class is in admin's school
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        klass = serializer.validated_data.get('klass')
        if school and klass and klass.school_id != school.id:
            raise ValidationError({'klass': 'Class must belong to your school'})
        serializer.save()

class ClassViewSet(viewsets.ModelViewSet):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['school']
    def perform_create(self, serializer):
        # Resolve school: prefer payload, else user's school
        school = serializer.validated_data.get('school') or getattr(self.request.user, 'school', None)
        if not school:
            raise ValidationError({'school': 'School is required. Set your user.school in Django admin or include "school" in the request.'})
        serializer.save(school=school)
    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        # NOTE: Previously we excluded empty Grade 9 classes here. This caused Grade 9 classes
        # to be hidden from the Manage Classes page when they had no students. We now return
        # all classes without special-casing Grade 9.
        # Optimize list: only basic fields needed by filters
        act = getattr(self, 'action', None)
        if act == 'list':
            qs = qs.only('id','name','grade_level')
        else:
            # For profile/detail views, prefetch related collections used in ClassSerializer
            qs = qs.prefetch_related('subjects', 'subject_teachers__teacher')
        return qs

    def get_serializer_class(self):
        from .serializers import ClassLiteSerializer, ClassSerializer as ClassDetailSerializer
        if getattr(self, 'action', None) == 'list':
            return ClassLiteSerializer
        return ClassDetailSerializer

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='add-student')
    def add_student(self, request, pk=None):
        """Allow only the class teacher (or admin/staff) to add a student to this class.
        Teachers cannot remove students; Student mutations remain admin-only in StudentViewSet.
        Expected body minimally includes: admission_no, name, dob (YYYY-MM-DD), gender.
        Optional guardian/contact fields supported.
        """
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))

        # Only the class teacher of this class (or admin/staff) can add
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can add students to this class'}, status=status.HTTP_403_FORBIDDEN)

        # Collect and validate payload
        data = getattr(request, 'data', {}) or {}
        required = ['admission_no', 'name', 'dob', 'gender']
        missing = [k for k in required if not str(data.get(k, '')).strip()]
        if missing:
            return Response({'detail': 'Missing required fields', 'missing': missing}, status=status.HTTP_400_BAD_REQUEST)

        payload = {
            'admission_no': str(data.get('admission_no')).strip(),
            'name': str(data.get('name')).strip(),
            'dob': data.get('dob'),
            'gender': str(data.get('gender')).strip(),
            'upi_number': str(data.get('upi_number') or '').strip(),
            'guardian_id': str(data.get('guardian_id') or '').strip(),
            'guardian_name': str(data.get('guardian_name') or '').strip(),
            'guardian_passport_no': str(data.get('guardian_passport_no') or '').strip(),
            'birth_certificate_no': str(data.get('birth_certificate_no') or '').strip(),
            'phone': str(data.get('phone') or '').strip(),
            'email': str(data.get('email') or '').strip(),
            'address': str(data.get('address') or '').strip(),
            'klass': klass,
            'school': getattr(klass, 'school', None),
        }

        # Create student; enforce unique admission_no within the model uniqueness
        try:
            with transaction.atomic():
                stu = Student.objects.create(**payload)
        except IntegrityError:
            return Response({'detail': 'Admission number already exists'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'detail': 'Failed to create student', 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        ser = StudentSerializer(stu, context={'request': request})
        return Response(ser.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='share-fees')
    def share_fees(self, request, pk=None):
        """Send fee balance notifications to guardians/students in this class.
        Access: Only class teacher of this class, or admin/staff.
        Body: { channel?: 'sms' | 'both', min_balance?: number, include_zero?: bool }
        """
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can send fees notifications for this class'}, status=status.HTTP_403_FORBIDDEN)

        channel = (str(request.data.get('channel') or 'sms') or 'sms').lower()
        include_zero = str(request.data.get('include_zero', 'false')).lower() in ('1','true','yes','on')
        try:
            min_balance = float(request.data.get('min_balance', 0) or 0)
        except Exception:
            min_balance = 0.0

        students = Student.objects.filter(klass=klass, is_active=True).select_related('klass').order_by('name')
        if not students.exists():
            return Response({'detail': 'No active students found in this class'}, status=status.HTTP_400_BAD_REQUEST)

        # Prepare messaging context
        school = getattr(klass, 'school', None)
        school_id = getattr(school, 'id', None)
        sender_id = resolve_default_sender_id(school_id) if school_id else None

        notified = 0
        sms_attempts = 0
        email_attempts = 0
        recipients_in_app = []
        items = []

        for stu in students:
            inv_qs = Invoice.objects.filter(student=stu)
            pay_qs = Payment.objects.filter(invoice__student=stu)
            total_billed = float(inv_qs.aggregate(s=Sum('amount'))['s'] or 0)
            total_paid = float(pay_qs.aggregate(s=Sum('amount'))['s'] or 0)
            balance = round(total_billed - total_paid, 2)

            if not include_zero and balance <= 0:
                continue
            if balance < min_balance:
                continue

            name = getattr(stu, 'name', '') or getattr(stu, 'admission_no', '') or f"Student {stu.id}"
            adm = getattr(stu, 'admission_no', '') or ''
            class_name = getattr(klass, 'name', '') or getattr(klass, 'grade_level', '') or 'Class'
            body = (
                f"Fees update: {name}" + (f" (ADM {adm})" if adm else '') +
                f" - {class_name}. Total billed: {total_billed:.2f}, Paid: {total_paid:.2f}, Balance: {balance:.2f}."
            )

            # Queue in-app message if student has linked user
            if sender_id and getattr(stu, 'user_id', None):
                recipients_in_app.append(stu.user_id)

            # SMS to guardian
            phone = getattr(stu, 'guardian_id', None)
            if phone:
                try:
                    ok_sms = send_sms(phone, body)
                except Exception:
                    ok_sms = False
                try:
                    log_delivery(
                        school_id=school_id,
                        channel='sms',
                        recipient=str(phone),
                        ok=bool(ok_sms),
                        message=body,
                        context=f"fees:class:{klass.id};student:{stu.id}",
                    )
                except Exception:
                    pass
                if ok_sms:
                    sms_attempts += 1

            # Optional email
            if channel == 'both':
                email = getattr(stu, 'email', None) or getattr(getattr(stu, 'user', None), 'email', None)
                if email:
                    try:
                        subj = f"Fees balance update - {class_name}"
                        ok_email = send_email_safe(subj, body, email, school_id=school_id)
                    except Exception:
                        ok_email = False
                    try:
                        log_delivery(
                            school_id=school_id,
                            channel='email',
                            recipient=str(email),
                            ok=bool(ok_email),
                            message=body,
                            context=f"fees:class:{klass.id};student:{stu.id}",
                        )
                    except Exception:
                        pass
                    if ok_email:
                        email_attempts += 1

            items.append({'student_id': stu.id, 'balance': balance})
            notified += 1

        # Mirror to chat/messages in bulk
        if sender_id and recipients_in_app:
            try:
                create_messages_for_users(
                    school_id=school_id,
                    sender_id=sender_id,
                    body=f"Fees balances have been updated for your account. Please check portal/SMS.",
                    recipient_user_ids=list(set(recipients_in_app)),
                    system_tag=f"fees_notice:class:{klass.id}",
                )
            except Exception:
                pass

        return Response({
            'detail': 'fees_notifications_queued',
            'class_id': klass.id,
            'students_notified': notified,
            'sms_sent_attempts': sms_attempts,
            'email_sent_attempts': email_attempts,
            'items': items,
        })

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='message-students')
    def message_students(self, request, pk=None):
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can message students in this class'}, status=status.HTTP_403_FORBIDDEN)

        body = str(request.data.get('message') or request.data.get('body') or '').strip()
        if not body:
            return Response({'detail': 'message is required'}, status=status.HTTP_400_BAD_REQUEST)

        students = (
            Student.objects
            .filter(klass=klass, is_active=True)
            .select_related('user', 'klass')
            .order_by('name')
        )
        if not students.exists():
            return Response({'detail': 'No active students found in this class'}, status=status.HTTP_400_BAD_REQUEST)

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        class_name = getattr(klass, 'name', '') or getattr(klass, 'grade_level', '') or 'Class'
        sender_name = getattr(user, 'username', None) or 'Teacher'

        # In-app (single broadcast message with many recipients)
        in_app_message_id = None
        in_app_recipients = 0
        try:
            recipient_user_ids = [int(s.user_id) for s in students if getattr(s, 'user_id', None)]
            recipient_user_ids = list(set(recipient_user_ids))
        except Exception:
            recipient_user_ids = []
        if school_id and recipient_user_ids:
            try:
                msg = Message.objects.create(
                    school_id=school_id,
                    sender_id=getattr(user, 'id', None),
                    body=body,
                    audience=Message.Audience.USERS,
                    system_tag=f"class_broadcast:class:{klass.id}",
                    is_broadcast=True,
                )
                recs = [MessageRecipient(message=msg, user_id=uid) for uid in recipient_user_ids]
                if recs:
                    MessageRecipient.objects.bulk_create(recs, ignore_conflicts=True)
                in_app_message_id = msg.id
                in_app_recipients = len(recipient_user_ids)
            except Exception:
                in_app_message_id = None
                in_app_recipients = 0

        sms_attempts = 0
        sms_ok = 0
        email_attempts = 0
        email_ok = 0

        for stu in students:
            # SMS to guardian phone
            phone = str(getattr(stu, 'guardian_id', None) or '').strip()
            if phone:
                sms_attempts += 1
                ok_sms = False
                try:
                    ok_sms = send_sms(phone, body)
                except Exception:
                    ok_sms = False
                if ok_sms:
                    sms_ok += 1
                try:
                    log_delivery(
                        school_id=school_id,
                        channel='sms',
                        recipient=str(phone),
                        ok=bool(ok_sms),
                        message=body,
                        context=f"classmsg:class:{klass.id};student:{stu.id}",
                    )
                except Exception:
                    pass

            # Email to student email (or linked user email)
            recipient_email = getattr(stu, 'email', None) or getattr(getattr(stu, 'user', None), 'email', None)
            if recipient_email:
                email_attempts += 1
                ok_email = False
                try:
                    subj = f"Message from {sender_name} - {class_name}"
                    ok_email = send_email_safe(subj, body, str(recipient_email), school_id=school_id)
                except Exception:
                    ok_email = False
                if ok_email:
                    email_ok += 1
                try:
                    log_delivery(
                        school_id=school_id,
                        channel='email',
                        recipient=str(recipient_email),
                        ok=bool(ok_email),
                        message=body,
                        context=f"classmsg:class:{klass.id};student:{stu.id}",
                    )
                except Exception:
                    pass

        return Response({
            'detail': 'class_message_sent',
            'class_id': klass.id,
            'in_app_message_id': in_app_message_id,
            'in_app_recipients': in_app_recipients,
            'sms_attempts': sms_attempts,
            'sms_ok': sms_ok,
            'email_attempts': email_attempts,
            'email_ok': email_ok,
        })

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='message-logs')
    def message_logs(self, request, pk=None):
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can view logs for this class'}, status=status.HTTP_403_FORBIDDEN)

        try:
            limit_param = request.query_params.get('limit')
            limit = int(limit_param) if (limit_param and str(limit_param).isdigit()) else 50
            limit = max(1, min(limit, 200))
        except Exception:
            limit = 50

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        if not school_id:
            return Response({'class_id': klass.id, 'items': []})

        items = []

        try:
            msg_qs = (
                Message.objects
                .filter(
                    school_id=school_id,
                    system_tag__in=[f"class_broadcast:class:{klass.id}", f"fees_notice:class:{klass.id}"],
                )
                .select_related('sender')
                .order_by('-created_at', '-id')
            )
            for m in msg_qs[:limit]:
                tag = str(getattr(m, 'system_tag', '') or '')
                category = 'class' if tag.startswith('class_broadcast:') else ('fees' if tag.startswith('fees_notice:') else 'other')
                items.append({
                    'id': f"inapp:{m.id}",
                    'category': category,
                    'channel': 'in_app',
                    'ok': True,
                    'recipient': 'students',
                    'message': (getattr(m, 'body', '') or ''),
                    'created_at': getattr(m, 'created_at', None),
                    'sender': getattr(getattr(m, 'sender', None), 'username', None),
                })
        except Exception:
            pass

        try:
            dl_qs = (
                DeliveryLog.objects
                .filter(
                    school_id=school_id,
                )
                .filter(
                    Q(context__contains=f"classmsg:class:{klass.id};") |
                    Q(context__contains=f"fees:class:{klass.id};")
                )
                .only('id', 'channel', 'recipient', 'ok', 'message_snippet', 'context', 'created_at')
                .order_by('-created_at', '-id')
            )
            for rec in dl_qs[:limit]:
                ctx = str(getattr(rec, 'context', '') or '')
                category = 'fees' if ctx.startswith('fees:') else ('class' if ctx.startswith('classmsg:') else 'other')
                items.append({
                    'id': f"dl:{rec.id}",
                    'category': category,
                    'channel': getattr(rec, 'channel', None),
                    'ok': bool(getattr(rec, 'ok', False)),
                    'recipient': getattr(rec, 'recipient', None),
                    'message': getattr(rec, 'message_snippet', None),
                    'created_at': getattr(rec, 'created_at', None),
                    'sender': None,
                })
        except Exception:
            pass

        items.sort(key=lambda x: (x.get('created_at') is not None, x.get('created_at')), reverse=True)
        return Response({'class_id': klass.id, 'items': items[:limit]})

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='retry-delivery-logs')
    def retry_delivery_logs(self, request, pk=None):
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can retry delivery logs for this class'}, status=status.HTTP_403_FORBIDDEN)

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        if not school_id:
            return Response({'detail': 'No school'}, status=status.HTTP_400_BAD_REQUEST)

        ids = request.data.get('ids') or request.data.get('id')
        if ids is None:
            return Response({'detail': 'ids is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(ids, (list, tuple)):
            ids = [ids]

        cleaned = []
        for x in ids:
            try:
                cleaned.append(int(x))
            except Exception:
                continue
        if not cleaned:
            return Response({'detail': 'No valid ids'}, status=status.HTTP_400_BAD_REQUEST)

        logs = DeliveryLog.objects.filter(id__in=cleaned, school_id=school_id, channel__in=['sms', 'email'])

        results = []
        for rec in logs:
            ctx = str(getattr(rec, 'context', '') or '')
            if (f"classmsg:class:{klass.id};" not in ctx) and (f"fees:class:{klass.id};" not in ctx):
                continue

            msg = str(getattr(rec, 'message_snippet', '') or '').strip()
            if not msg:
                results.append({'id': rec.id, 'ok': False, 'channel': rec.channel, 'recipient': rec.recipient, 'detail': 'empty_message'})
                continue

            ok = False
            try:
                if rec.channel == 'sms':
                    ok = send_sms(str(rec.recipient), msg)
                elif rec.channel == 'email':
                    ok = send_email_safe('Delivery retry', msg, str(rec.recipient), school_id=school_id)
            except Exception:
                ok = False

            try:
                new_ctx = (f"retry_of:{rec.id};" + (ctx or ''))[:100]
                log_delivery(
                    school_id=school_id,
                    channel=str(rec.channel),
                    recipient=str(rec.recipient),
                    ok=bool(ok),
                    message=msg,
                    context=new_ctx,
                )
            except Exception:
                pass

            results.append({'id': rec.id, 'ok': bool(ok), 'channel': rec.channel, 'recipient': rec.recipient})

        return Response({'results': results})

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='delete-old-logs')
    def delete_old_logs(self, request, pk=None):
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can delete logs for this class'}, status=status.HTTP_403_FORBIDDEN)

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        if not school_id:
            return Response({'detail': 'No school'}, status=status.HTTP_400_BAD_REQUEST)

        days_param = request.data.get('days') or request.query_params.get('days')
        before_param = request.data.get('before') or request.query_params.get('before')
        cutoff = None
        if before_param:
            try:
                from django.utils.dateparse import parse_datetime
                dt = parse_datetime(str(before_param))
                if dt:
                    cutoff = dt
            except Exception:
                cutoff = None
        if cutoff is None:
            try:
                days = int(days_param) if days_param is not None else 30
            except Exception:
                days = 30
            cutoff = timezone.now() - timezone.timedelta(days=max(1, days))

        qs = (
            DeliveryLog.objects
            .filter(school_id=school_id, created_at__lt=cutoff)
            .filter(
                Q(context__contains=f"classmsg:class:{klass.id};") |
                Q(context__contains=f"fees:class:{klass.id};")
            )
        )
        try:
            deleted_count, _ = qs.delete()
        except Exception:
            return Response({'detail': 'delete_failed'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'deleted': int(deleted_count), 'cutoff': cutoff})

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='fees-balances')
    def fees_balances(self, request, pk=None):
        """List students in this class with their fee balances.
        Query params: min_balance (number, default 0), include_zero (bool), only_positive (alias).
        Access: class teacher of this class or admin/staff.
        """
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can view balances for this class'}, status=status.HTTP_403_FORBIDDEN)

        try:
            min_balance = float(request.query_params.get('min_balance', 0) or 0)
        except Exception:
            min_balance = 0.0
        include_zero = str(request.query_params.get('include_zero', request.query_params.get('only_positive') and 'false') or 'false').lower() in ('1','true','yes','on')

        students = Student.objects.filter(klass=klass, is_active=True).select_related('klass').order_by('name')
        school = getattr(klass, 'school', None)
        school_id = getattr(school, 'id', None)

        items = []
        total_balance = 0.0
        for stu in students:
            inv_qs = Invoice.objects.filter(student=stu)
            pay_qs = Payment.objects.filter(invoice__student=stu)
            total_billed = float(inv_qs.aggregate(s=Sum('amount'))['s'] or 0)
            total_paid = float(pay_qs.aggregate(s=Sum('amount'))['s'] or 0)
            balance = round(total_billed - total_paid, 2)

            if not include_zero and balance <= 0:
                continue
            if balance < min_balance:
                continue

            items.append({
                'student': {
                    'id': getattr(stu, 'id', None),
                    'name': getattr(stu, 'name', None),
                    'admission_no': getattr(stu, 'admission_no', None),
                },
                'total_billed': round(total_billed, 2),
                'total_paid': round(total_paid, 2),
                'balance': balance,
                'guardian_phone': getattr(stu, 'guardian_id', None),
                'email': getattr(stu, 'email', None) or getattr(getattr(stu, 'user', None), 'email', None),
            })
            total_balance += balance if balance > 0 else 0.0

        return Response({
            'class_id': klass.id,
            'count': len(items),
            'total_balance': round(total_balance, 2),
            'items': items,
        })

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='fees-status')
    def fees_status(self, request, pk=None):
        """Return latest delivery status (SMS/Email) per student for this class' fees notifications.
        Access: class teacher of this class or admin/staff.
        Response: {
          class_id, items: [
            { student_id, sms: {ok, created_at} | null, email: {ok, created_at} | null }
          ]
        }
        Matches DeliveryLog.context pattern: 'fees:class:{klass.id};student:{stu.id}'.
        """
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can view delivery status for this class'}, status=status.HTTP_403_FORBIDDEN)

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        # Build a map of latest statuses per student per channel from DeliveryLog
        status_map = {}
        try:
            logs = (
                DeliveryLog.objects
                .filter(school_id=school_id, context__contains=f"fees:class:{klass.id};", channel__in=['sms', 'email'])
                .only('id', 'channel', 'ok', 'created_at', 'context')
                .order_by('-created_at', '-id')
            )
            for rec in logs:
                ctx = getattr(rec, 'context', '') or ''
                sid = None
                for part in ctx.split(';'):
                    part = part.strip()
                    if part.startswith('student:'):
                        try:
                            sid = int(part.split(':', 1)[1])
                        except Exception:
                            sid = None
                        break
                if not sid:
                    continue
                entry = status_map.setdefault(sid, {'sms': None, 'email': None})
                ch = getattr(rec, 'channel', '')
                if ch in ('sms', 'email') and entry[ch] is None:
                    entry[ch] = {
                        'ok': bool(getattr(rec, 'ok', False)),
                        'created_at': getattr(rec, 'created_at', None),
                    }
                # Early stop if both channels captured for this student
                if entry.get('sms') is not None and entry.get('email') is not None:
                    # continue scanning for others; no break because we need other students
                    pass
        except Exception:
            status_map = {}

        # Ensure we include all active students in class in the response
        items = []
        for stu in Student.objects.filter(klass=klass, is_active=True).only('id'):
            s = status_map.get(getattr(stu, 'id', None)) or {'sms': None, 'email': None}
            items.append({
                'student_id': getattr(stu, 'id', None),
                'sms': s.get('sms'),
                'email': s.get('email'),
            })

        return Response({
            'class_id': klass.id,
            'items': items,
        })

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='fees-resend')
    def fees_resend(self, request, pk=None):
        """Resend a single fees notification (SMS or Email) for a student in this class.
        Body: { student_id: <id>, channel: 'sms' | 'email' }
        Access: class teacher of this class or admin/staff.
        """
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can resend for this class'}, status=status.HTTP_403_FORBIDDEN)

        # Validate inputs
        try:
            student_id = int(request.data.get('student_id'))
        except Exception:
            return Response({'detail': 'student_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        channel = str(request.data.get('channel') or '').lower()
        if channel not in ('sms', 'email'):
            return Response({'detail': 'channel must be sms or email'}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure student belongs to this class
        try:
            stu = Student.objects.select_related('klass').get(id=student_id, klass=klass)
        except Student.DoesNotExist:
            return Response({'detail': 'Student not found in this class'}, status=status.HTTP_404_NOT_FOUND)

        # Compose latest balance message
        inv_qs = Invoice.objects.filter(student=stu)
        pay_qs = Payment.objects.filter(invoice__student=stu)
        total_billed = float(inv_qs.aggregate(s=Sum('amount'))['s'] or 0)
        total_paid = float(pay_qs.aggregate(s=Sum('amount'))['s'] or 0)
        balance = round(total_billed - total_paid, 2)
        name = getattr(stu, 'name', '') or getattr(stu, 'admission_no', '') or f"Student {stu.id}"
        adm = getattr(stu, 'admission_no', '') or ''
        class_name = getattr(klass, 'name', '') or getattr(klass, 'grade_level', '') or 'Class'
        body = (
            f"Fees update: {name}" + (f" (ADM {adm})" if adm else '') +
            f" - {class_name}. Total billed: {total_billed:.2f}, Paid: {total_paid:.2f}, Balance: {balance:.2f}."
        )

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        ok = False
        if channel == 'sms':
            phone = getattr(stu, 'guardian_id', None)
            if not phone:
                return Response({'detail': 'No guardian phone on record'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                ok = send_sms(phone, body)
            except Exception:
                ok = False
            try:
                log_delivery(
                    school_id=school_id,
                    channel='sms',
                    recipient=str(phone),
                    ok=bool(ok),
                    message=body,
                    context=f"fees:class:{klass.id};student:{stu.id}",
                )
            except Exception:
                pass
        else:
            email = getattr(stu, 'email', None) or getattr(getattr(stu, 'user', None), 'email', None)
            if not email:
                return Response({'detail': 'No email on record'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                subj = f"Fees balance update - {class_name}"
                ok = send_email_safe(subj, body, email, school_id=school_id)
            except Exception:
                ok = False
            try:
                log_delivery(
                    school_id=school_id,
                    channel='email',
                    recipient=str(email),
                    ok=bool(ok),
                    message=body,
                    context=f"fees:class:{klass.id};student:{stu.id}",
                )
            except Exception:
                pass

        return Response({'ok': bool(ok)})

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='mine')
    def mine(self, request):
        """Return classes relevant to the authenticated teacher.
        - If the requester is a teacher (non-staff), include classes where they are either:
          1) the class teacher (Class.teacher = user), OR
          2) assigned to teach any subject in the class via ClassSubjectTeacher.
        - Admins/staff get all classes in their school.
        """
        qs = self.get_queryset()
        user = request.user
        if getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            qs = qs.filter(Q(teacher=user) | Q(subject_teachers__teacher=user)).distinct()
        ser = self.get_serializer(qs, many=True)
        return Response(ser.data)

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='students')
    def students(self, request, pk=None):
        """Return the list of students enrolled in this class.
        Used by the Class details > Students tab.
        """
        klass = self.get_object()
        # Scope by the class directly; order by name for consistent display
        qs = Student.objects.filter(klass=klass).order_by('name')
        # Use lightweight serializer for list performance
        from .serializers import StudentListSerializer
        ser = StudentListSerializer(qs, many=True, context={'request': request})
        return Response(ser.data)

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='history')
    def history(self, request, pk=None):
        """Return class history events derived from StudentClassHistory.
        Includes promotions, assignments, moves, and graduations involving this class.
        """
        klass = self.get_object()
        try:
            rows = (
                StudentClassHistory.objects
                .filter(Q(from_class=klass) | Q(to_class=klass))
                .select_related('student', 'from_class', 'to_class')
                .order_by('created_at', 'id')
            )
            events = []
            students_in = []
            students_out = []
            for h in rows:
                ev = {
                    'id': getattr(h, 'id', None),
                    'student': {
                        'id': getattr(getattr(h, 'student', None), 'id', None),
                        'name': getattr(getattr(h, 'student', None), 'name', None),
                        'admission_no': getattr(getattr(h, 'student', None), 'admission_no', None),
                    },
                    'action': getattr(h, 'action', None),
                    'from': getattr(getattr(h, 'from_class', None), 'name', None),
                    'to': getattr(getattr(h, 'to_class', None), 'name', None) or ('Graduated' if getattr(h, 'action', '') == 'graduated' else None),
                    'year': getattr(h, 'year', None),
                    'term': getattr(h, 'term', None),
                    'note': getattr(h, 'note', ''),
                    'created_at': getattr(h, 'created_at', None),
                }
                events.append(ev)

                # Map into "students in" / "students out" buckets used by the AdminClassProfile UI
                action = ev.get('action')
                # Students coming into this class
                if getattr(h, 'to_class_id', None) == klass.id and action in ('assigned', 'promoted', 'moved'):
                    students_in.append({
                        'student_id': ev['student']['id'],
                        'student_name': ev['student']['name'],
                        'from': ev.get('from') or 'Unassigned',
                        'year': ev.get('year'),
                        'term': ev.get('term'),
                        'created_at': ev.get('created_at'),
                    })
                # Students leaving this class (including graduation / unassignment)
                if getattr(h, 'from_class_id', None) == klass.id and action in ('promoted', 'moved', 'graduated', 'unassigned'):
                    students_out.append({
                        'student_id': ev['student']['id'],
                        'student_name': ev['student']['name'],
                        'to': ev.get('to') or ('Unassigned' if action == 'unassigned' else None),
                        'year': ev.get('year'),
                        'term': ev.get('term'),
                        'created_at': ev.get('created_at'),
                    })
        except Exception:
            events = []
            students_in = []
            students_out = []
        # Lightweight summary
        summary = {
            'total_events': len(events),
            'promoted': len([e for e in events if e.get('action') == 'promoted']),
            'assigned': len([e for e in events if e.get('action') == 'assigned']),
            'moved': len([e for e in events if e.get('action') == 'moved']),
            'graduated': len([e for e in events if e.get('action') == 'graduated']),
            'unassigned': len([e for e in events if e.get('action') == 'unassigned']),
        }
        # Exams grouped by academic year and term for this class
        exams_by_term = []
        try:
            qs_exams = Exam.objects.filter(klass=klass).order_by('year', 'term', 'date', 'id')
            grouped = {}
            for ex in qs_exams:
                key = (getattr(ex, 'year', None), getattr(ex, 'term', None))
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(getattr(ex, 'name', '') or f"Exam #{getattr(ex, 'id', '')}")
            for (year, term), names in grouped.items():
                exams_by_term.append({
                    'year': year,
                    'term': term,
                    'exams': ', '.join([n for n in names if n]),
                })
        except Exception:
            exams_by_term = []

        payload = {
            'class': {'id': klass.id, 'name': klass.name, 'grade_level': klass.grade_level},
            'events': events,
            'summary': summary,
            # Structured history used by AdminClassProfile
            'students_in': students_in,
            'students_out': students_out,
            'exams_by_term': exams_by_term,
        }
        return Response(payload)

    @action(detail=True, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='share-results')
    def share_results(self, request, pk=None):
        """Send concise exam results summaries to parents/guardians for this class.
        Body: { exam_id: <optional exam id>, channel: 'sms' | 'both' }
        - SMS is sent to Student.guardian_id (phone).
        - Email (when requested) is sent to student.email or linked user.email.
        Summaries include total marks, average and class position.
        """
        klass = self.get_object()
        # Restrict teachers to their own class
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can share results for this class'}, status=status.HTTP_403_FORBIDDEN)
        school_id = getattr(klass, 'school_id', None)
        exam_id = request.data.get('exam_id') or None
        channel = (str(request.data.get('channel') or 'sms') or 'sms').lower()
        # Resolve exam: prefer provided id, else latest for this class
        exam = None
        if exam_id:
            try:
                exam = Exam.objects.get(id=int(exam_id), klass=klass)
            except (Exam.DoesNotExist, ValueError, TypeError):
                return Response({'detail': 'Exam not found for this class'}, status=status.HTTP_404_NOT_FOUND)
        else:
            exam = (
                Exam.objects
                .filter(klass=klass)
                .order_by('-date', '-id')
                .first()
            )
            if not exam:
                return Response({'detail': 'No exams found for this class'}, status=status.HTTP_400_BAD_REQUEST)

        # Build per-student totals and averages. Use the same filtering rules as the exam summary
        # so we do not miss historical results where the student may have moved classes.
        results = (
            ExamResult.objects
            .filter(exam=exam, student__is_active=True)
            .filter(subject__is_examinable=True)
            .select_related('student', 'subject')
        )
        if not results.exists():
            return Response({'detail': 'No results captured for this exam'}, status=status.HTTP_400_BAD_REQUEST)

        per_student = {}
        for r in results:
            s = r.student
            entry = per_student.setdefault(s.id, {
                'student': s,
                # aggregate per subject: { code: {marks_sum, denom_sum} }
                'subject_aggs': {},
            })
            try:
                sub_code = getattr(r.subject, 'code', None) or getattr(r.subject, 'name', None) or ''
            except Exception:
                sub_code = ''
            if not sub_code:
                continue
            # Determine denominator for this result
            try:
                marks_val = float(r.marks)
            except Exception:
                marks_val = 0.0
            denom_val = None
            try:
                if r.out_of is not None:
                    dv = float(r.out_of)
                    denom_val = dv if dv > 0 else None
            except Exception:
                denom_val = None
            if denom_val is None:
                try:
                    comp_max = float(getattr(getattr(r, 'component', None), 'max_marks', 0) or 0)
                    denom_val = comp_max if comp_max > 0 else None
                except Exception:
                    denom_val = None
            # Fallback: skip if we cannot infer a denominator (cannot compute percentage)
            if denom_val is None or denom_val <= 0:
                continue
            agg = entry['subject_aggs'].setdefault(sub_code, {'marks_sum': 0.0, 'denom_sum': 0.0})
            agg['marks_sum'] += marks_val
            agg['denom_sum'] += denom_val

        # Compute per-student subject percentages and totals; sort by total desc for positions (class-level)
        ordered = []
        for sid, data in per_student.items():
            subj_pct_ints = {}
            total_pct = 0.0
            count = 0
            for code, agg in (data.get('subject_aggs') or {}).items():
                ms = float(agg.get('marks_sum') or 0.0)
                ds = float(agg.get('denom_sum') or 0.0)
                if ds > 0:
                    pct = (ms / ds) * 100.0
                    subj_pct_ints[code] = int(round(pct))
                    total_pct += pct
                    count += 1
            avg_pct = (total_pct / count) if count else 0.0
            data['subject_pct_ints'] = subj_pct_ints
            ordered.append({
                'student': data['student'],
                'total': round(total_pct),  # whole-number total of percentages
                'average': round(avg_pct, 2),
            })
        ordered.sort(key=lambda x: x['total'], reverse=True)
        # Assign positions (simple ranking by total, ties share same position)
        last_total = None
        position = 0
        for idx, row in enumerate(ordered):
            if last_total is None or row['total'] < last_total:
                position = idx + 1
                last_total = row['total']
            row['position'] = position

        # Compute grade-level positions across all classes for the same exam cohort
        grade_positions = {}
        try:
            grade_tag = getattr(klass, 'grade_level', None) or getattr(exam, 'grade_level_tag', None)
            school_id = getattr(klass, 'school_id', None)
            # Identify exams forming the same cohort (same grade level, term, year, and name) within the school
            cohort_exams = Exam.objects.filter(
                klass__school_id=school_id,
                grade_level_tag=grade_tag,
                year=getattr(exam, 'year', None),
                term=getattr(exam, 'term', None),
                name=getattr(exam, 'name', None),
            ).only('id')
            if cohort_exams.exists():
                cohort_results = (
                    ExamResult.objects
                    .filter(exam__in=cohort_exams, student__is_active=True, subject__is_examinable=True)
                    .select_related('student')
                )
                per_student_grade = {}
                for r in cohort_results:
                    s = r.student
                    e = per_student_grade.setdefault(s.id, {'student': s, 'total': 0.0, 'count': 0})
                    e['total'] += float(r.marks)
                    e['count'] += 1
                ordered_grade = []
                for sid, data in per_student_grade.items():
                    cnt = data['count'] or 1
                    ordered_grade.append({'student': data['student'], 'total': round(data['total'], 2), 'avg': round(data['total']/cnt, 2)})
                ordered_grade.sort(key=lambda x: x['total'], reverse=True)
                last = None
                pos = 0
                for idx, row in enumerate(ordered_grade):
                    if last is None or row['total'] < last:
                        pos = idx + 1
                        last = row['total']
                    grade_positions[getattr(row['student'], 'id', None)] = {'position': pos, 'out_of': len(ordered_grade)}
        except Exception:
            grade_positions = {}

        # Prepare optional chat sender for in-app/email/SMS via unified Messages system
        sender_id = resolve_default_sender_id(school_id) if school_id else None
        recipient_user_ids = []

        # Resolve class/grade and class teacher info for richer messages
        grade_label = getattr(klass, 'grade_level', '') or ''
        class_name = getattr(klass, 'name', '') or grade_label
        teacher_obj = getattr(klass, 'teacher', None)
        teacher_name = ''
        teacher_phone = ''
        try:
            if teacher_obj:
                first = getattr(teacher_obj, 'first_name', '') or ''
                last = getattr(teacher_obj, 'last_name', '') or ''
                teacher_name = (first + ' ' + last).strip() or getattr(teacher_obj, 'username', '') or ''
                # Prefer TeacherProfile.phone, fallback to user.phone
                from .models import TeacherProfile as TP
                prof_phone = (
                    TP.objects.filter(user=teacher_obj).values_list('phone', flat=True).first()
                    if TP is not None else None
                )
                teacher_phone = prof_phone or getattr(teacher_obj, 'phone', '') or ''
        except Exception:
            teacher_name = teacher_name or ''
            teacher_phone = teacher_phone or ''

        # Send SMS / email per student
        exam_name = getattr(exam, 'name', 'Exam')
        sent_sms = 0
        sent_email = 0
        for row in ordered:
            stu = row['student']
            total = row['total']
            pos = row['position']
            name = getattr(stu, 'name', '') or getattr(stu, 'admission_no', '') or f"Student {stu.id}"
            adm = getattr(stu, 'admission_no', '') or ''
            # Subject performance summary like: MATH 80, ENG 75, SCI 78
            subj_marks = per_student.get(stu.id, {}).get('subject_pct_ints', {})
            try:
                items = sorted(subj_marks.items(), key=lambda kv: kv[0])
            except Exception:
                items = list(subj_marks.items())
            subj_parts = []
            for code, mark in items:
                try:
                    mark_val = int(round(float(mark)))
                except Exception:
                    mark_val = 0
                subj_parts.append(f"{code} {mark_val}")
            subjects_str = ', '.join(subj_parts)

            # Build tailored message including student, class, grade, subjects and class teacher
            base = f"{exam_name}: {name}"
            if adm:
                base += f" (ADM {adm})"
            if grade_label or class_name:
                base += f" - {grade_label or class_name}"
                if class_name and class_name != grade_label:
                    base += f" {class_name}"
            # Positions
            cls_total = len(ordered)
            grade_pos = grade_positions.get(getattr(stu, 'id', None)) or {}
            pos_grade = grade_pos.get('position')
            out_grade = grade_pos.get('out_of')
            pos_text = f"Class Pos {pos}/{cls_total}"
            if pos_grade and out_grade:
                pos_text += f"; Grade Pos {pos_grade}/{out_grade}"
            summary = f" Total {int(round(total))}, {pos_text}."
            subjects_clause = f" Subjects: {subjects_str}." if subjects_str else ''
            teacher_clause = ''
            if teacher_name:
                teacher_clause = f" Class teacher: {teacher_name}"
                if teacher_phone:
                    teacher_clause += f" ({teacher_phone})."
                else:
                    teacher_clause += '.'
            msg = base + summary + subjects_clause + ' ' + teacher_clause

            # Queue in-app + student-targeted email/SMS via Messages if the student has a linked user
            if sender_id and getattr(stu, 'user_id', None):
                recipient_user_ids.append(stu.user_id)

            # SMS to guardian
            phone = getattr(stu, 'guardian_id', None)
            if phone and channel in ('sms', 'both'):
                ok_sms = False
                try:
                    ok_sms = send_sms(phone, msg)
                except Exception:
                    ok_sms = False
                try:
                    log_delivery(
                        school_id=school_id,
                        channel='sms',
                        recipient=str(phone),
                        ok=bool(ok_sms),
                        message=msg,
                        context=f"results:exam:{exam.id};student:{stu.id}",
                    )
                except Exception:
                    pass
                if ok_sms:
                    sent_sms += 1

            # Email
            if channel == 'both':
                email = getattr(stu, 'email', None) or getattr(getattr(stu, 'user', None), 'email', None)
                if email:
                    ok_email = False
                    try:
                        ok_email = send_email_safe(exam_name, msg, email, school_id=school_id)
                    except Exception:
                        ok_email = False
                    try:
                        log_delivery(
                            school_id=school_id,
                            channel='email',
                            recipient=str(email),
                            ok=bool(ok_email),
                            message=msg,
                            context=f"results:exam:{exam.id};student:{stu.id}",
                        )
                    except Exception:
                        pass
                    if ok_email:
                        sent_email += 1

        # Mirror to chat/messages so students also see this in-app and via the generic delivery pipeline
        if sender_id and recipient_user_ids:
            try:
                # One shared body per exam; this will create a Message per user and queue email/SMS using user.phone/email
                body = f"{exam_name} results are available for your class {klass.name}. Please check your portal or SMS for details."
                create_messages_for_users(
                    school_id=school_id,
                    sender_id=sender_id,
                    body=body,
                    recipient_user_ids=list(set(recipient_user_ids)),
                    system_tag='exam_results',
                )
            except Exception:
                pass

        return Response({
            'detail': 'results_notifications_queued',
            'exam_id': exam.id,
            'students': len(ordered),
            'sms_sent_attempts': sent_sms,
            'email_sent_attempts': sent_email,
        })

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='results-status')
    def results_status(self, request, pk=None):
        """Return latest delivery status (SMS/Email) per student for a given exam in this class.
        Query: ?exam=<exam_id>. If absent, use the latest exam for the class.
        Access: class teacher of this class or admin/staff.
        Response: { exam_id, class_id, items: [ {student_id, sms:{ok,created_at}|null, email:{ok,created_at}|null} ],
                    totals: { sms: {sent, failed}, email: {sent, failed} } }
        """
        klass = self.get_object()
        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            if not (getattr(user, 'role', None) == 'teacher' and getattr(klass, 'teacher_id', None) == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can view delivery status for this class'}, status=status.HTTP_403_FORBIDDEN)

        # Resolve exam
        exam_id = request.query_params.get('exam')
        if exam_id:
            try:
                exam = Exam.objects.get(id=int(exam_id), klass=klass)
            except (Exam.DoesNotExist, ValueError, TypeError):
                return Response({'detail': 'Exam not found for this class'}, status=status.HTTP_404_NOT_FOUND)
        else:
            exam = (
                Exam.objects
                .filter(klass=klass)
                .order_by('-date', '-id')
                .first()
            )
            if not exam:
                return Response({'detail': 'No exams found for this class'}, status=status.HTTP_400_BAD_REQUEST)

        school_id = getattr(getattr(klass, 'school', None), 'id', None)
        status_map = {}
        sms_sent = sms_failed = email_sent = email_failed = 0
        try:
            logs = (
                DeliveryLog.objects
                .filter(school_id=school_id, context__contains=f"results:exam:{exam.id};", channel__in=['sms','email'])
                .only('id','channel','ok','created_at','context')
                .order_by('-created_at','-id')
            )
            for rec in logs:
                # Count totals by channel
                if rec.channel == 'sms':
                    if rec.ok:
                        sms_sent += 1
                    else:
                        sms_failed += 1
                elif rec.channel == 'email':
                    if rec.ok:
                        email_sent += 1
                    else:
                        email_failed += 1

                ctx = getattr(rec, 'context', '') or ''
                sid = None
                for part in ctx.split(';'):
                    part = part.strip()
                    if part.startswith('student:'):
                        try:
                            sid = int(part.split(':', 1)[1])
                        except Exception:
                            sid = None
                        break
                if not sid:
                    continue
                entry = status_map.setdefault(sid, {'sms': None, 'email': None})
                ch = getattr(rec, 'channel', '')
                if ch in ('sms','email') and entry[ch] is None:
                    entry[ch] = {'ok': bool(getattr(rec,'ok',False)), 'created_at': getattr(rec,'created_at',None)}
        except Exception:
            status_map = {}

        items = []
        for stu in Student.objects.filter(klass=klass, is_active=True).only('id'):
            s = status_map.get(getattr(stu, 'id', None)) or {'sms': None, 'email': None}
            items.append({'student_id': getattr(stu, 'id', None), 'sms': s.get('sms'), 'email': s.get('email')})

        return Response({
            'exam_id': exam.id,
            'class_id': klass.id,
            'items': items,
            'totals': {
                'sms': {'sent': sms_sent, 'failed': sms_failed},
                'email': {'sent': email_sent, 'failed': email_failed},
            }
        })

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='promote')
    def promote(self, request, pk=None):
        """Promote or graduate all students in this class.
        Rules:
        - Grade 9: mark as graduated (with fee balance clearance check).
        - Other grades: promote to next grade in SAME stream.
          If the next class exists but is NOT empty, block with 400.
        """
        from django.db import transaction
        from .models import AcademicYear as AcademicYearModel, Class as ClassModel, StudentClassHistory as HistoryModel
        import re

        klass = self.get_object()
        school = getattr(klass, 'school', None)
        if not school:
            return Response({'detail': 'Class has no associated school'}, status=status.HTTP_400_BAD_REQUEST)

        # Resolve current academic year to tag history years
        ay = AcademicYearModel.objects.filter(school=school, is_current=True).first()
        grad_year = None
        if ay and getattr(ay, 'end_date', None):
            try:
                grad_year = int(ay.end_date.year)
            except Exception:
                grad_year = None

        # Parse grade number
        current_grade = ClassModel.format_grade_level(klass.grade_level)
        m_named = re.search(r'\bgrade\s*(\d{1,2})\b', current_grade, flags=re.IGNORECASE)
        match = m_named if m_named else re.search(r'\b(\d{1,2})\b', current_grade)
        if not match:
            return Response({'detail': f'Could not determine numeric grade from "{current_grade}"'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            current_grade_num = int(match.group(1) if match.lastindex else match.group())
        except Exception:
            return Response({'detail': f'Failed to parse grade from "{current_grade}"'}, status=status.HTTP_400_BAD_REQUEST)

        summary = {
            'class_id': klass.id,
            'from': getattr(klass, 'name', ''),
            'mode': None,
            'students_moved': 0,
            'students_graduated': 0,
        }

        with transaction.atomic():
            if current_grade_num == 9:
                # Graduation path: graduate all students in this class.
                # Finance data (invoices/payments) is left untouched so any fee arrears remain recorded.
                moved_count = 0
                for stu in klass.student_set.select_for_update().all():
                    try:
                        HistoryModel.objects.create(
                            student=stu,
                            from_class=klass,
                            to_class=None,
                            action='graduated',
                            year=grad_year,
                            term=None,
                            note='Class-level graduation',
                        )
                    except Exception:
                        pass
                    stu.klass = None
                    stu.is_graduated = True
                    stu.is_active = False
                    stu.graduation_year = grad_year
                    stu.school = school
                    stu.save(update_fields=['klass', 'is_graduated', 'is_active', 'graduation_year', 'school'])
                    moved_count += 1
                summary['mode'] = 'graduated'
                summary['students_graduated'] = moved_count
                # If all students were graduated (no students left in the class), clear the class teacher
                if moved_count > 0 and not klass.student_set.exists():
                    klass.teacher = None
                    klass.save(update_fields=['teacher'])
            else:
                # Promotion path to next grade in same stream
                new_grade_num = current_grade_num + 1
                target_grade = ClassModel.format_grade_level(str(new_grade_num))
                target = ClassModel.objects.filter(
                    school=school,
                    stream=klass.stream,
                    grade_level=target_grade,
                ).first()

                # Enforce: if target exists and has students, block promotion
                if target and target.student_set.exists():
                    return Response(
                        {
                            'detail': 'Promotion blocked: next class already has students. Clear or use a different class.',
                            'next_class_id': target.id,
                            'next_class_name': getattr(target, 'name', ''),
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if target:
                    # Move students into existing (empty) target class
                    moved_count = 0
                    for stu in klass.student_set.select_for_update().all():
                        try:
                            HistoryModel.objects.create(
                                student=stu,
                                from_class=klass,
                                to_class=target,
                                action='promoted',
                                year=grad_year,
                                term=None,
                                note='Class-level promotion to next grade',
                            )
                        except Exception:
                            pass
                        stu.klass = target
                        stu.is_graduated = False
                        stu.school = school
                        stu.save(update_fields=['klass', 'is_graduated', 'school'])
                        moved_count += 1
                    summary['mode'] = 'moved_to_existing_class'
                    summary['to_id'] = target.id
                    summary['to'] = getattr(target, 'name', '')
                    summary['students_moved'] = moved_count
                else:
                    # In-place rename to new grade (no target exists)
                    from django.db import IntegrityError
                    try:
                        original_name = klass.name
                        klass.grade_level = str(new_grade_num)
                        klass.save(update_fields=['grade_level', 'name'])
                        summary['mode'] = 'renamed_class'
                        summary['from'] = original_name
                        summary['to'] = klass.name
                        summary['students_moved'] = klass.student_set.count()
                    except IntegrityError:
                        # Fallback: target created concurrently; move into it (must be empty due to previous check)
                        target = ClassModel.objects.filter(
                            school=school,
                            stream=klass.stream,
                            grade_level=target_grade,
                        ).first()
                        if not target:
                            raise
                        # Enforce emptiness on fallback as well
                        if target.student_set.exists():
                            return Response(
                                {
                                    'detail': 'Promotion blocked: next class already has students. Clear or use a different class.',
                                    'next_class_id': target.id,
                                    'next_class_name': getattr(target, 'name', ''),
                                },
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                        moved_count = 0
                        for stu in klass.student_set.select_for_update().all():
                            try:
                                HistoryModel.objects.create(
                                    student=stu,
                                    from_class=klass,
                                    to_class=target,
                                    action='promoted',
                                    year=grad_year,
                                    term=None,
                                    note='Class-level promotion (fallback move)',
                                )
                            except Exception:
                                pass
                            stu.klass = target
                            stu.is_graduated = False
                            stu.school = school
                            stu.save(update_fields=['klass', 'is_graduated', 'school'])
                            moved_count += 1
                        summary['mode'] = 'moved_to_existing_class_fallback'
                        summary['to_id'] = target.id
                        summary['to'] = getattr(target, 'name', '')
                        summary['students_moved'] = moved_count

        return Response({'detail': 'Class promotion completed', 'summary': summary})

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='add-students')
    def add_students(self, request, pk=None):
        """Admin: Add students without a class to this class.
        Body: { students: [<id>, ...] }
        Rules:
        - Only students with klass IS NULL are considered.
        - Scope to same school as the class.
        - Records StudentClassHistory(action='assigned').
        Returns { assigned: [ids], skipped: [{id, reason}], count }
        """
        klass = self.get_object()
        try:
            ids = request.data.get('students') or []
            if isinstance(ids, str):
                ids = [int(x.strip()) for x in ids.split(',') if x.strip()]
            ids = [int(x) for x in ids]
        except Exception:
            return Response({'detail': 'students must be an array of IDs or a comma-separated string'}, status=400)

        if not ids:
            return Response({'detail': 'No student IDs provided'}, status=400)

        school_id = getattr(klass, 'school_id', None)
        qs = Student.objects.filter(id__in=ids, klass__isnull=True)
        if school_id:
            qs = qs.filter(school_id=school_id)

        found_ids = set(qs.values_list('id', flat=True))
        assigned = []
        skipped = []

        # Prepare current term year/number best-effort
        year = None
        term_num = None
        try:
            from .models import Term
            t = Term.objects.filter(academic_year__school_id=school_id, is_current=True).first()
            if t:
                term_num = int(getattr(t, 'number', None) or 0) or None
                try:
                    year = int(getattr(t.academic_year.end_date, 'year', None) or getattr(t.academic_year.start_date, 'year', None))
                except Exception:
                    year = None
        except Exception:
            pass

        for sid in ids:
            if sid not in found_ids:
                # Determine reason
                reason = 'not_found_or_already_in_class'
                try:
                    stu = Student.objects.filter(id=sid).first()
                    if stu and stu.klass_id:
                        reason = 'already_in_class'
                    elif stu and school_id and (stu.school_id not in (None, school_id)):
                        reason = 'different_school'
                    elif not stu:
                        reason = 'not_found'
                except Exception:
                    pass
                skipped.append({'id': sid, 'reason': reason})
                continue
            stu = qs.filter(id=sid).first()
            if not stu:
                skipped.append({'id': sid, 'reason': 'not_found_or_already_in_class'})
                continue
            prev = getattr(stu, 'klass', None)
            stu.klass = klass
            # Persist school on student in case they previously had none
            if not getattr(stu, 'school_id', None):
                stu.school_id = school_id
            stu.save(update_fields=['klass','school'])
            try:
                StudentClassHistory.objects.create(
                    student=stu,
                    from_class=prev,
                    to_class=klass,
                    action='assigned',
                    year=year,
                    term=term_num,
                    note='Bulk add via class action'
                )
            except Exception:
                pass
            assigned.append(sid)

        return Response({'assigned': assigned, 'skipped': skipped, 'count': len(assigned)})


class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    # Allow teachers to read, admins to manage
    permission_classes = [IsTeacherOrAdmin]

    def get_permissions(self):
        """Loosen gate for unsafe methods; enforce with object-level checks.
        - SAFE methods: keep IsTeacherOrAdmin
        - UNSAFE (POST/PATCH/PUT/DELETE): allow authenticated and rely on perform_create/_can_manage_exam
        """
        from rest_framework.permissions import IsAuthenticated
        if self.request and self.request.method not in permissions.SAFE_METHODS:
            return [IsAuthenticated()]
        return [perm() if isinstance(perm, type) else perm for perm in self.permission_classes]

    def _is_admin(self, request):
        u = getattr(request, 'user', None)
        return bool(u and (u.role == 'admin' or u.is_staff or u.is_superuser))

    def _can_manage_exam(self, request, exam=None):
        """Return True if requester can modify the given exam.
        Rules:
        - Admin/staff: always true.
        - Teacher: allowed if they are the class teacher OR assigned to any subject in the class via ClassSubjectTeacher.
        - Others: false.
        Also scopes to same school when applicable.
        """
        if self._is_admin(request):
            return True
        user = getattr(request, 'user', None)
        if not user or getattr(user, 'role', None) != 'teacher':
            return False
        if exam is None:
            return False
        # Same school check
        school = getattr(user, 'school', None)
        try:
            if school and getattr(exam.klass, 'school_id', None) not in (None, getattr(school, 'id', None)):
                return False
        except Exception:
            # If we cannot determine, fail closed
            return False
        # Allow any teacher from the same school to manage (broad permission)
        if getattr(user, 'role', None) == 'teacher':
            return True
        # Fallback checks (kept for clarity)
        if getattr(exam.klass, 'teacher_id', None) == getattr(user, 'id', None):
            return True
        try:
            return ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user).exists()
        except Exception:
            return False

    def get_queryset(self):
        qs = super().get_queryset().select_related('klass')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        # Admins can see all; others are scoped to their school
        if school and not self._is_admin(self.request):
            qs = qs.filter(klass__school=school)

        # If requester is a teacher, restrict to classes they teach (class teacher or subject teacher).
        # However, for detail actions like summary/rank/report_card_pdf/compare_subjects, allow access to
        # published exams from the same school to avoid 404 before the per-action permission checks.
        user = getattr(self.request, 'user', None)
        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            try:
                action = getattr(self, 'action', None)
            except Exception:
                action = None
            is_detail_action = action in ('retrieve', 'summary', 'rank', 'report_card_pdf', 'compare_subjects')
            if is_detail_action:
                # Include exams taught by the teacher OR published ones.
                # Some datasets set textual status='published' without toggling the boolean flag.
                # Allow both forms to avoid false 404s on detail endpoints.
                try:
                    status_published = Q(status__iexact='published')
                except Exception:
                    status_published = Q()
                qs = qs.filter(
                    Q(klass__teacher=user) |
                    Q(klass__subject_teachers__teacher=user) |
                    Q(published=True) |
                    status_published
                ).distinct()
            else:
                qs = qs.filter(Q(klass__teacher=user) | Q(klass__subject_teachers__teacher=user)).distinct()

        # Default: limit to current academic year unless include_history=true (SAFE methods only, non-admins)
        if self.request.method in permissions.SAFE_METHODS and not self._is_admin(self.request):
            include_history = str(self.request.query_params.get('include_history', 'false')).lower() in ('1','true','yes')
            if not include_history and school:
                try:
                    ay = AcademicYear.objects.filter(school=school, is_current=True).first()
                    if ay:
                        qs = qs.filter(date__gte=ay.start_date, date__lte=ay.end_date)
                except Exception:
                    pass

        # When filtering by grade, use the persistent grade tag captured at exam creation
        # Fallback: include exams where grade_level_tag is empty/null but the class's current grade_level matches.
        grade = self.request.query_params.get('grade')
        if grade:
            try:
                from .models import Class as ClassModel
                formatted = ClassModel.format_grade_level(grade)
            except Exception:
                formatted = grade
            # Apply OR fallback when tag is missing
            try:
                missing_tag = Q(grade_level_tag__isnull=True) | Q(grade_level_tag="")
                qs = qs.filter(Q(grade_level_tag=formatted) | (missing_tag & Q(klass__grade_level=formatted)))
            except Exception:
                qs = qs.filter(grade_level_tag=formatted)
        return qs

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='by-name')
    def by_name(self, request):
        """Return all exams that share the same name, optionally filtered by year and term.
        Query params:
          - name: required, exam name (case-insensitive exact match)
          - year: optional int
          - term: optional int (1..3)
          - include_history: optional bool, default false (when false, limits to current Academic Year)
        Response: list of exams with class/grade labels for rendering a grouped view.
        """
        name = request.query_params.get('name')
        if not name:
            return Response({'detail': 'name is required'}, status=status.HTTP_400_BAD_REQUEST)

        qs = self.get_queryset().filter(name__iexact=name)

        # Optional filters
        year = request.query_params.get('year')
        term = request.query_params.get('term')
        if year:
            try:
                qs = qs.filter(year=int(year))
            except Exception:
                pass
        if term:
            try:
                qs = qs.filter(term=int(term))
            except Exception:
                pass

        # Order consistently by year desc, term asc, grade label, stream name, date
        qs = qs.order_by('-year', 'term', 'grade_level_tag', 'klass__stream__name', 'date', 'id')

        data = []
        for e in qs.select_related('klass', 'klass__stream'):
            data.append({
                'id': e.id,
                'name': e.name,
                'year': e.year,
                'term': e.term,
                'date': e.date,
                'total_marks': e.total_marks,
                'published': e.published,
                'published_at': e.published_at,
                'grade_level_tag': getattr(e, 'grade_level_tag', None),
                'klass': {
                    'id': getattr(e.klass, 'id', None),
                    'name': getattr(e.klass, 'name', None),
                    'grade_level': getattr(e.klass, 'grade_level', None),
                    'stream': getattr(getattr(e.klass, 'stream', None), 'name', None),
                }
            })
        return Response({'name': name, 'items': data})

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin], url_path='common-bulk-create')
    def common_bulk_create(self, request):
        """Admin-only: Create common exams across classes.
        Body:
        {
          "names": ["CAT 1","CAT 2","CAT 3","CAT 4","CAT 5"],  # required, 1..N
          "term": 1,                                                    # required (1..3)
          "year": 2025,                                                 # required
          "total_marks": 100,                                           # optional, default 100
          "date": "2025-02-10",                                       # optional default date for all
          "dates": ["2025-02-10","2025-03-01", ...],                # optional per-name dates; same length as names
          "grade": "Grade 4",                                         # optional filter; otherwise all classes
          "publish": false                                              # optional, default false
        }
        Creates at most one exam per (class, name, year, term). If an exam already exists, it is skipped.
        """
        try:
            names = request.data.get('names')
            term = request.data.get('term')
            year = request.data.get('year')
            total_marks = request.data.get('total_marks', 100)
            date_all = request.data.get('date')
            dates = request.data.get('dates')
            grade = request.data.get('grade')
            publish = bool(request.data.get('publish', False))

            if not isinstance(names, list) or len(names) == 0:
                return Response({'detail': 'names must be a non-empty array'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                term = int(term)
                year = int(year)
            except Exception:
                return Response({'detail': 'term and year are required integers'}, status=status.HTTP_400_BAD_REQUEST)
            if term not in (1, 2, 3):
                return Response({'detail': 'term must be 1, 2 or 3'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                total_marks = int(total_marks)
            except Exception:
                return Response({'detail': 'total_marks must be an integer'}, status=status.HTTP_400_BAD_REQUEST)

            # Resolve school scope
            school = getattr(getattr(request, 'user', None), 'school', None)
            if not school:
                return Response({'detail': 'Your user is not linked to a school'}, status=status.HTTP_400_BAD_REQUEST)

            # Build class queryset in scope
            cls_qs = Class.objects.filter(school=school)
            if grade:
                try:
                    from .models import Class as ClassModel
                    formatted = ClassModel.format_grade_level(grade)
                    cls_qs = cls_qs.filter(grade_level=formatted)
                except Exception:
                    cls_qs = cls_qs.filter(grade_level=grade)

            # Validate dates array if provided
            use_dates = None
            if isinstance(dates, list) and len(dates) == len(names):
                use_dates = dates
            else:
                if dates is not None:
                    return Response({'detail': 'dates length must match names length'}, status=status.HTTP_400_BAD_REQUEST)

            # Default date fallback: today
            default_date = None
            if date_all:
                default_date = str(date_all)
            else:
                try:
                    default_date = date.today().isoformat()
                except Exception:
                    default_date = None

            created = []
            skipped = []
            errors = []
            with transaction.atomic():
                for klass in cls_qs.select_related('stream'):
                    for idx, nm in enumerate(names):
                        nm_str = str(nm).strip()
                        if not nm_str:
                            continue
                        # Determine date to use
                        dval = None
                        if use_dates:
                            dval = use_dates[idx]
                        else:
                            dval = default_date
                        # Check if exam exists
                        exists = Exam.objects.filter(name__iexact=nm_str, year=year, term=term, klass=klass).first()
                        if exists:
                            skipped.append({'klass': getattr(klass, 'name', klass.id), 'name': exists.name, 'id': exists.id})
                            continue
                        try:
                            e = Exam(name=nm_str, year=year, term=term, klass=klass, date=dval, total_marks=total_marks)
                            e.save()
                            if publish:
                                e.published = True
                                e.published_at = timezone.now()
                                e.save(update_fields=['published', 'published_at', 'name'])
                            created.append({'id': e.id, 'klass': getattr(klass, 'name', klass.id), 'name': e.name, 'date': e.date})
                        except Exception as ex:
                            errors.append({'klass': getattr(klass, 'name', klass.id), 'name': nm_str, 'error': str(ex)})

            return Response({'created': created, 'skipped': skipped, 'errors': errors}, status=(status.HTTP_201_CREATED if created and not errors else status.HTTP_207_MULTI_STATUS))
        except Exception as e:
            return Response({'detail': 'bulk create failed', 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin], url_path='bulk-assign')
    def bulk_assign(self, request):
        """Admin-only: Assign a list of students to a class.
        Body: { "klass": <class_id>, "student_ids": [1,2,3] }
        Sets is_graduated=False and student.school to klass.school.
        Useful for repairing unassigned students after promotion.
        """
        try:
            klass_id = request.data.get('klass')
            student_ids = request.data.get('student_ids') or []
            if not klass_id or not isinstance(student_ids, (list, tuple)):
                return Response({'detail': 'klass and student_ids are required'}, status=400)
            klass = Class.objects.filter(id=klass_id).first()
            if not klass:
                return Response({'detail': 'Class not found'}, status=404)
            updated = 0
            for sid in student_ids:
                stu = Student.objects.filter(id=sid).select_related('klass').first()
                if not stu:
                    continue
                prev = getattr(stu, 'klass', None)
                # Record history if model exists
                try:
                    from .models import StudentClassHistory
                    StudentClassHistory.objects.create(
                        student=stu,
                        from_class=prev,
                        to_class=klass,
                        action='assigned',
                        note='Bulk assign',
                    )
                except Exception:
                    pass
                stu.klass = klass
                stu.is_graduated = False
                stu.school = klass.school
                stu.save(update_fields=['klass','is_graduated','school'])
                updated += 1
            return Response({'detail': 'Students assigned', 'updated': updated})
        except Exception as e:
            return Response({'detail': 'Bulk assign failed', 'error': str(e)}, status=500)

    def perform_create(self, serializer):
        # Only admins can create exams
        if not self._is_admin(self.request):
            raise ValidationError({'detail': 'Only admins can create exams'})
        # enforce school scoping by validating klass belongs to user's school
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        klass = serializer.validated_data.get('klass')
        if school and klass and klass.school_id != school.id:
            raise ValidationError({'klass': 'Class must belong to your school'})
        serializer.save()

    # Block non-admin updates/deletes
    def update(self, request, *args, **kwargs):
        if not self._is_admin(request):
            return Response({'detail': 'Only admins can modify exams'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        exam = self.get_object()
        if not self._can_manage_exam(request, exam):
            return Response({'detail': 'You do not have permission to modify this exam'}, status=status.HTTP_403_FORBIDDEN)
        # If teacher changes klass, enforce same-school scoping
        school = getattr(getattr(request, 'user', None), 'school', None)
        new_klass_id = request.data.get('klass')
        if new_klass_id is not None and school:
            k = Class.objects.filter(id=new_klass_id).first()
            if k and getattr(k, 'school_id', None) != getattr(school, 'id', None) and not self._is_admin(request):
                return Response({'klass': ['Class must belong to your school']}, status=status.HTTP_400_BAD_REQUEST)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        exam = self.get_object()
        if not self._can_manage_exam(request, exam):
            return Response({'detail': 'You do not have permission to delete this exam'}, status=status.HTTP_403_FORBIDDEN)
        # Non-admins cannot delete published exams
        if (getattr(exam, 'published', False) or str(getattr(exam, 'status', '')).lower() == 'published') and not self._is_admin(request):
            return Response({'detail': 'Cannot delete a published exam'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)

    def _build_summary(self, exam):
        # Prefer examinable subjects; if none flagged, fallback to all subjects
        subj_qs_ex = exam.klass.subjects.filter(is_examinable=True)
        use_examinable_only = subj_qs_ex.exists()
        class_subjects = list(
            (subj_qs_ex if use_examinable_only else exam.klass.subjects.all()).values('id', 'code', 'name')
        )
        # Aggregate results; when no subjects are flagged examinable, do not filter them out
        base_res = ExamResult.objects.filter(exam=exam, student__is_active=True)
        if use_examinable_only:
            base_res = base_res.filter(subject__is_examinable=True)
        res = base_res.select_related('student', 'subject', 'component')
        students_map = {}
        for r in res:
            # Skip rows with missing/non-numeric marks to avoid TypeErrors
            try:
                mval = float(r.marks)
            except Exception:
                continue
            s = r.student
            entry = students_map.setdefault(s.id, {
                'id': s.id,
                'name': getattr(s, 'name', str(s)),
                'total': 0.0,
                'count': 0,
                'marks': {},
                # Track component-level percentages to compute subject percentage as average of components
                'subject_percent_parts': {},
                # Also track weighted accumulator per subject: sum(marks) and sum(denominators)
                'subject_pct_acc': {},
            })
            # Compute component percentage part
            # Determine denominator: prefer explicit out_of saved with the result,
            # then component.max_marks, then exam.total_marks, else 100
            sid = str(r.subject_id)
            denom = None
            try:
                if getattr(r, 'out_of', None):
                    denom = float(r.out_of)
            except Exception:
                denom = None
            if not denom:
                comp = getattr(r, 'component', None)
                if comp and getattr(comp, 'max_marks', None) is not None:
                    denom = float(comp.max_marks)
            if not denom:
                if getattr(r.exam, 'total_marks', None) is not None:
                    denom = float(r.exam.total_marks)
            if not denom:
                denom = 100.0
            if denom and denom > 0:
                # Backward-compat: older rows may have stored percent-like marks (0..100) even when denom is smaller
                if mval <= 100.0 and mval > denom:
                    part_pct = mval
                    norm_marks = (mval / 100.0) * denom
                else:
                    part_pct = (mval / denom) * 100.0
                    norm_marks = mval
                parts = entry['subject_percent_parts'].setdefault(sid, [])
                parts.append(part_pct)
                # Weighted accumulator for precise subject percentage across components
                acc = entry['subject_pct_acc'].setdefault(sid, {'num': 0.0, 'den': 0.0})
                acc['num'] += norm_marks
                acc['den'] += denom
                # Aggregate per subject: sum component marks under the same subject
                prev = entry['marks'].get(sid, 0.0)
                entry['marks'][sid] = prev + norm_marks
                entry['total'] += norm_marks
                entry['count'] += 1
        students = []
        for sid, e in students_map.items():
            # Build subject percentage map by averaging component percentages for that subject
            subj_pct_map = {}
            for sub_id, parts in e['subject_percent_parts'].items():
                if parts:
                    # Prefer weighted percentage across components when we have denominators
                    acc = (e.get('subject_pct_acc') or {}).get(sub_id)
                    if acc and acc.get('den'):
                        try:
                            subj_pct_map[sub_id] = round((float(acc['num']) / float(acc['den'])) * 100.0, 2)
                        except Exception:
                            subj_pct_map[sub_id] = round(sum(parts) / len(parts), 2)
                    else:
                        subj_pct_map[sub_id] = round(sum(parts) / len(parts), 2)

            # Fill missing subjects with 0 so blanks are treated as zeroes
            try:
                all_subj_ids = [str(s['id']) for s in class_subjects]
            except Exception:
                all_subj_ids = list(subj_pct_map.keys())
            for sub_id in all_subj_ids:
                if str(sub_id) not in subj_pct_map:
                    subj_pct_map[str(sub_id)] = 0.0

            # Totals/averages to MATCH the Results page: use subject percentages across ALL class subjects
            pct_values = [subj_pct_map.get(str(i), 0.0) for i in all_subj_ids]
            pct_sum = sum(pct_values) if pct_values else 0.0
            pct_cnt = len(all_subj_ids)
            pct_avg = (pct_sum / pct_cnt) if pct_cnt else 0.0

            # Fallbacks in case percentages are unavailable for legacy data
            raw_avg = (e['total'] / e['count']) if e['count'] else 0.0
            total_field = round(pct_sum if pct_cnt else e['total'], 2)
            average_field = round(pct_avg if pct_cnt else raw_avg, 2)

            students.append({
                'id': e['id'],
                'name': e['name'],
                'total': total_field,
                'average': average_field,
                'marks': e['marks'],
                'subject_percentages': subj_pct_map,
            })
        # sort by total desc (which is based on percent-sum for consistency with Results page)
        students.sort(key=lambda x: x['total'], reverse=True)
        # assign positions
        position = 1
        last_total = None
        same_rank_count = 0
        for idx, st in enumerate(students):
            if last_total is None or st['total'] < last_total:
                position = idx + 1
                last_total = st['total']
                same_rank_count = 1
            else:
                same_rank_count += 1
            st['position'] = position
        # class mean (average of student averages, using percentage-based averages)
        class_mean = round(sum(s['average'] for s in students) / len(students), 2) if students else 0.0
        # subject means (by marks) and mean percentages (by averaging student subject percentages)
        subj_means = []
        subj_mean_percentages = []
        subj_ids = [s['id'] for s in class_subjects]
        for sid in subj_ids:
            vals = [st['marks'].get(str(sid)) for st in students if st['marks'].get(str(sid)) is not None]
            mean = round(sum(vals)/len(vals), 2) if vals else 0.0
            subj_means.append({'subject': sid, 'mean': mean})
            # Mean subject percentage across students
            pcts = [st.get('subject_percentages', {}).get(str(sid)) for st in students if st.get('subject_percentages', {}).get(str(sid)) is not None]
            mean_pct = round(sum(pcts)/len(pcts), 2) if pcts else 0.0
            subj_mean_percentages.append({'subject': sid, 'mean_percentage': mean_pct})
        return {
            'subjects': class_subjects,
            'students': students,
            'class_mean': class_mean,
            'subject_means': subj_means,
            'subject_mean_percentages': subj_mean_percentages,
        }

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='rank')
    def rank(self, request, pk=None):
        """Return the student's position in their class and in their grade for this exam.
        Query params: student=<id>
        Grade rank is computed across all classes in the same school with the same grade_level
        and exams that share the same (name, year, term).
        """
        exam = self.get_object()
        try:
            student_id = int(request.query_params.get('student'))
        except Exception:
            return Response({'detail': 'student query parameter is required'}, status=400)

        # Class position using local summary
        summary = self._build_summary(exam)
        class_list = summary.get('students', [])
        class_pos = None
        for st in class_list:
            if str(st.get('id')) == str(student_id):
                class_pos = st.get('position')
                break

        # Grade cohort: same school, same grade_level, same (name, year, term)
        school = getattr(getattr(request, 'user', None), 'school', None)
        grade_level = getattr(getattr(exam, 'klass', None), 'grade_level', None)
        same_grade_exams = Exam.objects.filter(
            name=exam.name,
            year=exam.year,
            term=exam.term,
            klass__grade_level=grade_level,
        )
        if school:
            same_grade_exams = same_grade_exams.filter(klass__school=school)

        # Build grade ranking using percentage-based totals from _build_summary to match Results/Report views
        totals = {}
        try:
            for ex in same_grade_exams:
                summary_ex = self._build_summary(ex)
                for st in summary_ex.get('students', []):
                    sid = st.get('id')
                    if sid is None:
                        continue
                    # Keep the best (should be unique per exam anyway). Use percent-total
                    totals[int(sid)] = float(st.get('total') or 0.0)
        except Exception:
            totals = {}

        # Sort by total desc and assign ranks with ties
        ordered = sorted(([sid, sc] for sid, sc in totals.items()), key=lambda x: x[1], reverse=True)
        grade_pos = None
        last_total = None
        position = 0
        for idx, (sid, sc) in enumerate(ordered, start=1):
            if last_total is None or sc < last_total:
                position = idx
                last_total = sc
            if str(sid) == str(student_id):
                grade_pos = position
                break

        return Response({
            'class': {'position': class_pos, 'size': len(class_list)},
            'grade': {'position': grade_pos, 'size': len(ordered)},
            'exam': {'id': exam.id, 'name': exam.name, 'year': exam.year, 'term': exam.term},
        })

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='compare')
    def compare_exams(self, request):
        """Compare two exams' analytics for the same class/grade.
        Query params: exam_a, exam_b
        Returns: { exam_a: summary, exam_b: summary, deltas: {class_mean_delta, subject_means: [{subject, mean_a, mean_b, delta}] } }
        Access rules are the same as for list/get: teachers restricted to their classes.
        """
        exam_a_id = request.query_params.get('exam_a')
        exam_b_id = request.query_params.get('exam_b')
        if not (exam_a_id and exam_b_id):
            return Response({'detail': 'Provide exam_a and exam_b query parameters'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam_a = self.get_queryset().get(pk=exam_a_id)
            exam_b = self.get_queryset().get(pk=exam_b_id)
        except Exam.DoesNotExist:
            return Response({'detail': 'One or both exams not found or not accessible'}, status=status.HTTP_404_NOT_FOUND)

        # Optional sanity: same class
        try:
            if getattr(exam_a, 'klass_id', None) != getattr(exam_b, 'klass_id', None):
                # Allow but note difference; many schools compare across classes of same grade
                pass
        except Exception:
            pass

        sum_a = self._build_summary(exam_a)
        sum_b = self._build_summary(exam_b)

        # Build subject means delta by subject id intersection
        subj_ids = set([s['id'] for s in sum_a['subjects']]) | set([s['id'] for s in sum_b['subjects']])
        def find_mean(summary, sid):
            for item in summary.get('subject_means', []):
                if str(item.get('subject')) == str(sid) or item.get('subject') == sid:
                    return item.get('mean')
            return None
        deltas = []
        for sid in subj_ids:
            ma = find_mean(sum_a, sid)
            mb = find_mean(sum_b, sid)
            if ma is None and mb is None:
                continue
            deltas.append({'subject': sid, 'mean_a': ma, 'mean_b': mb, 'delta': (None if (ma is None or mb is None) else round(mb - ma, 2))})

        class_mean_delta = None
        try:
            if sum_a.get('class_mean') is not None and sum_b.get('class_mean') is not None:
                class_mean_delta = round(sum_b['class_mean'] - sum_a['class_mean'], 2)
        except Exception:
            class_mean_delta = None

        return Response({
            'exam_a': {'id': getattr(exam_a, 'id', None), 'name': getattr(exam_a, 'name', None), 'year': getattr(exam_a, 'year', None), 'term': getattr(exam_a, 'term', None), 'summary': sum_a},
            'exam_b': {'id': getattr(exam_b, 'id', None), 'name': getattr(exam_b, 'name', None), 'year': getattr(exam_b, 'year', None), 'term': getattr(exam_b, 'term', None), 'summary': sum_b},
            'deltas': {
                'class_mean_delta': class_mean_delta,
                'subject_means': deltas,
            }
        })

    @action(detail=True, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='compare-subjects')
    def compare_subjects(self, request, pk=None):
        """Compare two subjects within a single exam.
        Query params: subject_a, subject_b
        Returns: { subject_a: {...}, subject_b: {...}, per_student: [{student, a_pct, b_pct, delta}], deltas: {mean_percentage_delta} }
        Uses percentage values computed in _build_summary per subject.
        """
        exam = self.get_object()
        subj_a = request.query_params.get('subject_a')
        subj_b = request.query_params.get('subject_b')
        if not (subj_a and subj_b):
            return Response({'detail': 'Provide subject_a and subject_b query parameters'}, status=status.HTTP_400_BAD_REQUEST)

        summary = self._build_summary(exam)
        # Helper to fetch mean percentage for a subject
        def mean_pct(sid):
            for item in summary.get('subject_mean_percentages', []):
                if str(item.get('subject')) == str(sid) or item.get('subject') == sid:
                    return item.get('mean_percentage')
            return None
        mean_a = mean_pct(subj_a)
        mean_b = mean_pct(subj_b)
        mean_delta = None if (mean_a is None or mean_b is None) else round(mean_b - mean_a, 2)

        # Build per-student comparison using subject_percentages map
        per_student = []
        for st in summary.get('students', []):
            a_pct = st.get('subject_percentages', {}).get(str(subj_a))
            b_pct = st.get('subject_percentages', {}).get(str(subj_b))
            delta = None if (a_pct is None or b_pct is None) else round(b_pct - a_pct, 2)
            per_student.append({'student_id': st.get('id'), 'student': st.get('name'), 'a_pct': a_pct, 'b_pct': b_pct, 'delta': delta})

        return Response({
            'exam': {'id': getattr(exam, 'id', None), 'name': getattr(exam, 'name', None), 'year': getattr(exam, 'year', None), 'term': getattr(exam, 'term', None)},
            'subject_a': {'id': subj_a, 'mean_percentage': mean_a},
            'subject_b': {'id': subj_b, 'mean_percentage': mean_b},
            'deltas': {'mean_percentage_delta': mean_delta},
            'per_student': per_student,
            'subjects': summary.get('subjects', []),
        })

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def summary(self, request, pk=None):
        """Return aggregated results for an exam.
        Access rules:
        - Admins/staff: always allowed.
        - Teachers: allowed if they are the class teacher or mapped via ClassSubjectTeacher for the exam's class,
          OR if the exam is published (read-only access for transparency).
        - Others: denied.
        """
        user = request.user

        # Fetch exam directly (do not rely on filtered get_queryset for teachers) then apply explicit access rules.
        try:
            exam = Exam.objects.select_related('klass').get(pk=pk)
        except Exam.DoesNotExist:
            return Response({'detail': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Enforce school scoping for non-admins
        if not self._is_admin(request):
            school = getattr(user, 'school', None)
            if school and getattr(exam.klass, 'school_id', None) not in (None, getattr(school, 'id', None)):
                return Response({'detail': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Admins always allowed
        if self._is_admin(request):
            data = self._build_summary(exam)
            data['exam'] = {'id': getattr(exam, 'id', None), 'name': getattr(exam, 'name', None), 'year': getattr(exam, 'year', None), 'term': getattr(exam, 'term', None), 'klass': getattr(exam, 'klass_id', None), 'total_marks': getattr(exam, 'total_marks', None)}
            return Response(data)

        # Teachers: allow if published or assigned to the class
        if getattr(user, 'role', None) == 'teacher':
            is_class_teacher = getattr(exam.klass, 'teacher_id', None) == getattr(user, 'id', None)
            is_subject_teacher = ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user).exists()
            is_published = bool(getattr(exam, 'published', False)) or str(getattr(exam, 'status', '')).lower() == 'published'
            if not (is_published or is_class_teacher or is_subject_teacher):
                return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
            data = self._build_summary(exam)
            data['exam'] = {'id': getattr(exam, 'id', None), 'name': getattr(exam, 'name', None), 'year': getattr(exam, 'year', None), 'term': getattr(exam, 'term', None), 'klass': getattr(exam, 'klass_id', None), 'total_marks': getattr(exam, 'total_marks', None)}
            return Response(data)

        return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='summary-csv')
    def summary_csv(self, request, pk=None):
        exam = self.get_object()
        user = request.user
        if not self._is_admin(request):
            if getattr(user, 'role', None) == 'teacher':
                is_class_teacher = getattr(exam.klass, 'teacher_id', None) == getattr(user, 'id', None)
                is_subject_teacher = ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user).exists()
                is_published = bool(getattr(exam, 'published', False)) or str(getattr(exam, 'status', '')).lower() == 'published'
                if not (is_published or is_class_teacher or is_subject_teacher):
                    return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
            else:
                return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
        school = getattr(request.user, 'school', None)
        data = self._build_summary(exam)
        # Build CSV
        import csv
        sio = StringIO()
        writer = csv.writer(sio)
        # Header section
        writer.writerow([school.name if school else '', exam.name, f"Year {exam.year}", f"Term {exam.term}"])
        if school and getattr(school, 'motto', ''):
            writer.writerow([school.motto])
        writer.writerow([])
        # Table header
        head = ['Position','Student'] + [(s.get('name') or s.get('code')) for s in data['subjects']] + ['Total','Average']
        writer.writerow(head)
        for st in data['students']:
            row = [st['position'], st['name']]
            for s in data['subjects']:
                row.append(st['marks'].get(str(s['id']), ''))
            row += [st['total'], st['average']]
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(['Class Mean', data['class_mean']])
        writer.writerow(['Subject Means'] + [f"{(s.get('name') or s.get('code'))}:{next((m['mean'] for m in data['subject_means'] if m['subject']==s['id']),0)}" for s in data['subjects']])
        csv_text = sio.getvalue()
        resp = HttpResponse(csv_text, content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="exam_{exam.id}_summary.csv"'
        return resp

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='report-card-pdf')
    def report_card_pdf(self, request, pk=None):
        """Generate a single student's report card PDF for this exam.
        Query params: student=<id>
        Permissions modeled after summary():
        - Admins/staff: allowed
        - Teachers: allowed if published or assigned to class/subject
        - Students: allowed only if published and requesting their own record
        """
        exam = self.get_object()
        try:
            student_id = int(request.query_params.get('student'))
        except Exception:
            return Response({'detail': 'student query parameter is required'}, status=400)

        user = request.user
        if not self._is_admin(request):
            if getattr(user, 'role', None) == 'teacher':
                is_class_teacher = getattr(exam.klass, 'teacher_id', None) == getattr(user, 'id', None)
                is_subject_teacher = ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user).exists()
                is_published = bool(getattr(exam, 'published', False)) or str(getattr(exam, 'status', '')).lower() == 'published'
                if not (is_published or is_class_teacher or is_subject_teacher):
                    return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
            elif getattr(user, 'role', None) == 'student':
                is_published = bool(getattr(exam, 'published', False)) or str(getattr(exam, 'status', '')).lower() == 'published'
                if not is_published:
                    return Response({'detail': 'Results not published yet'}, status=403)
                # ensure requested student is the logged-in user
                stu_user_id = Student.objects.filter(pk=student_id).values_list('user_id', flat=True).first()
                if not (stu_user_id and int(stu_user_id) == int(getattr(user, 'id', 0))):
                    return Response({'detail': 'Not allowed'}, status=403)
            else:
                return Response({'detail': 'You do not have permission to perform this action.'}, status=403)

        if not REPORTLAB_AVAILABLE:
            return Response({'detail': 'PDF generation library not installed. Please install reportlab.'}, status=500)

        data = self._build_summary(exam)
        subjects = data.get('subjects', [])
        st_row = next((s for s in data.get('students', []) if str(s.get('id')) == str(student_id)), None)
        if not st_row:
            return Response({'detail': 'No results found for this student in this exam'}, status=404)

        # Resolve stage-wide grading bands for this class (Primary 1-6, Junior 7-9)
        def _infer_stage(klass_obj):
            stg = getattr(klass_obj, 'stage', None)
            if stg:
                return stg
            try:
                import re
                gl = getattr(klass_obj, 'grade_level', '') or ''
                m = re.search(r"(\d{1,2})", str(gl))
                num = int(m.group(1)) if m else None
                if num is not None:
                    if 1 <= num <= 6:
                        return 'primary'
                    if 7 <= num <= 9:
                        return 'junior'
            except Exception:
                pass
            return None

        def _load_stage_bands(klass_obj):
            try:
                stage = _infer_stage(klass_obj)
                if not stage:
                    return []
                school = getattr(klass_obj, 'school', None)
                qs = StageGradingBand.objects.all()
                if school:
                    qs = qs.filter(school=school)
                qs = qs.filter(stage=stage)
                bands = list(qs.values('grade','min','max','order'))
                # Sort by min descending so higher grades match first
                bands.sort(key=lambda b: float(b.get('min') or -1), reverse=True)
                return bands
            except Exception:
                return []

        stage_bands = _load_stage_bands(getattr(exam, 'klass', None))

        def _letter_from_percentage(pct: float | int | None) -> str:
            # Use stage-wide bands when available; else default thresholds
            try:
                n = float(pct)
            except (TypeError, ValueError):
                return '-'
            try:
                if stage_bands:
                    for b in stage_bands:
                        try:
                            bmin = float(b.get('min')) if b.get('min') is not None else float('-inf')
                            bmax = float(b.get('max')) if b.get('max') is not None else float('inf')
                            if n >= bmin and n <= bmax:
                                return str(b.get('grade') or '-')
                        except Exception:
                            continue
            except Exception:
                pass
            # Fallback default scale
            if n >= 80: return 'A'
            if n >= 70: return 'B'
            if n >= 60: return 'C'
            if n >= 50: return 'D'
            return 'E'

        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.utils import ImageReader

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=18*mm, bottomMargin=14*mm)
        styles = getSampleStyleSheet()
        elements = []

        # Modern header band (simple swoosh using polygons)
        def _draw_header(canvas, doc):
            width, height = A4
            canvas.saveState()
            # light background band
            canvas.setFillColor(colors.Color(0.93, 0.96, 1))
            canvas.rect(0, height-60, width, 60, fill=1, stroke=0)
            # accent swoosh
            canvas.setFillColor(colors.Color(0.17, 0.53, 0.87))
            p = canvas.beginPath()
            p.moveTo(0, height)
            p.lineTo(0, height-50)
            p.lineTo(width*0.25, height-15)
            p.lineTo(width*0.5, height)
            p.close()
            canvas.drawPath(p, fill=1, stroke=0)
            canvas.restoreState()

        school = getattr(request.user, 'school', None)
        logo_path = None
        try:
            if school and getattr(school, 'logo', None) and getattr(school.logo, 'path', None):
                logo_path = school.logo.path  # Platypus Image expects a filename or file-like object
        except Exception:
            logo_path = None

        # Header: logo + name + motto + title
        head_cells = []
        if logo_path:
            head_cells.append([Image(logo_path, width=14*mm, height=14*mm), Paragraph(f"<b>{school.name}</b><br/><font size=9 color=grey>{getattr(school,'motto','') or ''}</font>", styles['Normal'])])
            head_tbl = Table(head_cells, colWidths=[16*mm, 150*mm])
            head_tbl.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('LEFTPADDING',(0,0),(-1,-1),0),
                ('RIGHTPADDING',(0,0),(-1,-1),0),
            ]))
            elements.append(head_tbl)
        else:
            elements.append(Paragraph(f"<b>{getattr(school,'name','')}</b>", styles['Title']))
            if getattr(school,'motto',''):
                elements.append(Paragraph(f"<font size=9 color=grey>{school.motto}</font>", styles['Normal']))
        elements.append(Spacer(1, 6))
        title = f"STUDENT REPORT CARD — {exam.name}"
        elements.append(Paragraph(f"<b>{title}</b>", styles['Heading3']))
        elements.append(Spacer(1, 6))

        stu = Student.objects.filter(pk=student_id).select_related('klass').first()
        # Summary block (modern layout): left = student info, right = metrics (positions, total, average)
        try:
            summary_local = self._build_summary(exam)
            class_students = summary_local.get('students', [])
            class_size = len(class_students)
            class_pos = next((s.get('position') for s in class_students if str(s.get('id')) == str(student_id)), None)

            # Grade rank across same school+grade_level and same exam name/year/term (percentage-based totals)
            grade_level = getattr(getattr(exam, 'klass', None), 'grade_level', None)
            same_grade_exams = Exam.objects.filter(
                name=exam.name,
                year=exam.year,
                term=exam.term,
                klass__grade_level=grade_level,
            )
            school_scope = getattr(getattr(self.request, 'user', None), 'school', None)
            if school_scope:
                same_grade_exams = same_grade_exams.filter(klass__school=school_scope)

            totals_map = {}
            try:
                for ex in same_grade_exams:
                    sum_ex = self._build_summary(ex)
                    for row in sum_ex.get('students', []):
                        sid2 = row.get('id')
                        if sid2 is None:
                            continue
                        totals_map[int(sid2)] = float(row.get('total') or 0.0)
            except Exception:
                totals_map = {}

            ordered = sorted(totals_map.items(), key=lambda x: x[1], reverse=True)
            grade_size = len(ordered)
            grade_pos = None
            last_total = None
            position = 0
            for idx, (sid, sc) in enumerate(ordered, start=1):
                if last_total is None or sc < last_total:
                    position = idx
                    last_total = sc
                if str(sid) == str(student_id):
                    grade_pos = position
                    break

            # Left: student info
            left_tbl = Table([
                ['Student', getattr(stu, 'name', '')],
                ['Admission No', getattr(stu, 'admission_no','')],
                ['Class', getattr(getattr(stu,'klass',None),'name','')],
            ], colWidths=[35*mm, 65*mm])
            left_tbl.setStyle(TableStyle([
                ('GRID',(0,0),(-1,-1),0.3, colors.lightgrey),
                ('BACKGROUND',(0,0),(0,-1), colors.whitesmoke),
                ('TEXTCOLOR',(0,0),(0,-1), colors.grey),
                ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ]))

            # Right: metrics
            right_tbl = Table([
                ['Class Position', f"{class_pos if class_pos is not None else '-'} / {class_size}"],
                ['Grade Position', f"{grade_pos if grade_pos is not None else '-'} / {grade_size}"],
                ['Total', f"{st_row.get('total', 0)}"],
                ['Average', f"{st_row.get('average', 0)}"],
            ], colWidths=[30*mm, 30*mm])
            right_tbl.setStyle(TableStyle([
                ('GRID',(0,0),(-1,-1),0.3, colors.lightgrey),
                ('BACKGROUND',(0,0),(0,-1), colors.whitesmoke),
                ('TEXTCOLOR',(0,0),(0,-1), colors.grey),
                ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
                ('ALIGN',(1,0),(1,-1),'RIGHT'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ]))

            # Parent: two columns
            summary_tbl = Table([[left_tbl, right_tbl]], colWidths=[100*mm, 60*mm])
            summary_tbl.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'TOP'),
                ('LEFTPADDING',(0,0),(-1,-1),0),
                ('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),
                ('BOTTOMPADDING',(0,0),(-1,-1),0),
            ]))
            elements.append(summary_tbl)
            elements.append(Spacer(1, 8))
        except Exception:
            # If anything fails in summary computation, continue without blocking PDF
            elements.append(Spacer(1, 4))

        # Teacher + Remarks block
        try:
            teacher_name = ''
            try:
                t = getattr(getattr(exam, 'klass', None), 'teacher', None)
                if t:
                    teacher_name = (getattr(t, 'first_name', '') + ' ' + getattr(t, 'last_name', '')).strip() or getattr(t, 'username', '')
            except Exception:
                teacher_name = ''
            avg_val = float(st_row.get('average') or 0)
            # Simple remark rubric if custom bands absent
            if avg_val >= 80:
                remark = 'Excellent performance — keep it up.'
            elif avg_val >= 70:
                remark = 'Very good work.'
            elif avg_val >= 60:
                remark = 'Good, aim higher.'
            elif avg_val >= 50:
                remark = 'Fair — effort needed.'
            else:
                remark = 'Needs improvement — consult your teacher.'
            tr_tbl = Table([
                ['Class Teacher', teacher_name or '-'],
                ['Remarks', remark],
            ], colWidths=[40*mm, 140*mm])
            tr_tbl.setStyle(TableStyle([
                ('GRID',(0,0),(-1,-1),0.3, colors.lightgrey),
                ('BACKGROUND',(0,0),(0,-1), colors.whitesmoke),
                ('TEXTCOLOR',(0,0),(0,-1), colors.grey),
                ('ALIGN',(0,0),(0,-1),'LEFT'),
            ]))
            elements.append(tr_tbl)
            elements.append(Spacer(1, 6))
        except Exception:
            pass

        # Per-subject performance table: Subject | Marks | Grade (using stage-wide bands when present)
        rows = [['Subject', 'Marks', 'Grade']]
        subject_percentages = st_row.get('subject_percentages') or {}
        for s in subjects:
            sid = str(s['id'])
            mark = st_row['marks'].get(sid)
            pct = subject_percentages.get(sid)
            grade = _letter_from_percentage(pct)
            rows.append([
                s.get('code') or s.get('name'),
                '' if mark is None else round(float(mark), 2),
                grade,
            ])
        # Totals row (overall marks and average percentage for the exam)
        total = st_row.get('total', 0)
        avg = st_row.get('average', 0)
        rows.append(['Total', total, ''])
        rows.append(['Average', avg, _letter_from_percentage(avg)])

        tbl = Table(rows, colWidths=[90*mm, 30*mm, 30*mm])
        tbl.setStyle(TableStyle([
            ('GRID',(0,0),(-1,-1),0.3, colors.lightgrey),
            ('BACKGROUND',(0,0),(-1,0), colors.whitesmoke),
            ('ALIGN',(1,1),(1,-1),'RIGHT'),
            ('ALIGN',(2,1),(2,-1),'CENTER'),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('BOTTOMPADDING',(0,0),(-1,0),6),
        ]))
        elements.append(tbl)

        # Signature lines section
        try:
            sign_tbl = Table([
                ['Class Teacher Signature', '', 'Principal Signature', '', "Parent's Signature", ''],
                ['', '', '', '', '', ''],
                ['Date', '', 'Date', '', 'Date', ''],
            ], colWidths=[35*mm, 25*mm, 35*mm, 25*mm, 35*mm, 25*mm])
            sign_tbl.setStyle(TableStyle([
                ('LINEABOVE',(1,1),(1,1),0.3, colors.grey),
                ('LINEABOVE',(3,1),(3,1),0.3, colors.grey),
                ('LINEABOVE',(5,1),(5,1),0.3, colors.grey),
                ('TOPPADDING',(0,1),(-1,1),10),
                ('BOTTOMPADDING',(0,1),(-1,1),2),
                ('TEXTCOLOR',(0,0),(-1,0), colors.grey),
            ]))
            elements.append(Spacer(1, 10))
            elements.append(sign_tbl)
        except Exception:
            pass

        def footer(canv, doc_):
            from reportlab.lib.pagesizes import A4 as PS
            canv.saveState()
            ts = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
            canv.setFont('Helvetica', 8)
            canv.drawString(12*mm, 10*mm, ts)
            page_num = canv.getPageNumber()
            txt = f"Page {page_num}"
            w = canv.stringWidth(txt, 'Helvetica', 8)
            canv.drawString(PS[0]-12*mm-w, 10*mm, txt)
            canv.restoreState()

        # Combine header band and footer so both render
        def draw_page(canv, doc_):
            try:
                _draw_header(canv, doc_)
            except Exception:
                pass
            footer(canv, doc_)

        doc.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)
        pdf = buffer.getvalue()
        buffer.close()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="exam_{exam.id}_student_{student_id}_report_card.pdf"'
        return resp

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='publish')
    def publish(self, request, pk=None):
        """Mark exam as published and kick off background notifications (email/SMS/PDF/chat).
        Returns immediately so the frontend can update status without waiting for delivery.
        """
        exam = self.get_object()
        if getattr(exam, 'published', False):
            return Response({'detail': 'Exam already published', 'published_at': getattr(exam, 'published_at', None)}, status=200)

        # Mark published first for immediate UI feedback
        exam.published = True
        exam.published_at = timezone.now()
        exam.save(update_fields=['published','published_at'])

        def _send_notifications(exam_id: int, actor_id: int | None):
            try:
                exam_local = Exam.objects.select_related('klass','klass__school').get(pk=exam_id)
                # Import here to avoid hard deps if communications app changes
                from communications.utils import send_email_with_attachment, send_email_safe, create_messages_for_users

                # Gather results grouped by student
                res = (
                    ExamResult.objects
                    .filter(exam=exam_local, subject__is_examinable=True)
                    .select_related('student','subject')
                )
                by_student = {}
                for r in res:
                    s = r.student
                    entry = by_student.setdefault(s.id, {
                        'student': s,
                        'marks': {},
                        'total': 0.0,
                        'count': 0,
                    })
                    entry['marks'][r.subject_id] = float(r.marks)
                    entry['total'] += float(r.marks)
                    entry['count'] += 1

                # Build a simple subject list for column order
                subjects = list(exam_local.klass.subjects.filter(is_examinable=True))

                # Send messages per student
                chat_user_ids = []
                # Collect in-app notifications for bulk insert
                notifications_bulk = []
                # Compute class-level positions from totals
                ordered_local = []
                for sid, data in by_student.items():
                    total_val = float(data['total'] or 0)
                    cnt = int(data['count'] or 1)
                    ordered_local.append({'student': data['student'], 'total': round(total_val,2), 'average': round(total_val/cnt, 2)})
                ordered_local.sort(key=lambda x: x['total'], reverse=True)
                last_total = None
                pos_counter = 0
                class_pos_map = {}
                for idx, row in enumerate(ordered_local):
                    if last_total is None or row['total'] < last_total:
                        pos_counter = idx + 1
                        last_total = row['total']
                    class_pos_map[getattr(row['student'], 'id', None)] = {'position': pos_counter, 'out_of': len(ordered_local)}

                # Compute grade-level positions across cohort exams (same grade/year/term/name)
                grade_positions = {}
                try:
                    grade_tag = getattr(exam_local, 'grade_level_tag', None) or getattr(exam_local.klass, 'grade_level', None)
                    cohort_exams = Exam.objects.filter(
                        klass__school_id=getattr(exam_local.klass, 'school_id', None),
                        grade_level_tag=grade_tag,
                        year=getattr(exam_local, 'year', None),
                        term=getattr(exam_local, 'term', None),
                        name=getattr(exam_local, 'name', None),
                    ).only('id')
                    if cohort_exams.exists():
                        cohort_results = (
                            ExamResult.objects
                            .filter(exam__in=cohort_exams, student__is_active=True, subject__is_examinable=True)
                            .select_related('student')
                        )
                        per_student_grade = {}
                        for r in cohort_results:
                            s = r.student
                            e = per_student_grade.setdefault(s.id, {'student': s, 'total': 0.0, 'count': 0})
                            e['total'] += float(r.marks)
                            e['count'] += 1
                        ordered_grade = []
                        for sid2, dat2 in per_student_grade.items():
                            cnt2 = dat2['count'] or 1
                            ordered_grade.append({'student': dat2['student'], 'total': round(dat2['total'], 2), 'avg': round(dat2['total']/cnt2, 2)})
                        ordered_grade.sort(key=lambda x: x['total'], reverse=True)
                        last = None
                        pos = 0
                        for idx, row in enumerate(ordered_grade):
                            if last is None or row['total'] < last:
                                pos = idx + 1
                                last = row['total']
                            grade_positions[getattr(row['student'], 'id', None)] = {'position': pos, 'out_of': len(ordered_grade)}
                except Exception:
                    grade_positions = {}

                for sid, data in by_student.items():
                    s = data['student']
                    total = data['total']
                    avg = round(total / data['count'], 2) if data['count'] else 0.0
                    # Build dashboard URL for students
                    try:
                        frontend_base = getattr(settings, 'FRONTEND_URL', 'http://localhost:5173')
                    except Exception:
                        frontend_base = 'http://localhost:5173'
                    dashboard_url = f"{frontend_base.rstrip('/')}/student"

                    # SMS with per-subject marks
                    subject_parts = []
                    try:
                        for subj in subjects:
                            code_or_name = getattr(subj, 'code', None) or getattr(subj, 'name', '')
                            mark = data['marks'].get(subj.id)
                            if mark is not None:
                                subject_parts.append(f"{code_or_name}:{round(float(mark),2)}")
                    except Exception:
                        subject_parts = []
                    subj_summary = ", ".join(subject_parts) if subject_parts else ""
                    # Compose richer SMS with student identifiers, class/grade and positions
                    name = getattr(s, 'name', '') or getattr(s, 'admission_no', '') or f"Student {getattr(s,'id','')}"
                    adm = getattr(s, 'admission_no', '') or ''
                    grade_label = getattr(exam_local.klass, 'grade_level', '') or ''
                    class_name = getattr(exam_local.klass, 'name', '') or grade_label
                    cls_pos = class_pos_map.get(getattr(s, 'id', None)) or {}
                    pos_text = ''
                    try:
                        cp = cls_pos.get('position')
                        co = cls_pos.get('out_of')
                        if cp and co:
                            pos_text = f" Class Pos {cp}/{co}"
                    except Exception:
                        pos_text = ''
                    try:
                        gp = (grade_positions.get(getattr(s,'id',None)) or {}).get('position')
                        go = (grade_positions.get(getattr(s,'id',None)) or {}).get('out_of')
                        if gp and go:
                            pos_text = (pos_text + '; ' if pos_text else '') + f"Grade Pos {gp}/{go}"
                    except Exception:
                        pass
                    meta = f"{exam_local.name}: {name}"
                    if adm:
                        meta += f" (ADM {adm})"
                    if grade_label or class_name:
                        meta += f" - {grade_label or class_name}"
                        if class_name and class_name != grade_label:
                            meta += f" {class_name}"
                    parts = []
                    if subj_summary:
                        parts.append(subj_summary)
                    parts.append(f"Total {round(total,2)}, Avg {avg}")
                    if pos_text:
                        parts.append(pos_text)
                    parts.append(f"Login: {dashboard_url}")
                    sms = f"{meta}. " + '. '.join(parts)

                    # Collect for chat mirror and in-app notifications
                    if getattr(s, 'user_id', None):
                        chat_user_ids.append(s.user_id)
                        try:
                            from communications.models import Notification
                            notifications_bulk.append(Notification(user_id=s.user_id, message=sms, type='in_app'))
                        except Exception:
                            pass

                    # Email with optional PDF attachment
                    recipient = getattr(s, 'email', None) or getattr(getattr(s, 'user', None), 'email', None)
                    # Email body mirrors SMS content with line breaks
                    body_lines = [
                        f"Dear {getattr(s,'name','Student')},",
                        "",
                        f"{exam_local.name} results:",
                        f"Student: {getattr(s,'name','')} (ADM {adm or '-'})",
                        f"Class: {class_name} · Grade: {grade_label}",
                        f"Subjects: {subj_summary}" if subj_summary else "Subjects: -",
                        f"Total: {round(total,2)} · Average: {avg}",
                    ]
                    cp = class_pos_map.get(getattr(s, 'id', None)) or {}
                    if cp.get('position') and cp.get('out_of'):
                        body_lines.append(f"Class Position: {cp['position']}/{cp['out_of']}")
                    gp = grade_positions.get(getattr(s,'id',None)) or {}
                    if gp.get('position') and gp.get('out_of'):
                        body_lines.append(f"Grade Position: {gp['position']}/{gp['out_of']}")
                    body_lines.extend([
                        "",
                        f"View your dashboard: {dashboard_url}",
                        "",
                        "Regards, School Administration",
                    ])
                    body = "\n".join(body_lines)

                    attachment_bytes = None
                    filename = f"results_{exam_local.id}_{s.id}.pdf"
                    if REPORTLAB_AVAILABLE:
                        try:
                            buf = BytesIO()
                            doc = SimpleDocTemplate(
                                buf,
                                pagesize=A4,
                                leftMargin=36,
                                rightMargin=36,
                                topMargin=48,
                                bottomMargin=36,
                            )
                            elements = []
                            styles = getSampleStyleSheet()
                            # Custom style tweaks
                            from reportlab.lib.styles import ParagraphStyle
                            brand = colors.HexColor('#111827')  # gray-900
                            accent = colors.HexColor('#2563eb')  # blue-600
                            subtle = colors.HexColor('#6b7280')  # gray-500

                            # Header with logo, school, and exam meta in two columns
                            header_rows = []
                            try:
                                school = getattr(exam_local.klass, 'school', None)
                                logo_img = None
                                if school and getattr(school, 'logo', None) and getattr(school.logo, 'path', None):
                                    logo_img = Image(school.logo.path, width=50, height=50)
                                title = Paragraph(
                                    f"<b>{(school.name if school else 'School')}</b>",
                                    ParagraphStyle('title', parent=styles['Title'], textColor=brand)
                                )
                                sub = Paragraph(
                                    f"{getattr(school,'motto','') or ''}",
                                    ParagraphStyle('sub', parent=styles['Normal'], textColor=subtle, fontSize=9)
                                )
                                meta = Paragraph(
                                    f"<b>{exam_local.name}</b> &nbsp; Year {exam_local.year} &nbsp; Term {exam_local.term} &nbsp; Class {exam_local.klass.name}",
                                    ParagraphStyle('meta', parent=styles['Normal'], textColor=brand, fontSize=10)
                                )
                                left = logo_img if logo_img else ''
                                right = [title, sub, meta]
                                header_rows.append([left, right])
                                hdr = Table(header_rows, colWidths=[60, doc.width-60])
                                hdr.setStyle(TableStyle([
                                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                                    ('LEFTPADDING',(0,0),(-1,-1),0),
                                    ('RIGHTPADDING',(0,0),(-1,-1),0),
                                ]))
                                elements.append(hdr)
                            except Exception:
                                elements.append(Paragraph(f"<b>{exam_local.name}</b>", styles['Title']))

                            elements.append(Spacer(1, 12))

                            # Student info strip
                            info = Table([
                                [
                                    Paragraph(f"<b>Student:</b> {getattr(s,'name','')}", styles['Normal']),
                                    Paragraph(f"<b>Adm No:</b> {getattr(s,'admission_no','-')}", styles['Normal']),
                                    Paragraph(f"<b>Date:</b> {timezone.localtime(timezone.now()).strftime('%Y-%m-%d')}", styles['Normal']),
                                ]
                            ], colWidths=[doc.width*0.45, doc.width*0.25, doc.width*0.30])
                            info.setStyle(TableStyle([
                                ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#f9fafb')),
                                ('BOX',(0,0),(-1,-1), 0.5, colors.HexColor('#e5e7eb')),
                                ('INNERGRID',(0,0),(-1,-1), 0.5, colors.HexColor('#e5e7eb')),
                                ('LEFTPADDING',(0,0),(-1,-1),6),
                                ('RIGHTPADDING',(0,0),(-1,-1),6),
                                ('TOPPADDING',(0,0),(-1,-1),6),
                                ('BOTTOMPADDING',(0,0),(-1,-1),6),
                            ]))
                            elements.append(info)

                            elements.append(Spacer(1, 12))

                            # Subjects table
                            rows = [["Subject", "Marks"]]
                            for subj in subjects:
                                rows.append([subj.code or subj.name, str(data['marks'].get(subj.id, ''))])
                            # Styled table
                            tbl = Table(rows, repeatRows=1, colWidths=[doc.width*0.7, doc.width*0.3])
                            tbl_style = [
                                ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#f3f4f6')),
                                ('TEXTCOLOR',(0,0),(-1,0), brand),
                                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                                ('ALIGN',(1,1),(-1,-1),'RIGHT'),
                                ('GRID',(0,0),(-1,-1), 0.25, colors.HexColor('#e5e7eb')),
                                ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#fbfdff')]),
                                ('LEFTPADDING',(0,0),(-1,-1),6),
                                ('RIGHTPADDING',(0,0),(-1,-1),6),
                                ('TOPPADDING',(0,0),(-1,-1),6),
                                ('BOTTOMPADDING',(0,0),(-1,-1),6),
                            ]
                            tbl.setStyle(TableStyle(tbl_style))
                            elements.append(tbl)

                            elements.append(Spacer(1, 10))

                            # Summary box
                            summary = Table([
                                [
                                    Paragraph('<b>Total</b>', styles['Normal']),
                                    Paragraph(f"{round(total,2)}", ParagraphStyle('num', parent=styles['Normal'], alignment=2, textColor=brand)),
                                ],
                                [
                                    Paragraph('<b>Average</b>', styles['Normal']),
                                    Paragraph(f"{avg}", ParagraphStyle('num2', parent=styles['Normal'], alignment=2, textColor=brand)),
                                ]
                            ], colWidths=[doc.width*0.7, doc.width*0.3])
                            summary.setStyle(TableStyle([
                                ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#f8fafc')),
                                ('BOX',(0,0),(-1,-1), 0.6, colors.HexColor('#cbd5e1')),
                                ('INNERGRID',(0,0),(-1,-1), 0.6, colors.HexColor('#e2e8f0')),
                                ('LEFTPADDING',(0,0),(-1,-1),8),
                                ('RIGHTPADDING',(0,0),(-1,-1),8),
                                ('TOPPADDING',(0,0),(-1,-1),6),
                                ('BOTTOMPADDING',(0,0),(-1,-1),6),
                            ]))
                            elements.append(summary)

                            elements.append(Spacer(1, 14))

                            # Footer note
                            foot = Paragraph(
                                f"Generated on {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')} · {(exam_local.klass.school.name if getattr(exam_local.klass,'school',None) else '')}",
                                ParagraphStyle('foot', parent=styles['Normal'], textColor=subtle, fontSize=8)
                            )
                            elements.append(foot)

                            doc.build(elements)
                            attachment_bytes = buf.getvalue()
                        except Exception:
                            attachment_bytes = None
                    # Send
                    try:
                        if recipient:
                            if attachment_bytes:
                                send_email_with_attachment(
                                    subject=f"{exam_local.name} Results",
                                    message=body,
                                    recipient=recipient,
                                    filename=filename,
                                    content=attachment_bytes,
                                    mimetype='application/pdf',
                                    school_id=getattr(getattr(exam_local, 'klass', None), 'school_id', None),
                                )
                            else:
                                send_email_safe(f"{exam_local.name} Results", body, recipient, school_id=getattr(getattr(exam_local, 'klass', None), 'school_id', None))
                    except Exception:
                        pass

                # Create in-app notifications in bulk (best-effort)
                try:
                    if notifications_bulk:
                        from communications.models import Notification as _Notif
                        _Notif.objects.bulk_create(notifications_bulk, ignore_conflicts=True)
                except Exception:
                    pass

                # Mirror to chat so students see it in Messages UI
                try:
                    if chat_user_ids:
                        body = f"Your exam results for {exam_local.name} (Year {exam_local.year}, Term {exam_local.term}) are now available."
                        create_messages_for_users(
                            school_id=getattr(exam_local.klass, 'school_id', None),
                            sender_id=actor_id,
                            body=body,
                            recipient_user_ids=chat_user_ids,
                            system_tag='results',
                            queue_delivery=False,
                        )
                except Exception:
                    pass
            except Exception:
                # Avoid crashing the thread
                import logging
                logging.getLogger(__name__).exception('Exam publish notifications failed for exam %s', exam_id)

        # Start background thread for notifications
        actor_id = getattr(request.user, 'id', None)
        t = threading.Thread(target=_send_notifications, args=(exam.id, actor_id), daemon=True)
        t.start()

        return Response({'detail': 'Published', 'published_at': exam.published_at})

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='summary-pdf')
    def summary_pdf(self, request, pk=None):
        exam = self.get_object()
        user = request.user
        if not self._is_admin(request):
            if getattr(user, 'role', None) == 'teacher':
                is_class_teacher = getattr(exam.klass, 'teacher_id', None) == getattr(user, 'id', None)
                is_subject_teacher = ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user).exists()
                is_published = bool(getattr(exam, 'published', False)) or str(getattr(exam, 'status', '')).lower() == 'published'
                if not (is_published or is_class_teacher or is_subject_teacher):
                    return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
            else:
                return Response({'detail': 'You do not have permission to perform this action.'}, status=status.HTTP_403_FORBIDDEN)
        school = getattr(request.user, 'school', None)
        data = self._build_summary(exam)
        if not REPORTLAB_AVAILABLE:
            return Response({'detail': 'PDF generation library not installed. Please install reportlab.'}, status=500)

        buffer = BytesIO()
        
        # Use landscape for exams with many subjects (>6)
        num_subjects = len(data['subjects'])
        if num_subjects > 6:
            from reportlab.lib.pagesizes import landscape
            pagesize = landscape(A4)
        else:
            pagesize = A4
        
        width, height = pagesize
        doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        elements = []
        styles = getSampleStyleSheet()

        # Header with logo and school name
        header_parts = []
        logo_img = None
        try:
            if school and getattr(school, 'logo', None) and getattr(school.logo, 'path', None):
                logo_img = Image(school.logo.path, width=50, height=50)
        except Exception:
            logo_img = None

        title_text = f"{school.name if school else 'School'} — {exam.name}"
        title_para = Paragraph(f"<b>{title_text}</b>", styles['Title'])
        motto_para = Paragraph(f"<font size=9>{getattr(school, 'motto', '') or ''}</font>", styles['Normal'])
        meta_para = Paragraph(f"<font size=9>Year: {exam.year} &nbsp;&nbsp; Term: {exam.term} &nbsp;&nbsp; Class: {exam.klass.name}</font>", styles['Normal'])

        # Build a two-column header row
        if logo_img:
            header_table = Table([[logo_img, [title_para, motto_para, meta_para]]], colWidths=[60, width-60-72])
        else:
            header_table = Table([[[title_para, motto_para, meta_para]]], colWidths=[width-72])
        header_table.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),0),
            ('RIGHTPADDING',(0,0),(-1,-1),0),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 12))

        # Build results table with calculated column widths
        table_head = ['Pos','Student'] + [(s.get('name') or s.get('code')) for s in data['subjects']] + ['Total','Avg']
        table_rows = [table_head]
        for st in data['students']:
            row = [st['position'], st['name']]
            for s in data['subjects']:
                row.append(st['marks'].get(str(s['id']), ''))
            row += [st['total'], st['average']]
            table_rows.append(row)

        # Calculate column widths to fit page
        available_width = width - 72  # Account for margins
        pos_width = 30  # Position column
        total_width = 40  # Total column
        avg_width = 40  # Average column
        
        # Remaining width for student name and subject columns
        remaining_width = available_width - pos_width - total_width - avg_width
        
        # Allocate widths
        if num_subjects > 0:
            # Student name gets 25% of remaining, subjects share the rest
            student_width = min(120, remaining_width * 0.25)
            subject_total_width = remaining_width - student_width
            subject_width = subject_total_width / num_subjects
        else:
            student_width = remaining_width
            subject_width = 0
        
        # Build column widths list
        col_widths = [pos_width, student_width]
        col_widths.extend([subject_width] * num_subjects)
        col_widths.extend([total_width, avg_width])
        
        # Adjust font size based on number of subjects
        if num_subjects > 10:
            font_size = 7
        elif num_subjects > 6:
            font_size = 8
        else:
            font_size = 9

        tbl = Table(table_rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR',(0,0),(-1,0), colors.HexColor('#111827')),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ALIGN',(0,0),(-1,0),'CENTER'),
            ('ALIGN',(1,1),(1,-1),'LEFT'),  # Student names left-aligned
            ('GRID',(0,0),(-1,-1), 0.25, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#fafafa')]),
            ('FONTSIZE',(0,0),(-1,-1), font_size),
            ('ALIGN',(2,1),(-1,-1),'CENTER'),  # Center all marks
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1), 3),
            ('RIGHTPADDING',(0,0),(-1,-1), 3),
            ('TOPPADDING',(0,0),(-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 10))

        # Class mean and subject means
        elements.append(Paragraph(f"<b>Class Mean:</b> {data['class_mean']}", styles['Normal']))
        subj_text = ' &nbsp; '.join([f"{(s.get('name') or s.get('code'))}: {next((m['mean'] for m in data['subject_means'] if m['subject']==s['id']),0)}" for s in data['subjects']])
        elements.append(Paragraph(f"<font size=8>{subj_text}</font>", styles['Normal']))

        # Footer with page numbers and timestamp
        def footer(canv, doc_):
            canv.saveState()
            page_num = canv.getPageNumber()
            ts = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
            footer_text_left = f"Generated: {ts}"
            footer_text_right = f"Page {page_num}"
            powered = "Powered by EDU-TRACK"
            canv.setFont('Helvetica', 8)
            # Left
            canv.drawString(36, 20, footer_text_left)
            # Right
            w = canv.stringWidth(footer_text_right, 'Helvetica', 8)
            canv.drawString(pagesize[0]-36-w, 20, footer_text_right)
            # Center tag
            pw = canv.stringWidth(powered, 'Helvetica', 8)
            canv.drawString((pagesize[0]-pw)/2, 20, powered)
            canv.restoreState()

        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        pdf = buffer.getvalue()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="exam_{exam.id}_summary.pdf"'
        return resp


class ExamResultViewSet(viewsets.ModelViewSet):
    queryset = ExamResult.objects.all()
    serializer_class = ExamResultSerializer
    # Students should be able to read their own published results; teachers/admins can access as scoped in get_queryset
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset().select_related('exam','student','subject')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(exam__klass__school=school)
        # optional filters (accept alternate parameter names for robustness)
        qp = self.request.query_params
        exam_id = qp.get('exam') or qp.get('exam_id')
        if exam_id:
            qs = qs.filter(exam_id=exam_id)
        student_id = qp.get('student') or qp.get('student_id')
        if student_id:
            qs = qs.filter(student_id=student_id)
        subject_id = qp.get('subject') or qp.get('subject_id')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        component_id = qp.get('component') or qp.get('component_id')
        if component_id:
            qs = qs.filter(component_id=component_id)
        # class/klass filters
        klass_id = qp.get('klass') or qp.get('class') or qp.get('klass_id') or qp.get('class_id')
        if klass_id:
            qs = qs.filter(exam__klass_id=klass_id)
        # Scope for non-admins
        user = getattr(self.request, 'user', None)
        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            # Teachers can read:
            #  - any published results school-wide, and
            #  - unpublished results only within classes they teach (class teacher or subject teacher mapping)
            if self.request.method in permissions.SAFE_METHODS:
                qs = qs.filter(
                    Q(exam__published=True) |
                    Q(exam__klass__subject_teachers__teacher=user)
                ).distinct()
            else:
                qs = qs.filter(Q(exam__klass__subject_teachers__teacher=user)).distinct()
        # If requester is a student (not staff), only show published exams and their own results
        if getattr(user, 'role', None) == 'student' and not (user.is_staff or user.is_superuser):
            qs = qs.filter(exam__published=True, student__user=user)
        return qs

    def perform_create(self, serializer):
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        exam = serializer.validated_data.get('exam')
        subject = serializer.validated_data.get('subject')
        component = serializer.validated_data.get('component')
        marks = serializer.validated_data.get('marks')
        out_of = serializer.validated_data.get('out_of')
        user = getattr(self.request, 'user', None)
        # Auto-assign sole component if subject has exactly one and none provided
        try:
            if subject is not None and component is None and subject.components.count() == 1:
                only_comp = subject.components.first()
                if only_comp:
                    serializer.validated_data['component'] = only_comp
                    component = only_comp
        except Exception:
            pass
        if school and exam and exam.klass.school_id != school.id:
            raise ValidationError({'exam': 'Exam must belong to your school'})
        # Component belongs to subject
        if component and subject and component.subject_id != subject.id:
            raise ValidationError({'component': 'Component does not belong to the selected subject'})
        # If the subject has components defined, require a component to be specified
        if subject is not None:
            try:
                if subject.components.exists() and component is None:
                    raise ValidationError({'component': 'This subject has components. Please select a component/paper to record marks for.'})
            except Exception:
                pass
        # Basic marks validation
        if exam and marks is not None:
            try:
                m = float(marks)
            except (TypeError, ValueError):
                raise ValidationError({'marks': 'Marks must be a number'})
            if m < 0:
                raise ValidationError({'marks': 'Marks cannot be negative'})
            # Teacher can provide out_of to scale raw marks to the component/exam scale
            # Determine target maximum for storage and validation
            target_max = None
            if component and getattr(component, 'max_marks', None) is not None:
                target_max = float(component.max_marks)
            elif component is not None and out_of is not None:
                # If component specified but no max_marks set, prefer provided out_of
                try:
                    target_max = float(out_of)
                except Exception:
                    target_max = None
            elif getattr(exam, 'total_marks', None) is not None:
                target_max = float(exam.total_marks)
            else:
                target_max = 100.0

            if out_of is not None:
                try:
                    oo = float(out_of)
                except (TypeError, ValueError):
                    raise ValidationError({'out_of': 'out_of must be a number'})
                if oo <= 0:
                    raise ValidationError({'out_of': 'out_of must be greater than 0'})
                if m > oo:
                    raise ValidationError({'marks': f'Marks cannot exceed out_of ({oo})'})
                # Store raw marks as entered; persist denominator for accurate percentage computation
                serializer.validated_data['marks'] = m
                try:
                    serializer.validated_data['out_of'] = float(oo)
                except Exception:
                    serializer.validated_data['out_of'] = None
            else:
                if target_max is not None and m > target_max:
                    raise ValidationError({'marks': f'Marks cannot exceed maximum ({target_max})'})
                # When no explicit out_of is provided, persist the effective denominator
                try:
                    serializer.validated_data['out_of'] = float(target_max)
                except Exception:
                    pass
        # If teacher, ensure they are allowed to submit for this class/subject
        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            allowed = False
            reason = ''
            if not allowed and exam and subject:
                # Exact subject assignment
                if ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user, subject=subject).exists():
                    allowed = True
                    reason = 'subject_teacher_exact'
            if not allowed:
                raise ValidationError({'detail': 'You are not assigned to this class/subject for this exam', 'code': 'not_assigned'})
        # Upsert to avoid unique_together conflicts (exam, student, subject)
        try:
            serializer.save()
        except IntegrityError:
            vd = serializer.validated_data
            with transaction.atomic():
                obj, _ = ExamResult.objects.update_or_create(
                    exam=vd['exam'],
                    student=vd['student'],
                    subject=vd['subject'],
                    component=vd.get('component'),
                    defaults={'marks': vd['marks'], 'out_of': vd.get('out_of')},
                )
            serializer.instance = obj

    def perform_update(self, serializer):
        user = getattr(self.request, 'user', None)
        instance = getattr(serializer, 'instance', None)
        exam = serializer.validated_data.get('exam') or getattr(instance, 'exam', None)
        subject = serializer.validated_data.get('subject') or getattr(instance, 'subject', None)
        component = serializer.validated_data.get('component') or getattr(instance, 'component', None)

        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school and exam and getattr(getattr(exam, 'klass', None), 'school_id', None) != school.id:
            raise ValidationError({'exam': 'Exam must belong to your school'})

        if component and subject and component.subject_id != subject.id:
            raise ValidationError({'component': 'Component does not belong to the selected subject'})

        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            allowed = False
            if not allowed and exam and subject:
                if ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user, subject=subject).exists():
                    allowed = True
            if not allowed:
                raise ValidationError({'detail': 'You are not assigned to this class/subject for this exam', 'code': 'not_assigned'})

        serializer.save()

    @action(detail=False, methods=['post'], permission_classes=[IsTeacherOrAdmin], url_path='bulk')
    def bulk_upsert(self, request):
        """Admin-only bulk upsert of exam results.
        Payload format:
        {
          "results": [
            {"exam": <id>, "student": <id>, "subject": <id>, "marks": <float>}, ...
          ]
        }
        """
        items = request.data.get('results')
        if not isinstance(items, list):
            return Response({'detail': 'results must be an array'}, status=400)
        successes = 0
        errors = []
        out_ids = []
        user = getattr(request, 'user', None)
        for idx, item in enumerate(items):
            if not isinstance(item, dict):
                errors.append({'index': idx, 'error': 'Invalid item'})
                continue
            # Parse IDs (accept either objects or IDs)
            try:
                exam_id = getattr(item.get('exam'), 'id', None) or int(item.get('exam'))
                student_id = getattr(item.get('student'), 'id', None) or int(item.get('student'))
                subject_id = getattr(item.get('subject'), 'id', None) or int(item.get('subject'))
                marks = item.get('marks')
                component_id = item.get('component')
                if component_id is not None:
                    component_id = getattr(component_id, 'id', None) or int(component_id)
                out_of = item.get('out_of')
            except Exception:
                errors.append({'index': idx, 'error': {'detail': 'Invalid identifiers in payload'}})
                continue
            # Fetch instances
            try:
                exam = Exam.objects.select_related('klass').get(pk=exam_id)
                student = Student.objects.get(pk=student_id)
                subject = Subject.objects.get(pk=subject_id)
                component = None
                if component_id is not None:
                    component = SubjectComponent.objects.get(pk=component_id)
                # Auto-assign sole component if subject has exactly one and none provided
                if component is None:
                    try:
                        if subject.components.count() == 1:
                            component = subject.components.first()
                    except Exception:
                        pass
            except Exam.DoesNotExist:
                errors.append({'index': idx, 'error': {'exam': 'Not found'}})
                continue
            except Student.DoesNotExist:
                errors.append({'index': idx, 'error': {'student': 'Not found'}})
                continue
            except Subject.DoesNotExist:
                errors.append({'index': idx, 'error': {'subject': 'Not found'}})
                continue
            except SubjectComponent.DoesNotExist:
                errors.append({'index': idx, 'error': {'component': 'Not found'}})
                continue
            # Basic scope + marks validations (similar to perform_create)
            school = getattr(getattr(request, 'user', None), 'school', None)
            if school and exam and exam.klass.school_id != school.id:
                errors.append({'index': idx, 'error': {'exam': 'Exam must belong to your school'}})
                continue
            # Component belongs to subject
            if component and component.subject_id != subject.id:
                errors.append({'index': idx, 'error': {'component': 'Component does not belong to the selected subject'}})
                continue
            # If the subject has components defined, require a component to be specified
            try:
                if subject.components.exists() and component is None:
                    errors.append({'index': idx, 'error': {'component': 'This subject has components. Please provide a component/paper for each mark.'}})
                    continue
            except Exception:
                pass
            # Block non-examinable subjects
            try:
                if hasattr(subject, 'is_examinable') and not bool(subject.is_examinable):
                    errors.append({'index': idx, 'error': {'subject': 'This subject is not examinable. Results cannot be recorded.'}})
                    continue
            except Exception:
                pass
            # Teacher permission: same rules as perform_create
            if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
                allowed = False
                if not allowed and exam and subject:
                    if ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user, subject=subject).exists():
                        allowed = True
                if not allowed:
                    errors.append({'index': idx, 'error': {'detail': 'You are not assigned to this class/subject for this exam'}})
                    continue
            # Allow out_of-only updates (no marks) to persist teacher-set denominators
            marks_missing = marks in (None, '', [])
            if marks_missing and out_of is not None:
                try:
                    oo = float(out_of)
                except (TypeError, ValueError):
                    errors.append({'index': idx, 'error': {'out_of': 'out_of must be a number'}})
                    continue
                if oo <= 0:
                    errors.append({'index': idx, 'error': {'out_of': 'out_of must be greater than 0'}})
                    continue
                try:
                    with transaction.atomic():
                        obj = ExamResult.objects.filter(
                            exam=exam,
                            student=student,
                            subject=subject,
                            component=component,
                        ).first()
                        if obj is None:
                            errors.append({'index': idx, 'error': {'detail': 'Cannot update out_of without existing marks for this student/subject/component'}})
                            continue
                        obj.out_of = float(oo)
                        obj.save(update_fields=['out_of'])
                    successes += 1
                    out_ids.append(obj.id)
                except Exception as ex:
                    errors.append({'index': idx, 'error': str(ex)})
                continue

            try:
                m = float(marks)
                if m < 0:
                    raise ValidationError({'marks': 'Marks cannot be negative'})
                # Determine target maximum
                target_max = None
                if component and getattr(component, 'max_marks', None) is not None:
                    target_max = float(component.max_marks)
                elif component is not None and out_of is not None:
                    target_max = float(out_of)
                elif getattr(exam, 'total_marks', None) is not None:
                    target_max = float(exam.total_marks)
                else:
                    target_max = 100.0

                if out_of is not None:
                    try:
                        oo = float(out_of)
                    except (TypeError, ValueError):
                        raise ValidationError({'out_of': 'out_of must be a number'})
                    if oo <= 0:
                        raise ValidationError({'out_of': 'out_of must be greater than 0'})
                    if m > oo:
                        raise ValidationError({'marks': f'Marks cannot exceed out_of ({oo})'})
                    # Store raw marks as entered; persist denominator for accurate percentage computation
                    out_of_to_store = float(oo)
                else:
                    if target_max is not None and m > target_max:
                        raise ValidationError({'marks': f'Marks cannot exceed maximum ({target_max})'})
                    out_of_to_store = float(target_max) if target_max is not None else None
            except ValidationError as ve:
                errors.append({'index': idx, 'error': ve.detail})
                continue
            except Exception:
                errors.append({'index': idx, 'error': {'marks': 'Marks must be a number'}})
                continue
            # Upsert (do not overwrite existing out_of unless explicitly provided)
            try:
                with transaction.atomic():
                    obj = ExamResult.objects.filter(
                        exam=exam,
                        student=student,
                        subject=subject,
                        component=component,
                    ).first()
                    if obj is None:
                        # Create: persist denominator so downstream aggregation/percentage is correct
                        obj = ExamResult.objects.create(
                            exam=exam,
                            student=student,
                            subject=subject,
                            component=component,
                            marks=m,
                            out_of=(out_of_to_store if out_of_to_store is not None else float(target_max or 100.0)),
                        )
                    else:
                        obj.marks = m
                        if out_of_to_store is not None:
                            obj.out_of = out_of_to_store
                        obj.save(update_fields=['marks'] + (['out_of'] if out_of_to_store is not None else []))
                successes += 1
                out_ids.append(obj.id)
            except Exception as ex:
                errors.append({'index': idx, 'error': str(ex)})
        status_code = 200 if not errors else 207  # 207 Multi-Status semantic
        return Response({'saved': successes, 'failed': len(errors), 'errors': errors, 'ids': out_ids}, status=status_code)

    @action(
        detail=False,
        methods=['post'],
        permission_classes=[IsTeacherOrAdmin],
        url_path='upload',
        parser_classes=[MultiPartParser, FormParser, JSONParser]
    )
    def upload_results(self, request):
        """Upload marks from a CSV/XLSX file or an image (OCR).
        Request (multipart/form-data or JSON for URLs):
        - file: uploaded file (csv, xlsx, xls, png, jpg, jpeg)
        - exam: exam ID (required)
        - subject: subject ID (required)
        - component: optional subject component ID
        - out_of: optional numeric total marks to scale from
        - commit: boolean; if false or omitted, returns a preview; if true, saves
        - column_map: optional JSON object mapping your headers to expected keys,
                       e.g. {"admission_no":"ADM", "name":"Student Name", "marks":"Score"}

        Matching logic per row (in order):
        1) student_id
        2) admission_no (exact)
        3) name (case-insensitive exact)
        """
        file = request.FILES.get('file')
        exam_id = request.data.get('exam')
        subject_id = request.data.get('subject')
        component_id = request.data.get('component')
        out_of = request.data.get('out_of')
        commit = str(request.data.get('commit', '')).lower() in ('1','true','yes','on')
        debug_flag = str(request.data.get('debug', '')).lower() in ('1','true','yes','on')
        column_map = request.data.get('column_map')
        # Parse column_map if sent as JSON string
        if isinstance(column_map, str):
            try:
                import json as _json
                column_map = _json.loads(column_map)
            except Exception:
                column_map = None

        # Validate core IDs
        try:
            exam = Exam.objects.select_related('klass').get(pk=int(exam_id))
            subject = Subject.objects.get(pk=int(subject_id))
            component = None
            if component_id not in (None, ''):
                component = SubjectComponent.objects.get(pk=int(component_id))
        except Exam.DoesNotExist:
            return Response({'detail': 'exam not found'}, status=404)
        except Subject.DoesNotExist:
            return Response({'detail': 'subject not found'}, status=404)
        except SubjectComponent.DoesNotExist:
            return Response({'detail': 'component not found'}, status=404)
        except (TypeError, ValueError):
            return Response({'detail': 'invalid identifiers'}, status=400)

        # School scope and component-subject consistency
        school = getattr(getattr(request, 'user', None), 'school', None)
        if school and exam.klass.school_id != getattr(school, 'id', None):
            return Response({'detail': 'Exam must belong to your school'}, status=403)
        if component and component.subject_id != subject.id:
            return Response({'detail': 'Component does not belong to the selected subject'}, status=400)
        # Auto-assign sole component if subject has exactly one and none provided
        try:
            if subject and not component and subject.components.count() == 1:
                component = subject.components.first()
        except Exception:
            pass
        # If the subject has components defined, require a component to be specified
        try:
            if subject and subject.components.exists() and not component:
                return Response({'detail': 'This subject has components. Please select a component/paper to upload marks for.'}, status=400)
        except Exception:
            pass
        # Block non-examinable subjects
        try:
            if hasattr(subject, 'is_examinable') and not bool(subject.is_examinable):
                return Response({'detail': 'This subject is not examinable. Results cannot be recorded.'}, status=400)
        except Exception:
            pass

        # Teacher permissions (same as bulk)
        user = getattr(request, 'user', None)
        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            allowed = False
            if not allowed and subject:
                if ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user, subject=subject).exists():
                    allowed = True
            if not allowed:
                return Response({'detail': 'You are not assigned to this class/subject for this exam'}, status=403)

        # Load roster for matching
        class_students = Student.objects.filter(klass=exam.klass)
        by_id = {s.id: s for s in class_students}
        by_adm = {str(getattr(s, 'admission_no', '')).strip(): s for s in class_students if getattr(s, 'admission_no', None)}
        # Normalized admission map to tolerate OCR quirks (O/0, I/1, dashes/spaces)
        def _norm_adm(x: str):
            if x is None:
                return None
            s = str(x).strip().upper()
            # Replace common OCR mistakes
            s = s.replace('O', '0')
            s = s.replace('I', '1').replace('L', '1')
            # Remove non-alphanumeric
            import re as _re
            s = _re.sub(r"[^A-Z0-9]", "", s)
            return s
        by_adm_norm = {}
        for k, stu in by_adm.items():
            nk = _norm_adm(k)
            if nk:
                by_adm_norm[nk] = stu
        def _norm_name(x):
            try:
                # lowercase, strip, collapse multiple spaces
                return ' '.join(str(x).strip().lower().split())
            except Exception:
                return None
        by_name = {_norm_name(getattr(s, 'name', '')): s for s in class_students if getattr(s, 'name', None)}

        # Helper to normalize columns
        def resolve_columns(header):
            hmap = {str(k).strip().lower(): k for k in header}
            def pick(*cands):
                for c in cands:
                    if c in hmap:
                        return hmap[c]
                return None
            if column_map and isinstance(column_map, dict):
                id_col = column_map.get('student_id')
                adm_col = column_map.get('admission_no')
                name_col = column_map.get('name')
                marks_col = column_map.get('marks')
            else:
                id_col = pick('student_id','id')
                adm_col = pick('admission_no','adm','adm_no','admission')
                name_col = pick('name','student','student_name')
                marks_col = pick('marks','score','points')
            return id_col, adm_col, name_col, marks_col

        # Parse incoming file (CSV/XLSX) or attempt OCR on images
        rows = []  # list of dicts {student_id?, admission_no?, name?, marks?}
        ocr_text = None
        raw_lines = []
        if file:
            fname = getattr(file, 'name', 'upload').lower()
            if fname.endswith('.csv'):
                # Robust CSV handling: try utf-8-sig, fall back to utf-8/latin1, sniff delimiter
                raw = file.read()
                try:
                    text = raw.decode('utf-8-sig')
                except Exception:
                    try:
                        text = raw.decode('utf-8')
                    except Exception:
                        text = raw.decode('latin1', errors='ignore')
                import csv as _csv
                try:
                    sniffer = _csv.Sniffer()
                    dialect = sniffer.sniff(text[:4096])
                    delim = dialect.delimiter
                except Exception:
                    # Common alternates if sniff fails
                    if '\t' in text and text.count('\t') > text.count(','):
                        delim = '\t'
                    elif ';' in text and text.count(';') > text.count(','):
                        delim = ';'
                    else:
                        delim = ','
                reader = _csv.DictReader(StringIO(text), delimiter=delim)
                id_col, adm_col, name_col, marks_col = resolve_columns(reader.fieldnames or [])
                for r in reader:
                    rows.append({
                        'student_id': r.get(id_col) if id_col else None,
                        'admission_no': r.get(adm_col) if adm_col else None,
                        'name': r.get(name_col) if name_col else None,
                        'marks': r.get(marks_col) if marks_col else None,
                    })
            elif fname.endswith('.xlsx') or fname.endswith('.xls'):
                try:
                    from openpyxl import load_workbook
                except Exception:
                    return Response({'detail': 'openpyxl is required to parse Excel files. Please install it on the server.'}, status=500)
                wb = load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                header = []
                data_started = False
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if i == 1:
                        header = [str(c).strip() if c is not None else '' for c in row]
                        id_col, adm_col, name_col, marks_col = resolve_columns(header)
                        data_started = True
                        continue
                    if not data_started:
                        continue
                    values = list(row)
                    def get(col_name):
                        if not col_name:
                            return None
                        idx = header.index(col_name)
                        return values[idx] if idx < len(values) else None
                    rows.append({
                        'student_id': get(id_col),
                        'admission_no': get(adm_col),
                        'name': get(name_col),
                        'marks': get(marks_col),
                    })
            elif any(fname.endswith(ext) for ext in ('.png','.jpg','.jpeg','.bmp','.webp','.tif','.tiff')):
                # OCR via pytesseract if available
                try:
                    from PIL import Image as _Image
                    import pytesseract as _pyt
                    # Try to auto-configure tesseract path on Windows via env var
                    import os as _os
                    tpath = _os.environ.get('TESSERACT_CMD') or _os.environ.get('TESSERACT_PATH')
                    if tpath:
                        try:
                            _pyt.pytesseract.tesseract_cmd = tpath
                        except Exception:
                            pass
                except Exception:
                    return Response({'detail': 'OCR not available. Install pytesseract and the Tesseract OCR binary to read images.'}, status=500)
                try:
                    img = _Image.open(file)
                    # Basic pre-processing to improve OCR
                    try:
                        from PIL import ImageOps as _ImageOps, ImageFilter as _ImageFilter, ImageEnhance as _ImageEnhance
                        img = img.convert('L')  # grayscale
                        img = _ImageOps.autocontrast(img)
                        # Slight upscale for small screenshots
                        try:
                            if img.width < 1600:
                                scale = 1600 / float(img.width)
                                img = img.resize((int(img.width*scale), int(img.height*scale)), _Image.LANCZOS)
                        except Exception:
                            pass
                        img = img.filter(_ImageFilter.SHARPEN)
                        img = _ImageEnhance.Contrast(img).enhance(1.2)
                    except Exception:
                        pass
                    # If tesseract is not in PATH and not configured above, this will raise an error
                    config = '--oem 3 --psm 6'
                    text = _pyt.image_to_string(img, lang='eng', config=config)
                    if not text or not text.strip():
                        # Fallback: reconstruct lines from word-level data
                        try:
                            from pytesseract import Output as _Output
                            data = _pyt.image_to_data(img, lang='eng', config=config, output_type=_Output.DICT)
                            lines = {}
                            for i in range(len(data.get('text', []))):
                                word = str(data['text'][i] or '').strip()
                                if not word:
                                    continue
                                key = (data.get('block_num',[0])[i], data.get('par_num',[0])[i], data.get('line_num',[0])[i])
                                lines.setdefault(key, []).append(word)
                            text = '\n'.join(' '.join(words) for _, words in sorted(lines.items()))
                        except Exception:
                            pass
                    ocr_text = text
                    import re as _re
                    # Strategy:
                    # - Split by lines
                    # - For each non-empty line, attempt several patterns
                    #   1) CSV/TSV-like: id,name,marks OR admission,name,marks
                    #   2) Free text with last number as marks: <anything name> <number>
                    #   3) Hyphen/colon separated: name - number, name: number
                    number_re = _re.compile(r"(?P<marks>\d+(?:\.\d+)?)\s*$")
                    # Common separators: commas, tabs, multiple spaces, dashes, colons
                    sep_re = _re.compile(r"[,\t]|\s{2,}|\s-\s|\s–\s|\s—\s|:\s")
                    for raw in text.splitlines():
                        l = str(raw or '').strip()
                        if not l:
                            continue
                        raw_lines.append(l)
                        # Try CSV/TSV or obvious separators first
                        parts = [p.strip() for p in sep_re.split(l) if str(p).strip()]
                        sid = None; adm = None; nm = None; mk = None
                        if len(parts) >= 2:
                            # If the last part is numeric, take it as marks
                            if parts[-1].replace('.', '', 1).isdigit():
                                mk = parts[-1]
                                # Reconstruct a plausible name from remaining parts, ignoring an initial numeric id
                                rest = parts[:-1]
                                if rest and str(rest[0]).isdigit():
                                    # first token is student_id
                                    sid = rest[0]
                                    # if the next token looks like an admission number, take it
                                    if len(rest) >= 2:
                                        # simple heuristic: contains letters+digits or a hyphen
                                        t1 = rest[1]
                                        if any(c.isalpha() for c in t1) and any(c.isdigit() for c in t1):
                                            adm = t1
                                            nm = ' '.join(rest[2:])
                                        else:
                                            nm = ' '.join(rest[1:])
                                    else:
                                        nm = ''
                                else:
                                    # if first token looks like an admission, keep it and name is the rest
                                    if rest and any(c.isalpha() for c in rest[0]) and any(c.isdigit() for c in rest[0]):
                                        adm = rest[0]
                                        nm = ' '.join(rest[1:])
                                    else:
                                        nm = ' '.join(rest)
                        # If still no marks, fallback: find trailing number with regex
                        if mk is None:
                            m = number_re.search(l)
                            if m:
                                mk = m.group('marks')
                                pre = l[:m.start()].strip()
                                # Tokenize on whitespace to extract id/admission/name
                                toks = [t for t in pre.split() if t]
                                if toks:
                                    # Case A: first token numeric -> student_id
                                    if toks[0].isdigit():
                                        sid = toks[0]
                                        rest = toks[1:]
                                        if rest:
                                            # If next token looks like an admission (letters+digits or contains '-')
                                            t1 = rest[0]
                                            if any(c.isalpha() for c in t1) and any(c.isdigit() for c in t1):
                                                adm = t1
                                                nm = ' '.join(rest[1:])
                                            else:
                                                nm = ' '.join(rest)
                                        else:
                                            nm = ''
                                    else:
                                        # Case B: start with admission then name
                                        if any(c.isalpha() for c in toks[0]) and any(c.isdigit() for c in toks[0]):
                                            adm = toks[0]
                                            nm = ' '.join(toks[1:])
                                        else:
                                            # Case C: just name before number
                                            nm = pre
                        # Append if we extracted at least a name or id and marks
                        if mk is not None and (nm or adm or sid):
                            rows.append({'student_id': sid, 'admission_no': adm, 'name': nm, 'marks': mk})
                except Exception as ex:
                    # Provide clearer guidance for missing binary
                    msg = str(ex)
                    help_hint = 'Tesseract is not installed or not on PATH. Install Tesseract OCR and, if needed, set env var TESSERACT_CMD to the tesseract.exe full path.'
                    return Response({'detail': f'OCR parse failed: {msg}', 'hint': help_hint}, status=400)
            else:
                return Response({'detail': 'Unsupported file type. Use CSV, XLSX/XLS, or an image.'}, status=400)
        else:
            return Response({'detail': 'file is required'}, status=400)

        # Validate marks and compute scaling
        def coerce_float(val):
            if val is None:
                return None
            try:
                return float(str(val).strip())
            except Exception:
                return None

        # Determine target maximum for validation
        target_max = None
        if component and getattr(component, 'max_marks', None) is not None:
            target_max = float(component.max_marks)
        elif component is not None and out_of not in (None, ''):
            # If component specified but has no max_marks, prefer provided out_of to avoid scaling to exam total
            try:
                target_max = float(out_of)
            except Exception:
                target_max = None
        elif getattr(exam, 'total_marks', None) is not None:
            target_max = float(exam.total_marks)
        else:
            target_max = 100.0

        # Parse out_of
        out_of_val = None
        if out_of not in (None, ''):
            try:
                out_of_val = float(out_of)
                if out_of_val <= 0:
                    return Response({'detail': 'out_of must be greater than 0'}, status=400)
            except Exception:
                return Response({'detail': 'out_of must be a number'}, status=400)

        preview = []
        to_save = []
        # Optional: fuzzy name matching helper
        try:
            from difflib import SequenceMatcher as _SeqMatch
        except Exception:
            _SeqMatch = None

        def _best_name_match(name_str):
            if not name_str:
                return None
            key = ' '.join(str(name_str).strip().lower().split())
            if key in by_name:
                return by_name[key]
            if _SeqMatch is None:
                return None
            # Fallback: pick best similarity
            best = None
            best_ratio = 0.0
            for k, stu in by_name.items():
                if not k:
                    continue
                ratio = _SeqMatch(None, key, k).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best = stu
            return best if best_ratio >= 0.8 else None

        for i, r in enumerate(rows):
            raw_marks = coerce_float(r.get('marks'))
            sid_raw = r.get('student_id')
            adm_raw = r.get('admission_no')
            name_raw = r.get('name')

            # Match student
            matched = None
            sid_int = None
            try:
                if sid_raw is not None and str(sid_raw).strip() != '':
                    sid_int = int(str(sid_raw).strip())
            except Exception:
                sid_int = None
            if sid_int and sid_int in by_id:
                matched = by_id[sid_int]
            elif adm_raw is not None:
                raw_key = str(adm_raw).strip()
                if raw_key in by_adm:
                    matched = by_adm[raw_key]
                else:
                    nk = _norm_adm(raw_key)
                    if nk in by_adm_norm:
                        matched = by_adm_norm[nk]
            elif name_raw is not None:
                # Try exact then fuzzy name match
                candidate = _best_name_match(name_raw)
                if candidate:
                    matched = candidate

            error = None
            scaled = None
            if raw_marks is None:
                error = {'marks': 'Marks must be a number'}
            elif raw_marks < 0:
                error = {'marks': 'Marks cannot be negative'}
            else:
                if out_of_val is not None:
                    if raw_marks > out_of_val:
                        error = {'marks': f'Marks cannot exceed out_of ({out_of_val})'}
                    else:
                        # Store raw marks as entered; keep out_of for percentage computation
                        scaled = raw_marks
                else:
                    if target_max is not None and raw_marks > target_max:
                        error = {'marks': f'Marks cannot exceed maximum ({target_max})'}
                    else:
                        scaled = raw_marks

            preview.append({
                'index': i,
                'student': getattr(matched, 'id', None),
                'student_name': getattr(matched, 'name', None) if matched else None,
                'input': {'student_id': sid_raw, 'admission_no': adm_raw, 'name': name_raw, 'marks': r.get('marks')},
                'scaled_marks': None if scaled is None else round(float(scaled), 2),
                'error': error if (error or matched is None) else None,
            })

            if matched and scaled is not None and not error:
                to_save.append((matched, float(scaled)))

        if not commit:
            # Return preview only
            resp = {
                'exam': exam.id,
                'subject': subject.id,
                'component': getattr(component, 'id', None),
                'target_max': target_max,
                'rows': preview,
                'would_save': len(to_save),
                'total_rows': len(rows),
            }
            if debug_flag and ocr_text is not None:
                resp['ocr_text'] = ocr_text
                resp['ocr_lines'] = raw_lines
            return Response(resp)

        # Commit upserts
        successes = 0
        errors = []
        saved_ids = []
        for idx, (stu, mval) in enumerate(to_save):
            try:
                with transaction.atomic():
                    obj = ExamResult.objects.filter(
                        exam=exam,
                        student=stu,
                        subject=subject,
                        component=component,
                    ).first()
                    if obj is None:
                        obj = ExamResult.objects.create(
                            exam=exam,
                            student=stu,
                            subject=subject,
                            component=component,
                            marks=mval,
                            out_of=(float(out_of_val) if out_of_val is not None else float(target_max or 100.0)),
                        )
                    else:
                        obj.marks = mval
                        # Only overwrite out_of if uploader explicitly supplied it
                        if out_of_val is not None:
                            obj.out_of = float(out_of_val)
                        elif getattr(obj, 'out_of', None) in (None, ''):
                            try:
                                obj.out_of = float(target_max or 100.0)
                            except Exception:
                                pass
                        obj.save(update_fields=['marks'] + (['out_of'] if out_of_val is not None else []))
                successes += 1
                saved_ids.append(obj.id)
            except Exception as ex:
                errors.append({'index': idx, 'error': str(ex)})

        status_code = 200 if not errors else 207
        return Response({'saved': successes, 'failed': len(errors), 'errors': errors, 'ids': saved_ids})

    @action(
        detail=False,
        methods=['get'],
        permission_classes=[IsTeacherOrAdmin],
        url_path='upload-template'
    )
    def upload_template(self, request):
        """Download a CSV template for an exam+subject with the class roster prefilled.
        Query params:
        - exam: exam ID (required)
        - subject: subject ID (required)
        - component: optional subject component ID (for information only)
        Columns: student_id,admission_no,name,marks
        """
        exam_id = request.query_params.get('exam')
        subject_id = request.query_params.get('subject')
        component_id = request.query_params.get('component')
        try:
            exam = Exam.objects.select_related('klass').get(pk=int(exam_id))
            subject = Subject.objects.get(pk=int(subject_id))
            component = None
            if component_id not in (None, '', 'null'):
                component = SubjectComponent.objects.get(pk=int(component_id))
        except (TypeError, ValueError):
            return Response({'detail': 'invalid identifiers'}, status=400)
        except Exam.DoesNotExist:
            return Response({'detail': 'exam not found'}, status=404)
        except Subject.DoesNotExist:
            return Response({'detail': 'subject not found'}, status=404)
        except SubjectComponent.DoesNotExist:
            return Response({'detail': 'component not found'}, status=404)

        # School scope and teacher permission
        school = getattr(getattr(request, 'user', None), 'school', None)
        if school and exam.klass.school_id != getattr(school, 'id', None):
            return Response({'detail': 'Exam must belong to your school'}, status=403)
        user = getattr(request, 'user', None)
        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            allowed = False
            if not allowed and subject:
                if ClassSubjectTeacher.objects.filter(klass=exam.klass, teacher=user, subject=subject).exists():
                    allowed = True
            if not allowed:
                return Response({'detail': 'You are not assigned to this class/subject for this exam'}, status=403)

        # Build CSV
        sio = StringIO()
        writer = csv.writer(sio)
        comps = list(SubjectComponent.objects.filter(subject=subject))
        if comps and component is None:
            # Multi-component template: one column per component
            comp_cols = []
            for c in comps:
                label = (getattr(c, 'code', None) or getattr(c, 'name', None) or f"comp_{c.id}")
                comp_cols.append((c.id, str(label)))
            header = ['student_id','admission_no','name'] + [lbl for _, lbl in comp_cols]
            writer.writerow(header)
            for s in Student.objects.filter(klass=exam.klass, is_active=True).order_by('name'):
                row = [s.id, getattr(s,'admission_no',''), getattr(s,'name','')]
                row += ['' for _ in comp_cols]
                writer.writerow(row)
        else:
            # Single-paper (whole subject or a specific component) template
            writer.writerow(['student_id','admission_no','name','marks'])
            for s in Student.objects.filter(klass=exam.klass, is_active=True).order_by('name'):
                writer.writerow([s.id, getattr(s,'admission_no',''), getattr(s,'name',''), ''])
        csv_text = sio.getvalue()
        resp = HttpResponse(csv_text, content_type='text/csv; charset=utf-8')
        subj_code = getattr(subject, 'code', subject.id)
        comp_tag = f"_comp{component.id}" if component else ("_all_components" if comps and component is None else '')
        resp['Content-Disposition'] = f'attachment; filename="upload_template_exam{exam.id}_subj{subj_code}{comp_tag}.csv"'
        return resp

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    filter_backends = [DjangoFilterBackend]

    # Allow server-side filtering by class, gender, active/graduation status, and year
    filterset_fields = ['klass', 'gender', 'is_graduated', 'graduation_year', 'is_active']
    # Support JSON (default axios), form, and multipart (for photo uploads)
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_permissions(self):
        """Permissions matrix for Student endpoints.
        - my, my_update: any authenticated user (student portal)
        - list/retrieve and other safe reads: any authenticated user (client UI may further hide)
        - mutations: admin only
        """
        act = getattr(self, 'action', None)
        if act in ('my', 'my_update'):
            return [permissions.IsAuthenticated()]
        # Allow the specialized teacher_update action for class teachers (enforced inside the action)
        if act in ('teacher_update',):
            return [IsTeacherOrAdmin()]
        if act in ('list', 'retrieve') or self.request.method in permissions.SAFE_METHODS:
            # Broaden to any authenticated to avoid role mismatches breaking list in deploy
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        qs = super().get_queryset()
        user = getattr(self.request, 'user', None)
        school = getattr(user, 'school', None)
        if school:
            # Include:
            # - students currently in classes of this school
            # - students explicitly scoped to this school (e.g., graduated)
            # - unassigned, not-graduated students with no school set yet (common on import)
            qs = qs.filter(
                Q(klass__school=school) |
                Q(school=school) |
                Q(klass__isnull=True, school__isnull=True, is_graduated=False)
            )
        # Support multiple alias query params for filtering by class id
        try:
            qp = self.request.query_params
            class_param = None
            for k in ('class', 'klass', 'class_id', 'classId'):
                v = qp.get(k)
                if v not in (None, ''):
                    class_param = v
                    break
            if class_param not in (None, ''):
                qs = qs.filter(klass_id=int(class_param))
        except Exception:
            pass
        # Optional grade filter via related class grade_level
        grade = self.request.query_params.get('grade')
        if grade:
            qs = qs.filter(klass__grade_level=grade)
        # Optional: by default, include both active and inactive. Frontend can pass is_active to filter.
        # Text search by name or admission number (case-insensitive)
        q = self.request.query_params.get('q')
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(admission_no__icontains=q))

        # Optimize field loading depending on action
        act = getattr(self, 'action', None)
        if act == 'list':
            # Only fields needed by StudentListSerializer
            qs = (
                qs.select_related('klass')
                  .only(
                      'id','admission_no','name','dob','gender','upi_number','guardian_id','klass','is_graduated','graduation_year','photo',
                      'klass__id','klass__name','klass__grade_level'
                  )
            )
        else:
            # For detailed views/updates, include related klass
            qs = qs.select_related('klass')
        return qs

    def _build_remote_photo_kwargs(self, serializer):
        request = getattr(self, 'request', None)
        if not request:
            return {}
        # Skip when file upload already provided/processed
        try:
            if serializer.validated_data.get('photo'):
                return {}
        except Exception:
            pass
        try:
            files = getattr(request, 'FILES', None)
            if files and files.get('photo'):
                return {}
        except Exception:
            pass

        photo_url = None
        try:
            data = getattr(request, 'data', {})
            photo_url = data.get('photo_url') or data.get('avatar_url')
        except Exception:
            photo_url = None

        if not photo_url:
            return {}

        remote = self._download_remote_photo(photo_url, getattr(serializer, 'instance', None))
        if not remote:
            return {}
        return {'photo': remote}

    def _download_remote_photo(self, url, instance=None):
        try:
            resp = requests.get(str(url), timeout=15)
            resp.raise_for_status()
        except Exception:
            return None

        content_type = str(resp.headers.get('content-type', '')).lower()
        ext = '.jpg'
        if 'png' in content_type:
            ext = '.png'
        elif 'jpeg' in content_type or 'jpg' in content_type:
            ext = '.jpg'
        elif 'webp' in content_type:
            ext = '.webp'

        base = 'student'
        try:
            if instance and getattr(instance, 'id', None):
                base = f"student_{instance.id}"
            else:
                base = f"student_{uuid.uuid4().hex[:8]}"
        except Exception:
            base = f"student_{uuid.uuid4().hex[:8]}"

        filename = f"{base}{ext}"
        try:
            return ContentFile(resp.content, name=filename)
        except Exception:
            return None

    @action(detail=True, methods=['patch'], permission_classes=[IsTeacherOrAdmin], url_path='teacher-update')
    def teacher_update(self, request, pk=None):
        """Allow the class teacher to update limited student fields.
        Restrictions:
        - Cannot change admission_no, name, or klass (prevents moves/removals).
        - Requester must be the class teacher of this student.
        Admin/staff may still use normal update endpoints.
        """
        student = self.get_object()

        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        if not is_admin:
            # Must be teacher and the assigned class teacher of this student's class
            if getattr(user, 'role', None) != 'teacher':
                return Response({'detail': 'Not allowed'}, status=status.HTTP_403_FORBIDDEN)
            klass_id = getattr(getattr(student, 'klass', None), 'id', None)
            teacher_id = getattr(getattr(student, 'klass', None), 'teacher_id', None)
            if not (klass_id and teacher_id == getattr(user, 'id', None)):
                return Response({'detail': 'Only the class teacher can edit this student'}, status=status.HTTP_403_FORBIDDEN)

        # Build data allowing only specific fields
        incoming = dict(getattr(request, 'data', {}) or {})
        # Blocked keys regardless
        blocked = {'admission_no', 'name', 'klass'}
        # Allowed editable keys for teachers
        allowed = {
            'dob','gender','upi_number','guardian_id','guardian_name','guardian_passport_no','birth_certificate_no',
            'phone','email','address','photo','boarding_status','is_active'
        }
        # Remove any blocked or non-allowed keys when not admin
        if not is_admin:
            cleaned = {k: v for k, v in incoming.items() if k in allowed}
        else:
            # Admins can use this action too, but still respect blocked keys
            cleaned = {k: v for k, v in incoming.items() if k not in blocked}

        # If nothing to update
        if not cleaned:
            return Response({'detail': 'No editable fields provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Support remote photo url if provided
        if 'photo_url' in incoming and 'photo' not in cleaned:
            remote = self._download_remote_photo(incoming.get('photo_url'), instance=student)
            if remote:
                cleaned['photo'] = remote

        serializer = self.get_serializer(student, data=cleaned, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='history')
    def history(self, request, pk=None):
        """Return student's academic and finance history in one payload.
        Sections:
        - classes: progression inferred from exams and current class/graduation
        - exams: per-exam totals and approximate percentage
        - fees: invoices/payments grouped by year/term with totals and overall summary

        Access rules:
        - Students can only view their own history and only published exams.
        - Teachers/Admins can view students within their school.
        """
        # Resolve student object within scoped queryset
        student = self.get_object()

        user = getattr(request, 'user', None)
        is_admin = bool(user and (getattr(user, 'role', None) == 'admin' or user.is_staff or user.is_superuser))
        role = getattr(user, 'role', None)

        # Students can only view their own
        if role == 'student' and not is_admin:
            if getattr(student, 'user_id', None) != getattr(user, 'id', None):
                return Response({'detail': 'Forbidden'}, status=403)

        # School scoping for non-admin staff
        school = getattr(user, 'school', None)
        if school and not is_admin and role != 'student':
            # Ensure student belongs to same school (via current class or stored school)
            s_school_id = getattr(getattr(student, 'klass', None), 'school_id', None) or getattr(student, 'school_id', None)
            if s_school_id is not None and s_school_id != getattr(school, 'id', None):
                return Response({'detail': 'Student is not in your school scope'}, status=403)

        # ---------- Classes progression (inferred + recorded) ----------
        try:
            exams_q = (
                Exam.objects
                .filter(results__student=student)
                .select_related('klass', 'klass__stream')
                .values('year', 'term', 'grade_level_tag', 'klass__name')
                .distinct()
                .order_by('year', 'term', 'grade_level_tag')
            )
            classes_progression = []
            for row in exams_q:
                classes_progression.append({
                    'year': row.get('year'),
                    'term': row.get('term'),
                    'grade': row.get('grade_level_tag'),
                    'class_name': row.get('klass__name'),
                    'source': 'exam',
                })
        except Exception:
            classes_progression = []

        # Merge explicit history records
        try:
            hist_rows = (
                StudentClassHistory.objects
                .filter(student=student)
                .select_related('from_class','to_class','from_class__stream','to_class__stream')
                .order_by('created_at','id')
            )
            for h in hist_rows:
                classes_progression.append({
                    'year': getattr(h, 'year', None),
                    'term': getattr(h, 'term', None),
                    'grade': getattr(getattr(h.to_class, 'grade_level', None), 'strip', lambda: None)() if getattr(h, 'to_class', None) else None,
                    'class_name': getattr(getattr(h, 'to_class', None), 'name', None) or 'Graduated' if getattr(h, 'action', '') == 'graduated' else None,
                    'action': getattr(h, 'action', None),
                    'from': getattr(getattr(h, 'from_class', None), 'name', None),
                    'to': getattr(getattr(h, 'to_class', None), 'name', None) or ('Graduated' if getattr(h, 'action', '') == 'graduated' else None),
                    'source': 'record',
                    'created_at': getattr(h, 'created_at', None),
                })
        except Exception:
            pass

        # Append current class/graduation snapshot
        try:
            classes_progression.append({
                'year': getattr(getattr(student, 'graduation_year', None), 'year', None) if getattr(student, 'is_graduated', False) else None,
                'term': None,
                'grade': getattr(getattr(student, 'klass', None), 'grade_level', None) if getattr(student, 'klass', None) else ('Graduated' if getattr(student, 'is_graduated', False) else None),
                'class_name': getattr(getattr(student, 'klass', None), 'name', ('Graduated' if getattr(student, 'is_graduated', False) else None)),
            })
        except Exception:
            pass

        # ---------- Exam performance ----------
        try:
            qp = getattr(request, 'query_params', {})
            year_filter = qp.get('year')
            term_filter = qp.get('term')
            ay_label_filter = qp.get('academic_year') or qp.get('academic_year_label')
            include_subjects = str(qp.get('include_subjects', 'false')).lower() in ('1','true','yes','on')

            res_q = (
                ExamResult.objects
                .filter(student=student)
                .select_related('exam', 'exam__klass')
            )
            if role == 'student' and not is_admin:
                res_q = res_q.filter(exam__published=True)
            if year_filter not in (None, ''):
                try:
                    res_q = res_q.filter(exam__year=int(year_filter))
                except Exception:
                    pass
            if term_filter not in (None, ''):
                try:
                    res_q = res_q.filter(exam__term=int(term_filter))
                except Exception:
                    pass
            if ay_label_filter not in (None, ''):
                try:
                    ssid = getattr(getattr(student, 'klass', None), 'school_id', None) or getattr(student, 'school_id', None)
                    ay_obj = AcademicYear.objects.filter(school_id=ssid, label=ay_label_filter).first()
                    if ay_obj:
                        res_q = res_q.filter(exam__date__gte=getattr(ay_obj, 'start_date', None), exam__date__lte=getattr(ay_obj, 'end_date', None))
                except Exception:
                    pass
            exam_rows = (
                res_q
                .values('exam_id', 'exam__name', 'exam__year', 'exam__term', 'exam__total_marks', 'exam__klass__name', 'exam__date')
                .annotate(total=Sum('marks'), subjects=Sum(1))
                .order_by('exam__year', 'exam__term', 'exam_id')
            )
            exams = []
            for r in exam_rows:
                total = float(r.get('total') or 0)
                subj_count = int(r.get('subjects') or 0)
                out_of_each = float(r.get('exam__total_marks') or 100)
                approx_pct = None
                try:
                    denom = out_of_each * subj_count if subj_count else None
                    approx_pct = round((total / denom) * 100.0, 2) if denom else None
                except Exception:
                    approx_pct = None
                exams.append({
                    'exam': {
                        'id': r.get('exam_id'),
                        'name': r.get('exam__name'),
                        'year': r.get('exam__year'),
                        'term': r.get('exam__term'),
                        'klass_name': r.get('exam__klass__name'),
                        'total_marks': out_of_each,
                        'date': r.get('exam__date'),
                    },
                    'total_marks_obtained': round(total, 2),
                    'subjects_count': subj_count,
                    'approx_percentage': approx_pct,
                })
            if include_subjects and exams:
                try:
                    exam_ids = [e['exam']['id'] for e in exams if e.get('exam', {}).get('id') is not None]
                    details = (
                        ExamResult.objects
                        .filter(student=student, exam_id__in=exam_ids)
                        .select_related('subject', 'component')
                    )
                    by_exam = {}
                    for r in details:
                        rec = by_exam.setdefault(getattr(r, 'exam_id', None), [])
                        rec.append({
                            'subject': {
                                'id': getattr(getattr(r, 'subject', None), 'id', None),
                                'code': getattr(getattr(r, 'subject', None), 'code', None),
                                'name': getattr(getattr(r, 'subject', None), 'name', None),
                            },
                            'component': ({
                                'id': getattr(getattr(r, 'component', None), 'id', None),
                                'code': getattr(getattr(r, 'component', None), 'code', None),
                                'name': getattr(getattr(r, 'component', None), 'name', None),
                                'max_marks': getattr(getattr(r, 'component', None), 'max_marks', None),
                            } if getattr(r, 'component', None) else None),
                            'marks': getattr(r, 'marks', None),
                        })
                    for e in exams:
                        ex_id = e.get('exam', {}).get('id')
                        if ex_id in by_exam:
                            e['subjects'] = by_exam.get(ex_id) or []
                except Exception:
                    pass

            exams_grouped_map = {}
            for item in exams:
                y = item['exam'].get('year')
                t = item['exam'].get('term')
                key = (y, t)
                grp = exams_grouped_map.setdefault(key, {
                    'year': y,
                    'term': t,
                    'items': [],
                    'total_exams': 0,
                    'total_marks_obtained': 0.0,
                    '_total_denominator': 0.0,
                })
                grp['items'].append(item)
                grp['total_exams'] += 1
                try:
                    grp['total_marks_obtained'] += float(item.get('total_marks_obtained') or 0)
                    subj_ct = float(item.get('subjects_count') or 0)
                    per_subj_max = float(item.get('exam', {}).get('total_marks') or 0)
                    denom = (subj_ct * per_subj_max) if subj_ct and per_subj_max else 0.0
                    grp['_total_denominator'] += denom
                except Exception:
                    pass
            exams_grouped = []
            for g in sorted(exams_grouped_map.values(), key=lambda x: ((x['year'] or 0), (x['term'] or 0))):
                try:
                    g['approx_percentage_mean'] = round(((g['total_marks_obtained'] / g.get('_total_denominator', 0)) * 100.0), 2) if g.get('_total_denominator', 0) else None
                except Exception:
                    g['approx_percentage_mean'] = None
                g.pop('_total_denominator', None)
                exams_grouped.append(g)

            try:
                student_school_id = getattr(getattr(student, 'klass', None), 'school_id', None) or getattr(student, 'school_id', None)
                ay_list = []
                if student_school_id:
                    ay_list = list(AcademicYear.objects.filter(school_id=student_school_id).values('label', 'start_date', 'end_date'))
                def find_ay_label(d):
                    if not d:
                        return None
                    for ay in ay_list:
                        s = ay.get('start_date')
                        e = ay.get('end_date')
                        if s and e and s <= d <= e:
                            return ay.get('label')
                    return None
                for item in exams:
                    d = item.get('exam', {}).get('date')
                    item['exam']['academic_year_label'] = find_ay_label(d)
                grouped_by_ay = {}
                for item in exams:
                    label = item.get('exam', {}).get('academic_year_label') or 'Unknown'
                    term = item.get('exam', {}).get('term')
                    node = grouped_by_ay.setdefault(label, {'academic_year_label': label, 'terms': {}})
                    tnode = node['terms'].setdefault(term, {'term': term, 'items': [], 'total_exams': 0, 'total_marks_obtained': 0.0, '_total_denominator': 0.0})
                    tnode['items'].append(item)
                    tnode['total_exams'] += 1
                    try:
                        tnode['total_marks_obtained'] += float(item.get('total_marks_obtained') or 0)
                        subj_ct = float(item.get('subjects_count') or 0)
                        per_subj_max = float(item.get('exam', {}).get('total_marks') or 0)
                        denom = (subj_ct * per_subj_max) if subj_ct and per_subj_max else 0.0
                        tnode['_total_denominator'] += denom
                    except Exception:
                        pass
                exams_grouped_academic_year = []
                for lbl, node in grouped_by_ay.items():
                    terms_vals = list(node['terms'].values())
                    for tv in terms_vals:
                        try:
                            tv['approx_percentage_mean'] = round(((tv['total_marks_obtained'] / tv.get('_total_denominator', 0)) * 100.0), 2) if tv.get('_total_denominator', 0) else None
                        except Exception:
                            tv['approx_percentage_mean'] = None
                        tv.pop('_total_denominator', None)
                    terms_list = sorted(terms_vals, key=lambda x: (x['term'] or 0))
                    exams_grouped_academic_year.append({'academic_year_label': lbl, 'terms': terms_list})
                try:
                    def label_sort_key(lbl):
                        try:
                            nums = [int(x) for x in str(lbl).split('/') if x.isdigit()]
                            return (max(nums) if nums else 0)
                        except Exception:
                            return 0
                    exams_grouped_academic_year.sort(key=lambda x: label_sort_key(x['academic_year_label']))
                except Exception:
                    pass
            except Exception:
                exams_grouped_academic_year = []
        except Exception:
            exams = []
            exams_grouped = []
            exams_grouped_academic_year = []

        # ---------- Fees: allocations and payments ----------
        try:
            from finance.models import Invoice, Payment
            # Invoices grouped by year/term
            inv_q = Invoice.objects.filter(student=student)
            if school:
                inv_q = inv_q.select_related('student__klass').filter(student__klass__school=school)
            inv_grouped = inv_q.values('year', 'term').annotate(billed=Sum('amount')).order_by('year', 'term')

            pay_q = Payment.objects.filter(invoice__student=student)
            if school:
                pay_q = pay_q.select_related('invoice__student__klass').filter(invoice__student__klass__school=school)
            pay_grouped = pay_q.values('invoice__year', 'invoice__term').annotate(paid=Sum('amount')).order_by('invoice__year', 'invoice__term')

            # Merge grouped by (year, term)
            by_term = {}
            for r in inv_grouped:
                key = (r.get('year'), r.get('term'))
                by_term[key] = {'year': r.get('year'), 'term': r.get('term'), 'billed': float(r.get('billed') or 0), 'paid': 0.0}
            for r in pay_grouped:
                key = (r.get('invoice__year'), r.get('invoice__term'))
                entry = by_term.setdefault(key, {'year': r.get('invoice__year'), 'term': r.get('invoice__term'), 'billed': 0.0, 'paid': 0.0})
                entry['paid'] = float(r.get('paid') or 0)
            by_term_list = sorted(by_term.values(), key=lambda x: (x['year'] or 0, x['term'] or 0))

            # Detailed lists (limited fields)
            invoices = list(inv_q.order_by('-created_at').values('id', 'amount', 'status', 'category_id', 'year', 'term', 'due_date', 'created_at'))
            payments = list(pay_q.order_by('-created_at').values('id', 'invoice_id', 'amount', 'method', 'reference', 'created_at'))

            # Summary totals
            total_billed = float(inv_q.aggregate(s=Sum('amount'))['s'] or 0)
            total_paid = float(pay_q.aggregate(s=Sum('amount'))['s'] or 0)
            balance = round(total_billed - total_paid, 2)
            fees = {
                'by_term': by_term_list,
                'invoices': invoices,
                'payments': payments,
                'summary': {
                    'total_billed': round(total_billed, 2),
                    'total_paid': round(total_paid, 2),
                    'balance': balance,
                }
            }
        except Exception:
            fees = {'by_term': [], 'invoices': [], 'payments': [], 'summary': {'total_billed': 0.0, 'total_paid': 0.0, 'balance': 0.0}}

        return Response({
            'student': {
                'id': student.id,
                'name': getattr(student, 'name', None),
                'admission_no': getattr(student, 'admission_no', None),
                'klass': getattr(getattr(student, 'klass', None), 'name', None),
                'is_graduated': getattr(student, 'is_graduated', False),
                'graduation_year': getattr(student, 'graduation_year', None),
            },
            'classes': classes_progression,
            'exams': exams,
            'exams_grouped': exams_grouped,
            'exams_grouped_academic_year': exams_grouped_academic_year,
            'fees': fees,
        })

    def get_serializer_class(self):
        # Use lightweight serializer for list to reduce payload size
        from .serializers import StudentListSerializer, StudentSerializer as StudentDetailSerializer
        if getattr(self, 'action', None) == 'list':
            return StudentListSerializer
        return StudentDetailSerializer

    def perform_create(self, serializer):
        """Ensure school scoping on create: derive school from klass or request.user.school."""
        user = getattr(self.request, 'user', None)
        school = getattr(user, 'school', None)
        klass = serializer.validated_data.get('klass')
        # If klass provided, ensure it belongs to the same school
        if klass and school and klass.school_id != getattr(school, 'id', None):
            raise ValidationError({'klass': 'Class must belong to your school'})
        # Persist school for scoping (when klass is later cleared on graduation)
        payload_school = getattr(klass, 'school', None) or school
        photo_kwargs = self._build_remote_photo_kwargs(serializer)
        serializer.save(school=payload_school, **photo_kwargs)

    def perform_update(self, serializer):
        """Validate admin edit and maintain school scoping when class changes.
        - If klass is set, it must belong to admin's school; set student.school from klass.school.
        - If klass is cleared (None), keep existing student.school or fallback to admin's school.
        """
        user = getattr(self.request, 'user', None)
        school = getattr(user, 'school', None)
        klass = serializer.validated_data.get('klass', serializer.instance.klass)
        # If setting a new class, validate school and update school field
        if klass is not None:
            if school and getattr(klass, 'school_id', None) not in (None, getattr(school, 'id', None)):
                raise ValidationError({'klass': 'Class must belong to your school'})
            photo_kwargs = self._build_remote_photo_kwargs(serializer)
            serializer.save(school=getattr(klass, 'school', None), **photo_kwargs)
        else:
            # klass cleared (e.g., graduation or temporary unassignment)
            current_school = getattr(serializer.instance, 'school', None)
            photo_kwargs = self._build_remote_photo_kwargs(serializer)
            serializer.save(school=current_school or school, **photo_kwargs)

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='set-active')
    def set_active(self, request, pk=None):
        """Admin: toggle a student's active status. Body: { "is_active": true|false }"""
        stu = self.get_object()
        try:
            val = request.data.get('is_active')
            if isinstance(val, str):
                val = str(val).lower() in ('1','true','yes','on')
            else:
                val = bool(val)
        except Exception:
            return Response({'detail': 'is_active is required as boolean'}, status=400)
        stu.is_active = bool(val)
        # When deactivating, student should not belong to any class henceforth
        if stu.is_active is False:
            stu.klass = None
            # Keep school as-is for scoping
            stu.save(update_fields=['is_active', 'klass'])
        else:
            stu.save(update_fields=['is_active'])
        # Linked user.is_active will be synced by signal
        return Response({'detail': 'updated', 'id': stu.id, 'is_active': stu.is_active})

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin], url_path='bulk-update')
    def bulk_update(self, request):
        data = getattr(request, 'data', {}) or {}

        otp_err = self._verify_bulk_otp(request, data)
        if otp_err is not None:
            return otp_err

        ids = data.get('student_ids') or data.get('students') or data.get('ids')
        updates = data.get('updates') or {}

        if not isinstance(ids, (list, tuple)) or not ids:
            return Response({'detail': 'student_ids must be a non-empty list'}, status=400)
        if not isinstance(updates, dict) or not updates:
            return Response({'detail': 'updates must be a non-empty object'}, status=400)

        id_list = []
        for v in ids:
            try:
                id_list.append(int(v))
            except Exception:
                return Response({'detail': 'student_ids must contain only integers'}, status=400)

        allowed_fields = {'gender', 'klass', 'boarding_status'}
        cleaned = {k: v for k, v in updates.items() if k in allowed_fields}
        if not cleaned:
            return Response({'detail': 'No supported fields provided. Allowed: gender, klass, boarding_status'}, status=400)

        update_kwargs = {}

        if 'gender' in cleaned:
            gender = cleaned.get('gender')
            if isinstance(gender, str):
                gender = gender.strip()
            if gender not in ('Male', 'Female'):
                return Response({'detail': 'gender must be "Male" or "Female"'}, status=400)
            update_kwargs['gender'] = gender

        if 'boarding_status' in cleaned:
            b = cleaned.get('boarding_status')
            if isinstance(b, str):
                b = b.strip().lower()
            if b not in ('day', 'boarding'):
                return Response({'detail': 'boarding_status must be "day" or "boarding"'}, status=400)
            update_kwargs['boarding_status'] = b

        school = getattr(getattr(request, 'user', None), 'school', None)
        if 'klass' in cleaned:
            klass_id = cleaned.get('klass')
            if klass_id in (None, ''):
                return Response({'detail': 'klass is required'}, status=400)
            try:
                klass_id = int(klass_id)
            except Exception:
                return Response({'detail': 'klass must be an integer'}, status=400)
            klass = Class.objects.filter(pk=klass_id).only('id', 'school_id').first()
            if not klass:
                return Response({'detail': 'Class not found'}, status=404)
            if school and getattr(klass, 'school_id', None) != getattr(school, 'id', None):
                return Response({'detail': 'Class must belong to your school'}, status=403)
            update_kwargs['klass_id'] = klass_id
            update_kwargs['school_id'] = getattr(klass, 'school_id', None)

        qs = self.get_queryset().filter(id__in=id_list)
        found_ids = list(qs.values_list('id', flat=True))
        if not found_ids:
            return Response({'detail': 'No students found for the provided ids'}, status=404)

        try:
            with transaction.atomic():
                updated_count = qs.update(**update_kwargs)
        except Exception:
            return Response({'detail': 'Bulk update failed'}, status=400)

        return Response({'updated_count': int(updated_count or 0), 'student_ids': found_ids})

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin], url_path='bulk-otp/request')
    def bulk_otp_request(self, request):
        user = getattr(request, 'user', None)
        if not (user and getattr(user, 'is_authenticated', False)):
            return Response({'detail': 'Not authenticated'}, status=401)

        recipient = (getattr(user, 'email', None) or '').strip()
        if not recipient:
            return Response({'detail': 'Your account does not have an email address set.'}, status=400)

        code = f"{secrets.randbelow(1000000):06d}"
        ttl = 10 * 60
        key = self._bulk_otp_cache_key(user)
        attempts_key = self._bulk_otp_attempts_cache_key(user)
        cache.set(key, code, timeout=ttl)
        cache.set(attempts_key, 0, timeout=ttl)

        subject = 'Edu-Track Bulk Action Verification Code'
        message = (
            'Use this 6-digit verification code to confirm the bulk students action:\n\n'
            f'{code}\n\n'
            'This code expires in 10 minutes. If you did not request this, you can safely ignore this email.'
        )
        html_message = render_to_string(
            'verification_code_email.html',
            {
                'brand': 'EduTrack',
                'title': 'Verify your email address',
                'intro': 'To confirm the bulk students action, please enter the verification code below in the app.',
                'code': code,
                'footer': 'This code expires in 10 minutes. If you did not request this, you can safely ignore this email.',
            },
        )

        ok = False
        try:
            ok = send_email_safe_html(subject, message, html_message, recipient, school_id=getattr(user, 'school_id', None))
        except Exception:
            ok = False

        if not ok:
            cache.delete(key)
            cache.delete(attempts_key)
            return Response({'detail': 'Could not send verification email. Please try again.'}, status=400)

        return Response({'detail': 'Verification code sent'} )

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin], url_path='bulk-delete')
    def bulk_delete(self, request):
        data = getattr(request, 'data', {}) or {}

        otp_err = self._verify_bulk_otp(request, data)
        if otp_err is not None:
            return otp_err

        ids = data.get('student_ids') or data.get('students') or data.get('ids')
        if not isinstance(ids, (list, tuple)) or not ids:
            return Response({'detail': 'student_ids must be a non-empty list'}, status=400)

        id_list = []
        for v in ids:
            try:
                id_list.append(int(v))
            except Exception:
                return Response({'detail': 'student_ids must contain only integers'}, status=400)

        qs = self.get_queryset().filter(id__in=id_list)
        found_ids = list(qs.values_list('id', flat=True))
        if not found_ids:
            return Response({'detail': 'No students found for the provided ids'}, status=404)

        try:
            with transaction.atomic():
                qs.delete()
        except Exception:
            return Response({'detail': 'Bulk delete failed'}, status=400)

        return Response({'deleted_count': len(found_ids), 'student_ids': found_ids})

    def _bulk_otp_cache_key(self, user):
        school_id = getattr(getattr(user, 'school', None), 'id', None) or 0
        return f"bulk_students_otp:{getattr(user, 'id', 0)}:{school_id}"

    def _bulk_otp_attempts_cache_key(self, user):
        school_id = getattr(getattr(user, 'school', None), 'id', None) or 0
        return f"bulk_students_otp_attempts:{getattr(user, 'id', 0)}:{school_id}"

    def _verify_bulk_otp(self, request, data):
        user = getattr(request, 'user', None)
        if not (user and getattr(user, 'is_authenticated', False)):
            return Response({'detail': 'Not authenticated'}, status=401)

        provided = data.get('otp_code') or data.get('code')
        provided = str(provided or '').strip()
        if not provided:
            return Response({'detail': 'Verification code is required.'}, status=403)

        key = self._bulk_otp_cache_key(user)
        attempts_key = self._bulk_otp_attempts_cache_key(user)

        expected = cache.get(key)
        if not expected:
            return Response({'detail': 'Verification code expired. Request a new code.'}, status=403)

        expected = str(expected).strip()
        if provided != expected:
            try:
                attempts = int(cache.get(attempts_key) or 0)
            except Exception:
                attempts = 0
            attempts += 1
            cache.set(attempts_key, attempts, timeout=10 * 60)
            if attempts >= 5:
                cache.delete(key)
                cache.delete(attempts_key)
                return Response({'detail': 'Too many failed attempts. Request a new code.'}, status=403)
            return Response({'detail': 'Invalid verification code.'}, status=403)

        cache.delete(key)
        cache.delete(attempts_key)
        return None

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='my')
    def my(self, request):
        """Return the student record linked to the authenticated user, if any."""
        user = request.user
        qs = self.get_queryset().filter(user=user)
        student = qs.first()
        if not student:
            return Response({'detail': 'Student record not found for this user'}, status=404)
        ser = self.get_serializer(student)
        return Response(ser.data)

    @action(detail=False, methods=['patch', 'post'], permission_classes=[permissions.IsAuthenticated], url_path='my/update')
    def my_update(self, request):
        """Allow the authenticated student to update limited contact fields: email, phone, address.
        Accepts PATCH or POST with any subset of these fields.
        """
        user = request.user
        student = self.get_queryset().filter(user=user).first()
        if not student:
            return Response({'detail': 'Student record not found for this user'}, status=404)
        # Whitelist editable fields — use guardian phone (guardian_id) instead of student phone
        allowed_fields = ['email', 'address', 'guardian_id']
        payload = {k: v for k, v in request.data.items() if k in allowed_fields}
        # Backward compatibility: if 'phone' provided, treat it as guardian phone
        if 'phone' in request.data and 'guardian_id' not in payload:
            payload['guardian_id'] = request.data.get('phone')
        if not payload:
            return Response({'detail': 'No editable fields provided. Allowed: email, address, guardian_id (guardian phone).'}, status=400)
        # If guardian phone is being updated, clear student.phone to "remove" student phone number
        if 'guardian_id' in payload:
            payload['phone'] = ''
        serializer = self.get_serializer(student, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

class CompetencyViewSet(viewsets.ModelViewSet):
    queryset = Competency.objects.all()
    serializer_class = CompetencySerializer

class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAdmin]
    def perform_create(self, serializer):
        school = getattr(self.request.user, 'school', None)
        serializer.save(school=serializer.validated_data.get('school', school))
    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(Q(school=school) | Q(school__isnull=True))
        return qs

    @action(detail=True, methods=['get'], permission_classes=[IsAdmin], url_path='stats')
    def stats(self, request, pk=None):
        subject = self.get_object()
        school = getattr(request.user, 'school', None)
        classes_qs = Class.objects.all()
        if school:
            classes_qs = classes_qs.filter(school=school)
        classes_qs = classes_qs.filter(subjects=subject)

        # Latest exam per class
        latest_exams = []
        for c in classes_qs:
            e = Exam.objects.filter(klass=c).order_by('-date','-id').first()
            if e:
                latest_exams.append((c, e))

        # Compute avg per class for this subject
        by_grade = {}
        for c, e in latest_exams:
            res = ExamResult.objects.filter(exam=e, subject=subject)
            if res.exists():
                avg = res.aggregate(m=Avg('marks'))['m'] or 0
                g = c.grade_level
                agg = by_grade.setdefault(g, {'grade_level': g, 'sum': 0.0, 'count': 0})
                agg['sum'] += float(avg)
                agg['count'] += 1

        avg_by_grade = [
            { 'grade_level': g, 'average': round(v['sum']/v['count'], 2) if v['count'] else 0.0, 'classes': v['count'] }
            for g, v in by_grade.items()
        ]
        try:
            avg_by_grade.sort(key=lambda x: float(x['grade_level']))
        except Exception:
            avg_by_grade.sort(key=lambda x: str(x['grade_level']))

        # Teachers
        teachers = TeacherProfile.objects.all()
        if school:
            teachers = teachers.filter(Q(user__school=school) | Q(klass__school=school))
        teachers = teachers.filter(Q(subjects__icontains=subject.code) | Q(subjects__icontains=subject.name))
        tser = TeacherProfileSerializer(teachers, many=True)

        grading = [
            {'grade':'A','min':80,'max':100},
            {'grade':'B','min':70,'max':79},
            {'grade':'C','min':60,'max':69},
            {'grade':'D','min':50,'max':59},
            {'grade':'E','min':0,'max':49},
        ]

        return Response({
            'subject': {'id': subject.id, 'code': subject.code, 'name': subject.name},
            'avg_by_grade': avg_by_grade,
            'teachers': tser.data,
            'grading': grading,
        })

class SubjectComponentViewSet(viewsets.ModelViewSet):
    queryset = SubjectComponent.objects.all()
    serializer_class = SubjectComponentSerializer
    # Allow teachers to READ components, but only admins can CREATE/UPDATE/DELETE
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subject']

    def get_queryset(self):
        qs = super().get_queryset().select_related('subject')
        # Optional scope by user's school via the parent subject
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(subject__school=school)
        return qs

    def get_permissions(self):
        # Safe methods (list/retrieve) are available to teachers and admins
        if getattr(self, 'action', None) in ('list', 'retrieve') or self.request.method in permissions.SAFE_METHODS:
            return [IsTeacherOrAdmin()]
        # Mutations require admin rights
        return [IsAdmin()]

class AssessmentViewSet(viewsets.ModelViewSet):
    queryset = Assessment.objects.all()
    serializer_class = AssessmentSerializer
    permission_classes = [IsTeacherOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['student']

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='my')
    def my(self, request):
        """Return assessments for the authenticated student."""
        user = request.user
        # Find linked student
        student = Student.objects.filter(user=user).first()
        if not student:
            return Response({'detail': 'Student record not found for this user'}, status=404)
        qs = self.get_queryset().filter(student=student)
        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)
        ser = self.get_serializer(qs, many=True)
        return Response(ser.data)

class LessonPlanViewSet(viewsets.ModelViewSet):
    queryset = LessonPlan.objects.all()
    serializer_class = LessonPlanSerializer
    permission_classes = [IsTeacherOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['klass','subject','term','week','date']

    def get_queryset(self):
        qs = super().get_queryset().select_related('klass','subject','teacher')
        user = getattr(self.request, 'user', None)
        # Teachers only see their plans; admins see all within school
        if user and getattr(user, 'role', None) == 'teacher' and not (user.is_staff or user.is_superuser):
            qs = qs.filter(teacher=user)
        else:
            school = getattr(user, 'school', None)
            if school:
                qs = qs.filter(klass__school=school)
        return qs

    def perform_create(self, serializer):
        # Default teacher to the requester; validate class belongs to their school
        user = self.request.user
        klass = serializer.validated_data.get('klass')
        school = getattr(user, 'school', None)
        if school and klass and klass.school_id != school.id and not (user.is_staff or user.is_superuser):
            raise ValidationError({'klass': 'Class must belong to your school'})
        serializer.save(teacher=user)


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [IsAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        return qs

    def perform_create(self, serializer):
        school = getattr(self.request.user, 'school', None)
        if not school:
            raise ValidationError({'school': 'School is required. Set your user.school in Django admin.'})
        serializer.save(school=school)


class TimetableEntryViewSet(viewsets.ModelViewSet):
    queryset = TimetableEntry.objects.all()
    serializer_class = TimetableEntrySerializer
    permission_classes = [IsAdminOrTeacherReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['term','klass','subject','teacher','day_of_week','room']

    def get_queryset(self):
        qs = super().get_queryset().select_related('klass','subject','teacher','room','term','klass__stream')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(klass__school=school)
        return qs

    def perform_create(self, serializer):
        # Validate school scoping for klass/room consistency
        school = getattr(self.request.user, 'school', None)
        klass = serializer.validated_data.get('klass')
        room = serializer.validated_data.get('room')
        if school and klass and klass.school_id != school.id:
            raise ValidationError({'klass': 'Class must belong to your school'})
        if school and room and room.school_id != school.id:
            raise ValidationError({'room': 'Room must belong to your school'})
        serializer.save()

class SubjectGradingBandViewSet(viewsets.ModelViewSet):
    queryset = SubjectGradingBand.objects.all()
    serializer_class = SubjectGradingBandSerializer
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subject']

    def get_permissions(self):
        # Allow authenticated users to READ grading bands; restrict writes to admins
        if self.request and self.request.method in permissions.SAFE_METHODS:
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]

class StageGradingBandViewSet(viewsets.ModelViewSet):
    queryset = StageGradingBand.objects.all()
    serializer_class = StageGradingBandSerializer
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['stage']

    def get_permissions(self):
        # Allow authenticated users to READ stage grading bands; restrict writes to admins
        if self.request and self.request.method in permissions.SAFE_METHODS:
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        return qs

    def perform_create(self, serializer):
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if not school:
            raise ValidationError({'detail': 'No school context'})
        serializer.save(school=school)

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsTeacherOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['student']

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='my')
    def my(self, request):
        """Return attendance entries for the authenticated student."""
        user = request.user
        student = Student.objects.filter(user=user).first()
        if not student:
            return Response({'detail': 'Student record not found for this user'}, status=404)
        qs = self.get_queryset().filter(student=student).order_by('-date')
        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)
        ser = self.get_serializer(qs, many=True)
        return Response(ser.data)

class TeacherProfileViewSet(viewsets.ModelViewSet):
    queryset = TeacherProfile.objects.all()
    serializer_class = TeacherProfileSerializer
    permission_classes = [IsAdmin]

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='release')
    def release(self, request, pk=None):
        """Release a teacher from the school.
        Clears class/timetable assignments and disables the teacher's portal access."""
        profile = self.get_object()
        teacher_user = getattr(profile, 'user', None)
        if not teacher_user:
            return Response({'detail': 'Teacher profile is not linked to a user.'}, status=status.HTTP_400_BAD_REQUEST)

        school = getattr(request.user, 'school', None)
        if school and teacher_user.school_id and teacher_user.school_id != school.id:
            return Response({'detail': 'You can only release teachers from your school.'}, status=status.HTTP_403_FORBIDDEN)

        summary = {}
        with transaction.atomic():
            classes_qs = Class.objects.filter(teacher=teacher_user)
            if school:
                classes_qs = classes_qs.filter(school=school)
            summary['classes_unassigned'] = classes_qs.update(teacher=None)

            cst_qs = ClassSubjectTeacher.objects.filter(teacher=teacher_user)
            if school:
                cst_qs = cst_qs.filter(klass__school=school)
            summary['subject_assignments_removed'] = cst_qs.count()
            if summary['subject_assignments_removed']:
                cst_qs.delete()

            timetable_qs = TimetableEntry.objects.filter(teacher=teacher_user)
            if school:
                timetable_qs = timetable_qs.filter(klass__school=school)
            summary['timetable_entries_cleared'] = timetable_qs.update(teacher=None)

            availability_qs = TeacherAvailability.objects.filter(teacher=teacher_user)
            summary['availability_removed'] = availability_qs.delete()[0]

            profile_updates = []
            if profile.klass_id is not None:
                profile.klass = None
                profile_updates.append('klass')
            if profile.subjects:
                profile.subjects = ''
                profile_updates.append('subjects')
            if getattr(profile, 'can_manage_timetable', False):
                profile.can_manage_timetable = False
                profile_updates.append('can_manage_timetable')
            if profile_updates:
                profile.save(update_fields=profile_updates)

            if teacher_user.is_active:
                teacher_user.is_active = False
                teacher_user.save(update_fields=['is_active'])

        return Response({'detail': 'Teacher released successfully.', 'summary': summary})

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='mine')
    def mine(self, request):
        """Return the authenticated user's TeacherProfile.
        - Teachers: read-only access to their profile.
        - Admins: can also use this to quickly fetch their linked profile if any.
        """
        user = getattr(request, 'user', None)
        if not user or not user.is_authenticated:
            return Response({'detail': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        prof = TeacherProfile.objects.filter(user=user).first()
        if not prof:
            return Response({'detail': 'Teacher profile not found'}, status=status.HTTP_404_NOT_FOUND)
        ser = self.get_serializer(prof)
        return Response(ser.data)
    def get_queryset(self):
        qs = super().get_queryset().select_related('user', 'klass')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            # TeacherProfile does not have a direct school field; scope by either the user's school
            # or the assigned class's school. Additionally, include unscoped teachers (user.school is null)
            # so admins can discover and assign them to classes in their school.
            qs = qs.filter(Q(user__school=school) | Q(klass__school=school) | Q(user__school__isnull=True))
        # Optional filter: subject (id or code), matches teacher's subjects string by code or name
        subj_param = self.request.query_params.get('subject')
        if subj_param:
            try:
                # Try as ID
                subj = Subject.objects.filter(id=int(subj_param)).first()
            except (ValueError, TypeError):
                subj = Subject.objects.filter(Q(code__iexact=subj_param) | Q(name__iexact=subj_param)).first()
            if subj:
                qs = qs.filter(Q(subjects__icontains=subj.code) | Q(subjects__icontains=subj.name))
            else:
                # Fallback: plain contains search
                qs = qs.filter(subjects__icontains=subj_param)
        return qs


class AcademicYearViewSet(viewsets.ModelViewSet):
    queryset = AcademicYear.objects.all()
    serializer_class = AcademicYearSerializer
    permission_classes = [IsAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        return qs

    def perform_create(self, serializer):
        school = getattr(self.request.user, 'school', None)
        if not school:
            raise ValidationError({'school': 'No school associated with your account.'})
        serializer.save(school=school)

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='current')
    def current(self, request):
        school = getattr(request.user, 'school', None)
        if not school:
            return Response({'detail': 'No school associated with user'}, status=400)
        today = timezone.localdate()
        # Prefer calendar-based detection
        obj = AcademicYear.objects.filter(school=school, start_date__lte=today, end_date__gte=today).first()
        # Fallback to flag if date-based not found
        if not obj:
            obj = AcademicYear.objects.filter(school=school, is_current=True).first()
        if not obj:
            return Response({'detail': 'Current academic year not found for today and no fallback set'}, status=404)
        return Response(self.get_serializer(obj).data)

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='mine')
    def mine(self, request):
        """List all academic years for the authenticated user's school (most recent first)."""
        school = getattr(request.user, 'school', None)
        if not school:
            return Response({'detail': 'No school associated with user'}, status=400)
        qs = AcademicYear.objects.filter(school=school).order_by('-start_date')
        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)
        ser = self.get_serializer(qs, many=True)
        return Response(ser.data)

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='set-current')
    def set_current(self, request, pk=None):
        ay = self.get_object()
        ay.is_current = True
        ay.save()
        # Promotion is no longer triggered from here
        return Response({'detail': 'Current academic year updated', 'promoted': False})

    


class TermViewSet(viewsets.ModelViewSet):
    queryset = Term.objects.all()
    serializer_class = TermSerializer
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['academic_year', 'number']

    def get_queryset(self):
        qs = super().get_queryset().select_related('academic_year')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(academic_year__school=school)
        return qs

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='current')
    def current(self, request):
        school = getattr(request.user, 'school', None)
        if not school:
            return Response({'detail': 'No school associated with user'}, status=400)
        today = timezone.localdate()
        # Determine current AY by date, fallback to is_current
        ay = AcademicYear.objects.filter(school=school, start_date__lte=today, end_date__gte=today).first()
        if not ay:
            ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        if not ay:
            return Response({'detail': 'Current academic year not found for today and no fallback set'}, status=404)
        # Determine current term by date, fallback to is_current
        term = Term.objects.filter(academic_year=ay, start_date__lte=today, end_date__gte=today).first()
        if not term:
            term = Term.objects.filter(academic_year=ay, is_current=True).first()
        if not term:
            return Response({'detail': 'Current term not found for today and no fallback set'}, status=404)
        return Response(self.get_serializer(term).data)

    @action(detail=False, methods=['get'], permission_classes=[IsTeacherOrAdmin], url_path='of-current-year')
    def of_current_year(self, request):
        """List all terms for the current academic year for the user's school."""
        school = getattr(request.user, 'school', None)
        if not school:
            return Response({'detail': 'No school associated with user'}, status=400)
        today = timezone.localdate()
        ay = AcademicYear.objects.filter(school=school, start_date__lte=today, end_date__gte=today).first()
        if not ay:
            ay = AcademicYear.objects.filter(school=school, is_current=True).first()
        if not ay:
            return Response({'detail': 'Current academic year not found for today and no fallback set'}, status=404)
        qs = Term.objects.filter(academic_year=ay).order_by('number')
        page = self.paginate_queryset(qs)
        if page is not None:
            ser = self.get_serializer(page, many=True)
            return self.get_paginated_response(ser.data)
        ser = self.get_serializer(qs, many=True)
        return Response(ser.data)

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='set-current')
    def set_current(self, request, pk=None):
        term = self.get_object()
        term.is_current = True
        term.save()
        return Response({'detail': 'Current term updated'})

# ===== Bulk import endpoints =====
@api_view(["POST"])
@permission_classes([IsAdmin])
@parser_classes([MultiPartParser])
def import_students(request):
    """CSV columns: admission_no,name,dob(YYYY-MM-DD),gender,guardian_id,class_id(optional)
    Uses request.user.school for scoping and allows linking to class by ID.
    """
    file = request.FILES.get('file')
    if not file:
        return Response({'detail': 'file is required'}, status=400)
    text = io.StringIO(file.read().decode('utf-8'))
    reader = csv.DictReader(text)
    created = 0
    errors = []
    for i, row in enumerate(reader, start=1):
        try:
            klass_id = row.get('class_id') or None
            klass = Class.objects.filter(id=klass_id).first() if klass_id else None
            Student.objects.create(
                admission_no=row['admission_no'],
                name=row['name'],
                dob=row['dob'],
                gender=row.get('gender',''),
                guardian_id=row.get('guardian_id',''),
                klass=klass,
            )
            created += 1
        except Exception as e:
            errors.append({'row': i, 'error': str(e)})
    return Response({'created': created, 'errors': errors}, status=201)


@api_view(["POST"])
@permission_classes([IsAdmin])
@parser_classes([MultiPartParser])
def import_competencies(request):
    """CSV columns: code,title,description,levels (comma-separated)"""
    file = request.FILES.get('file')
    if not file:
        return Response({'detail': 'file is required'}, status=400)
    text = io.StringIO(file.read().decode('utf-8'))
    reader = csv.DictReader(text)
    created = 0
    updated = 0
    for row in reader:
        levels = [s.strip() for s in (row.get('levels') or '').split(',') if s.strip()]
        obj, is_created = Competency.objects.update_or_create(
            code=row['code'],
            defaults={
                'title': row.get('title',''),
                'description': row.get('description',''),
                'level_scale': levels or ["Emerging","Developing","Proficient","Mastered"],
            }
        )
        created += 1 if is_created else 0
        updated += 0 if is_created else 1
    return Response({'created': created, 'updated': updated}, status=201)


# ===== Timetable Planning ViewSets =====
class TimetableTemplateViewSet(viewsets.ModelViewSet):
    queryset = TimetableTemplate.objects.all()
    serializer_class = TimetableTemplateSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        return qs

    def perform_create(self, serializer):
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if not school:
            raise ValidationError({'school': 'School is required'})
        serializer.save(school=school)


class PeriodSlotTemplateViewSet(viewsets.ModelViewSet):
    queryset = PeriodSlotTemplate.objects.all()
    serializer_class = PeriodSlotTemplateSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]

    def get_queryset(self):
        qs = super().get_queryset().select_related('template')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(template__school=school)
        template_id = self.request.query_params.get('template')
        if template_id:
            qs = qs.filter(template_id=template_id)
        return qs


class TimetablePlanViewSet(viewsets.ModelViewSet):
    queryset = TimetablePlan.objects.all()
    serializer_class = TimetablePlanSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['term','status']

    def get_queryset(self):
        qs = super().get_queryset().select_related('term','template')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(school=school)
        return qs

    def perform_create(self, serializer):
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if not school:
            raise ValidationError({'school': 'School is required'})
        serializer.save(school=school, created_by=getattr(self.request, 'user', None))

    @action(detail=True, methods=['post'], url_path='generate', permission_classes=[IsAdmin])
    def generate(self, request, pk=None):
        plan = self.get_object()
        try:
            from .services.timetable_generator import generate as generate_timetable
        except Exception as e:
            return Response({
                'version_id': None,
                'placed_count': 0,
                'unplaced': [],
                'detail': f'Generator unavailable: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            result = generate_timetable(plan)
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'version_id': None,
                'placed_count': 0,
                'unplaced': [],
                'detail': f'Generation failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class TimetableClassConfigViewSet(viewsets.ModelViewSet):
    queryset = TimetableClassConfig.objects.all()
    serializer_class = TimetableClassConfigSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['plan','klass']

    def get_queryset(self):
        qs = super().get_queryset().select_related('plan','klass','room_preference')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(plan__school=school)
        return qs


class ClassSubjectQuotaViewSet(viewsets.ModelViewSet):
    queryset = ClassSubjectQuota.objects.all()
    serializer_class = ClassSubjectQuotaSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['plan','klass','subject']

    def get_queryset(self):
        qs = super().get_queryset().select_related('plan','klass','subject')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(plan__school=school)
        return qs


class TeacherAvailabilityViewSet(viewsets.ModelViewSet):
    queryset = TeacherAvailability.objects.all()
    serializer_class = TeacherAvailabilitySerializer
    permission_classes = [IsAdminOrTeacherReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['teacher','day_of_week']

    def get_queryset(self):
        qs = super().get_queryset().select_related('teacher')
        return qs


class TimetableVersionViewSet(viewsets.ModelViewSet):
    queryset = TimetableVersion.objects.all()
    serializer_class = TimetableVersionSerializer
    permission_classes = [IsAdminOrTeacherReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['plan','is_current']

    def get_queryset(self):
        qs = super().get_queryset().select_related('plan')
        school = getattr(getattr(self.request, 'user', None), 'school', None)
        if school:
            qs = qs.filter(plan__school=school)
        return qs

    @action(detail=True, methods=['post'], url_path='publish')
    def publish(self, request, pk=None):
        version = self.get_object()
        version.is_current = True
        version.save(update_fields=['is_current'])
        # mark plan status updated
        TimetablePlan.objects.filter(pk=version.plan_id).update(status='published')
        # Notify entire school via broadcast message (also triggers email/SMS delivery)
        try:
            from communications.utils import resolve_default_sender_id, create_broadcast_message
            school_id = getattr(getattr(version, 'plan', None), 'school_id', None)
            if school_id:
                sender_id = resolve_default_sender_id(school_id)
                term = getattr(getattr(version, 'plan', None), 'term', None)
                term_text = f"Term {getattr(term, 'number', '')}" if term else ""
                body = f"A new timetable has been published {('for ' + term_text) if term_text else ''}. Please check the Timetable section."
                if sender_id:
                    create_broadcast_message(school_id=school_id, sender_id=sender_id, body=body)
        except Exception:
            pass
        # Additionally: send personalized teacher messages with direct timetable link
        try:
            from django.contrib.auth import get_user_model
            from communications.utils import resolve_default_sender_id, create_personalized_messages_for_users
            plan = getattr(version, 'plan', None)
            school_id = getattr(plan, 'school_id', None)
            if school_id:
                User = get_user_model()
                teacher_users = User.objects.filter(school_id=school_id, role='teacher', is_active=True).only('id','first_name','last_name','username')
                if teacher_users.exists():
                    sender_id = resolve_default_sender_id(school_id)
                    if sender_id:
                        # Build absolute link to teacher timetable with planId
                        try:
                            base_url = request.build_absolute_uri('/')
                        except Exception:
                            base_url = '/'
                        base_url = base_url.rstrip('/')
                        link = f"{base_url}/teacher/timetable?planId={getattr(plan, 'id', '')}"
                        term = getattr(plan, 'term', None)
                        term_label = getattr(term, 'name', None) or (f"T{getattr(term,'number','')}" if getattr(term,'number',None) else '')
                        pairs = []
                        for u in teacher_users:
                            try:
                                name = (f"{getattr(u,'first_name','')} {getattr(u,'last_name','')}").strip() or getattr(u,'username','Teacher')
                                msg = (
                                    f"Hello {name}, a new timetable {('for ' + term_label) if term_label else ''} has been published. "
                                    f"View your personalized timetable here: {link}"
                                )
                                pairs.append((u.id, msg))
                            except Exception:
                                continue
                        if pairs:
                            create_personalized_messages_for_users(
                                school_id=school_id,
                                sender_id=sender_id,
                                user_body_pairs=pairs,
                                system_tag='timetable_publish',
                                queue_delivery=True,
                            )
        except Exception:
            pass
        return Response({'detail': 'published'})
